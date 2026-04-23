import csv, io, sys, json, os
from collections import defaultdict
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

# ===== ファイルパス（毎週更新） =====
TRAFFIC_FILE = r"C:\Users\KEISUKE SHIMOMOTO\Downloads\eBay-ListingsTrafficReport-Mar-15-2026-21_56_24-0700-13288691549.csv"
ADS_FILE     = r"C:\Users\KEISUKE SHIMOMOTO\Downloads\rioxxrinaxjapan_advertising_sales_report_20260315.csv"
TRANS_FILE   = r"C:\Users\KEISUKE SHIMOMOTO\Downloads\Transaction-Mar-16-2026-01_18_10-0700-13288723972.csv"

# 出力ファイル名は日付を自動挿入（手動更新不要）
TODAY        = date.today()
OUTPUT       = rf"C:\Users\KEISUKE SHIMOMOTO\Desktop\reffort\commerce\ebay\analytics\eBay週次レポート_v3_{TODAY.strftime('%Y%m%d')}.xlsx"

# 週次履歴ファイル（前週比較のためのJSONデータ）
HISTORY_FILE = r"C:\Users\KEISUKE SHIMOMOTO\Desktop\reffort\commerce\ebay\analytics\weekly_history.json"

# ===== 期間設定 =====
PERIOD_LABEL = "2/13〜3/15（31日間）"
WEEKS = [
    ('W1\n2/13-19', date(2026,2,13), date(2026,2,19), 7),
    ('W2\n2/20-26', date(2026,2,20), date(2026,2,26), 7),
    ('W3\n2/27-3/5',date(2026,2,27), date(2026,3,5),  7),
    ('W4\n3/6-15',  date(2026,3,6),  date(2026,3,15), 10),
]
# PLP費用（Transaction Reportに含まれないため手入力）
PLP_FEE_TOTAL = 491.0

# 育成候補の閾値（出品数が減るにつれて下げる。AIが自動提案する）
IKUSEI_PV_THRESHOLD = 50   # PV閾値（現在50）
IKUSEI_MAX          = 20   # 最大表示件数

# ===== 週次履歴の読み書き =====
def load_history():
    """前週データを読み込む（ファイルがなければ空を返す）"""
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_history(history):
    """週次データをJSONに保存（最大12週分保持）"""
    if len(history) > 12:
        oldest_key = sorted(history.keys())[0]
        del history[oldest_key]
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

def get_last_week_data(history, current_key):
    """現在より古い最新のデータを返す"""
    past_keys = sorted([k for k in history.keys() if k < current_key])
    if not past_keys:
        return None
    return history[past_keys[-1]]

# 履歴を読み込み
history = load_history()
HISTORY_DATE_KEY = TODAY.strftime('%Y-%m-%d')
last_data = get_last_week_data(history, HISTORY_DATE_KEY)
LAST_PERIOD = last_data.get('period', '—') if last_data else '—'

print(f"📅 前週データ: {'あり（' + LAST_PERIOD + '）' if last_data else 'なし（初回実行）'}")

# ===== ユーティリティ =====
def to_f(s):
    try: return float(str(s).replace(',','').replace('%','').replace('$','').strip())
    except: return 0.0

def clean_id(s):
    return s.strip().replace('=','').replace('"','')

def parse_tx_date(s):
    for fmt in ('%b %d, %Y', '%b  %d, %Y'):
        try: return datetime.strptime(s.strip(), fmt).date()
        except: pass
    return None

def days_listed(s):
    try:
        d = datetime.strptime(s.strip(), '%Y-%m-%d').date()
        return (TODAY - d).days
    except: return 0

def pct_change(new, old):
    if old == 0: return '—'
    c = (new - old) / old * 100
    return f'▲{c:.1f}%' if c >= 0 else f'▼{abs(c):.1f}%'

# ===== 前週データ参照ヘルパー =====
def get_prev_category(item_id):
    """前週のカテゴリを返す"""
    if not last_data:
        return '—'
    cat_map = [
        ('🔥 コア売れ筋', 'top15_ids'),
        ('⭐ 準売れ筋',   '準売れ筋_ids'),
        ('🌱 育成候補',   '育成_ids'),
        ('⚠️ 要調査',     '要調査_ids'),
        ('🗑 削除L1',     '削除L1_ids'),
        ('🗑 削除L2',     '削除L2_ids'),
    ]
    for cat, key in cat_map:
        if item_id in last_data.get(key, []):
            return cat
    if item_id in last_data.get('all_ids', []):
        return '（圏外）'
    return '🆕 新規'

def get_prev_sold(item_id):
    """前週の販売数を返す（なければNone）"""
    if not last_data:
        return None
    return last_data.get('per_item', {}).get(item_id, {}).get('sold', None)

def get_prev_cvr(item_id):
    """前週のCVRを返す（なければNone）"""
    if not last_data:
        return None
    return last_data.get('per_item', {}).get(item_id, {}).get('cvr', None)

def fmt_delta_sold(current_sold, item_id):
    """前週比販売数の変化を文字列で返す（▲+3 / ▼-2 / ±0 / —）"""
    prev = get_prev_sold(item_id)
    if prev is None:
        return '—'
    delta = int(current_sold) - int(prev)
    if delta > 0:  return f'▲+{delta}'
    elif delta < 0: return f'▼{delta}'
    return '±0'

# ===== 標準列幅（全シート共通・統一感のために定数化）=====
W_RANK     = 6    # 順位・優先度
W_TITLE    = 52   # 商品タイトル
W_ITEM_ID  = 14   # Item ID
W_SCORE    = 11   # スコア・ポテンシャル
W_SOLD     = 8    # 販売数
W_CVR      = 8    # CVR
W_PV       = 9    # PV
W_IMPS     = 12   # インプレッション
W_QTY      = 8    # 在庫数
W_DAYS     = 8    # 掲載日数
W_PROMO    = 13   # 広告状態
W_ORGANIC  = 10   # Organic比
W_PREV_CAT = 14   # 前週カテゴリ
W_AD_SALES = 12   # 広告売上
W_MEMO     = 20   # メモ・確認欄
W_REASON   = 28   # 原因・推測

# ===== スタイル定数 =====
FONT = 'メイリオ'
C_GREEN_BG  = 'E8F5E9'
C_BLUE_BG   = 'E3F2FD'
C_YELLOW_BG = 'FFFDE7'
C_ORANGE_BG = 'FFF3E0'
C_RED_BG    = 'FFEBEE'
C_ALERT_BG  = 'FCE4EC'   # コア落ちアラート
C_GRAY_BG   = 'F5F5F5'
C_DARK      = '1A237E'
C_GREEN_HDR = '1B5E20'
C_BLUE_HDR  = '0D47A1'
C_YELLOW_HDR= 'E65100'
C_ORANGE_HDR= 'BF360C'
C_RED_HDR   = 'B71C1C'
C_ALERT_HDR = 'AD1457'   # コア落ち

def hdr_font(color='FFFFFF', size=10, bold=True):
    return Font(name=FONT, size=size, bold=bold, color=color)

def body_font(size=9, bold=False, color='000000'):
    return Font(name=FONT, size=size, bold=bold, color=color)

def hdr_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def body_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def border_thin():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def apply_header_row(ws, row_num, headers, bg_color, txt_color='FFFFFF', row_height=22):
    ws.row_dimensions[row_num].height = row_height
    for col_idx, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=h)
        cell.font = hdr_font(color=txt_color)
        cell.fill = hdr_fill(bg_color)
        cell.alignment = center()
        cell.border = border_thin()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

def apply_body_row(ws, row_num, values, bg_color=None, bold=False, height=20):
    ws.row_dimensions[row_num].height = height
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=v)
        cell.font = body_font(bold=bold)
        if bg_color:
            cell.fill = body_fill(bg_color)
        cell.border = border_thin()
        if isinstance(v, (int, float)):
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = left()

# ===== Traffic Report 読み込み =====
with open(TRAFFIC_FILE, encoding='utf-8-sig') as f:
    lines = f.readlines()
data = io.StringIO(''.join(lines[5:]))

items = []
for row in csv.DictReader(data):
    title = row.get('Listing title','').strip()
    if not title: continue
    start_str = row.get('Item Start Date','').strip()
    total_imps = to_f(row.get('Total impressions', 0))
    pl_imps    = to_f(row.get('Total Promoted Listings impressions (applies to eBay site only)', 0))
    off_imps   = to_f(row.get('Total Promoted Offsite impressions (applies to off-eBay only)', 0))
    org_imps   = to_f(row.get('Total organic impressions on eBay site', 0))
    items.append({
        'title':    title,
        'id':       clean_id(row.get('eBay item ID','')),
        'start':    start_str,
        'days':     days_listed(start_str),
        'imps':     total_imps,
        'pl_imps':  pl_imps,
        'off_imps': off_imps,
        'org_imps': org_imps,
        'pv':       to_f(row.get('Total page views', 0)),
        'sold':     to_f(row.get('Quantity sold', 0)),
        'cvr':      to_f(row.get('Sales conversion rate = Quantity sold/Total page views', 0)),
        'ctr':      to_f(row.get('Click-through rate = Page views from eBay site/Total impressions', 0)),
        'qty':      to_f(row.get('Quantity available', 0)),
        'promo':    row.get('Current promoted listings status','').strip(),
    })

# ===== Advertising Report 読み込み =====
with open(ADS_FILE, encoding='utf-8-sig') as f:
    ads_lines = f.readlines()
ads_data = io.StringIO(''.join(ads_lines[2:]))
ad_by_item = defaultdict(lambda: {'plg_s':0,'plg_fee':0,'plp_s':0,'off_s':0})

for row in csv.DictReader(ads_data):
    iid  = row.get('Item ID','').strip()
    ad   = row.get('Ad type','')
    st   = row.get('Campaign strategy','')
    sale_type = row.get('Sale type','')
    s    = to_f(row.get('Sales','0'))
    fee  = to_f(row.get('Ad fees (General)','-'))
    if 'Offsite' in ad:
        ad_by_item[iid]['off_s'] += s
    elif 'Priority' in st:
        ad_by_item[iid]['plp_s'] += s
    elif 'General' in st and 'Halo' not in sale_type:
        ad_by_item[iid]['plg_s']  += s
        ad_by_item[iid]['plg_fee']+= fee

for item in items:
    ad = ad_by_item.get(item['id'], {})
    item['plg_s']    = ad.get('plg_s', 0)
    item['plg_fee']  = ad.get('plg_fee', 0)
    item['plp_s']    = ad.get('plp_s', 0)
    item['off_s']    = ad.get('off_s', 0)
    item['ad_sales'] = item['plg_s'] + item['plp_s'] + item['off_s']

# ===== Transaction Report 読み込み（週別集計）=====
with open(TRANS_FILE, encoding='utf-8-sig') as f:
    trans_lines = f.readlines()
trans_data = io.StringIO(''.join(trans_lines[11:]))

week_stats = []
for wlabel, wstart, wend, wdays in WEEKS:
    week_stats.append({
        'label': wlabel, 'start': wstart, 'end': wend, 'days': wdays,
        'orders':0, 'gross':0.0, 'fvf':0.0, 'intl':0.0,
        'plg_fee':0.0, 'off_fee':0.0, 'qty':0,
    })

total_orders = 0
total_gross = total_fvf = total_intl = total_plg_fee_tx = total_off_fee = 0.0
total_qty = 0

for row in csv.DictReader(trans_data):
    cur = row.get('Transaction currency','').strip()
    if cur != 'USD': continue
    t     = row.get('Type','').strip()
    d_str = row.get('Transaction creation date','').strip()
    d     = parse_tx_date(d_str)
    if not d: continue
    desc  = row.get('Description','').strip()

    if t == 'Order':
        gross = to_f(row.get('Gross transaction amount', 0))
        fvf   = to_f(row.get('Final Value Fee - fixed', 0)) + to_f(row.get('Final Value Fee - variable', 0))
        intl  = to_f(row.get('International fee', 0))
        qty   = int(to_f(row.get('Quantity', 0)))
        total_orders += 1; total_gross += gross
        total_fvf += fvf; total_intl += intl; total_qty += qty
        for ws in week_stats:
            if ws['start'] <= d <= ws['end']:
                ws['orders'] += 1; ws['gross'] += gross
                ws['fvf'] += fvf; ws['intl'] += intl; ws['qty'] += qty
                break
    elif t == 'Other fee' and 'Promoted Listings - General fee' in desc:
        fee = abs(to_f(row.get('Net amount', 0)))
        total_plg_fee_tx += fee
        for ws in week_stats:
            if ws['start'] <= d <= ws['end']:
                ws['plg_fee'] += fee; break
    elif t == 'Other fee' and 'Promoted Offsite fee' in desc:
        fee = abs(to_f(row.get('Net amount', 0)))
        total_off_fee += fee
        for ws in week_stats:
            if ws['start'] <= d <= ws['end']:
                ws['off_fee'] += fee; break

for ws in week_stats:
    ws['net'] = ws['gross'] - ws['fvf'] - ws['intl'] - ws['plg_fee'] - ws['off_fee']

# ===== Traffic集計 =====
total_items   = len(items)
total_imps_t  = sum(i['imps'] for i in items)
total_pv_t    = sum(i['pv'] for i in items)
total_sold_t  = sum(i['sold'] for i in items)
total_pl_imps = sum(i['pl_imps'] for i in items)
total_off_imps= sum(i['off_imps'] for i in items)
total_org_imps= sum(i['org_imps'] for i in items)
zero_sold_cnt = sum(1 for i in items if i['sold'] == 0)
overall_ctr   = total_pv_t / total_imps_t * 100 if total_imps_t > 0 else 0
overall_cvr   = total_sold_t / total_pv_t * 100 if total_pv_t > 0 else 0

# ===== 商品分類 =====
# 🔥 コア売れ筋TOP15
top15     = sorted(items, key=lambda x: -x['sold'])[:15]
top15_ids = {i['id'] for i in top15}

# ⭐ 準売れ筋: sold>0 かつ TOP15外
# ポテンシャルスコア = CVR × PV（高いほど「今すぐ伸ばせる」商品）
# 優先度: S=上位10件（最重点）/ A=11-30件（重点）/ B=31-80件（普通）/ C=それ以下
_準売れ筋_raw = [i for i in items if i['sold']>0 and i['id'] not in top15_ids]
for i in _準売れ筋_raw:
    i['pot_score'] = round(i['cvr'] * i['pv'], 1)   # ポテンシャルスコア計算
準売れ筋 = sorted(_準売れ筋_raw, key=lambda x: -x['pot_score'])  # スコア降順で並べ替え
for idx, i in enumerate(準売れ筋):                  # 優先度ラベルを付与
    if   idx < 10:  i['priority'] = 'S 🔥'
    elif idx < 30:  i['priority'] = 'A ⭐'
    elif idx < 80:  i['priority'] = 'B 🌱'
    else:           i['priority'] = 'C —'
準売れ筋_ids = {i['id'] for i in 準売れ筋}

# ⚠️ 要調査: sold=0 & インプ500以上 & 掲載90日以上（全件記録し上位50件表示）
要調査_all  = [i for i in items if i['sold']==0 and i['imps']>=500 and i['days']>=90]
要調査      = sorted(要調査_all, key=lambda x: -x['imps'])[:50]
要調査_ids  = {i['id'] for i in 要調査_all}   # 全件のIDセット（育成除外に使用）

# 🌱 育成候補（v3改訂）
#   条件: sold=0 & PV>=閾値 & 要調査に含まれない（ツールバグ疑いを除外）& 上位20件
#   【v3変更点】PV20→50に引き上げ、要調査との重複を除去、スタッフが対処可能な件数に絞る
育成        = sorted(
    [i for i in items if i['sold']==0 and i['pv']>=IKUSEI_PV_THRESHOLD
     and i['id'] not in 要調査_ids],
    key=lambda x: -x['pv']
)[:IKUSEI_MAX]
育成_ids    = {i['id'] for i in 育成}

# 🗑 削除候補L1: 完全死蔵（即削除）
削除L1      = [i for i in items if i['imps']==0 and i['pv']==0 and i['sold']==0]
削除L1_ids  = {i['id'] for i in 削除L1}

# 🗑 削除候補L2: 高齢・ほぼ非表示・売上ゼロ（要確認削除）
削除L2      = sorted(
    [i for i in items if i['id'] not in 削除L1_ids
     and i['imps']<50 and i['pv']<5 and i['sold']==0 and i['days']>=180],
    key=lambda x: -x['days']
)
削除L2_ids  = {i['id'] for i in 削除L2}

# ===== コア落ちアラート計算 =====
# 前週TOP15にいたが今週TOP15から外れた商品を特定する
items_by_id = {i['id']: i for i in items}

def get_current_category(item_id):
    """現在のカテゴリを返す"""
    if item_id in top15_ids:    return '🔥 コア売れ筋'
    if item_id in 準売れ筋_ids: return '⭐ 準売れ筋'
    if item_id in 育成_ids:     return '🌱 育成候補'
    if item_id in 要調査_ids:   return '⚠️ 要調査'
    if item_id in 削除L1_ids:   return '🗑 削除L1'
    if item_id in 削除L2_ids:   return '🗑 削除L2'
    return '（その他）'

def auto_reason(item):
    """コア落ちの原因を自動推測する"""
    if item is None:            return '🗑 出品リストから消えた（削除済？）'
    if item['qty'] == 0:        return '⚠️ 在庫切れ（補充or出品終了を検討）'
    if item['imps'] == 0:       return '🔧 在庫ツールバグ？（Cowatech確認）'
    if item['sold'] == 0 and item['pv'] > 20:
                                return '💰 価格・競合問題？（比較調査を）'
    if item['sold'] == 0 and item['pv'] == 0:
                                return '🔧 露出ゼロ（ツールor出品問題）'
    return '📉 競合激化 or 需要低下（価格・販促見直し）'

if last_data:
    prev_top15_ids = set(last_data.get('top15_ids', []))
    コア落ち_ids   = prev_top15_ids - top15_ids
    コア落ち = []
    for iid in コア落ち_ids:
        item      = items_by_id.get(iid)
        prev_info = last_data.get('per_item', {}).get(iid, {})
        prev_sold = prev_info.get('sold', 0)
        prev_cvr  = prev_info.get('cvr', 0.0)
        prev_title= prev_info.get('title', iid)
        curr_sold = item['sold'] if item else 0
        curr_cvr  = item['cvr']  if item else 0.0
        コア落ち.append({
            'title':      item['title'] if item else prev_title,
            'id':         iid,
            'prev_sold':  prev_sold,
            'curr_sold':  curr_sold,
            'delta_sold': int(curr_sold) - int(prev_sold),
            'prev_cvr':   prev_cvr,
            'curr_cvr':   curr_cvr,
            'qty':        int(item['qty']) if item else 0,
            'curr_cat':   get_current_category(iid),
            'reason':     auto_reason(item),
        })
    コア落ち.sort(key=lambda x: -x['prev_sold'])
else:
    コア落ち = []

# ===== AIからの動的提案を生成 =====
suggestions = []

# 出品数の変化チェック
if last_data:
    last_total = last_data.get('total_items', 0)
    if last_total > 0:
        reduction = last_total - total_items
        if reduction >= 50:
            new_thresh = max(20, IKUSEI_PV_THRESHOLD - 10)
            suggestions.append(
                f'📉 出品数が前回比▼{reduction}件減少（{last_total}→{total_items}件）。'
                f'大掃除が順調に進んでいます。育成候補のPV閾値を{IKUSEI_PV_THRESHOLD}→{new_thresh}に'
                f'下げることを推奨します（次回のスクリプト設定変更で反映）。'
            )

# コア落ちチェック
if コア落ち:
    suggestions.append(
        f'🚨 先週のコア売れ筋から{len(コア落ち)}件が脱落しました。'
        f'「🚨 コア落ち」シートを確認してください。'
    )

# 育成候補の件数チェック
if len(育成) < 5:
    suggestions.append(
        f'💡 育成候補が{len(育成)}件と少なくなっています。'
        f'出品数削減が進んでいる証拠です。PV閾値を{IKUSEI_PV_THRESHOLD}→{max(20, IKUSEI_PV_THRESHOLD-10)}に'
        f'下げることを推奨します（IKUSEI_PV_THRESHOLDを変更してください）。'
    )
elif len(育成) >= IKUSEI_MAX:
    suggestions.append(
        f'💡 育成候補が上限{IKUSEI_MAX}件に達しています。スタッフが全件対応できない場合は'
        f'PV閾値を{IKUSEI_PV_THRESHOLD}→{IKUSEI_PV_THRESHOLD+10}に上げて絞ることを検討してください。'
    )

# 要調査の割合チェック
要調査_ratio = len(要調査_all) / total_items * 100 if total_items > 0 else 0
if 要調査_ratio > 50:
    suggestions.append(
        f'🚨 要調査が全出品の{要調査_ratio:.0f}%（{len(要調査_all)}件）を占めています。'
        f'在庫ツールのバグが広範囲に影響している可能性が極めて高い。'
        f'Cowatechへの修正依頼を最優先にしてください。'
    )

if not suggestions:
    suggestions.append('✅ 特に問題なし。引き続き現在の方針で運営を継続してください。')

# ===== Excel ワークブック作成 =====
wb = Workbook()

# ─────────────────────────────────────────────────────────────
# Sheet 1: 📊 サマリー
# ─────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = '📊 サマリー'
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 34   # 指標名（KPI）/ 週別トレンドのラベル
ws1.column_dimensions['B'].width = 18   # 数値 / W1
ws1.column_dimensions['C'].width = 30   # 内訳・補足 / W2（テキスト表示のため幅広に）
ws1.column_dimensions['D'].width = 15   # W3
ws1.column_dimensions['E'].width = 15   # W4
ws1.column_dimensions['F'].width = 18   # 合計
ws1.column_dimensions['G'].width = 13   # W3→W4変化

def s1_section(ws, row_num, title, color='E8EAF6'):
    ws.merge_cells(f'A{row_num}:G{row_num}')
    c = ws.cell(row=row_num, column=1, value=title)
    c.font = Font(name=FONT, size=10, bold=True, color=C_DARK)
    c.fill = body_fill(color)
    c.alignment = left()
    ws.row_dimensions[row_num].height = 20
    return row_num + 1

row = 1

# タイトル
ws1.merge_cells(f'A{row}:G{row}')
c = ws1.cell(row=row, column=1,
    value=f'📊 eBay 週次レポート  {PERIOD_LABEL}  （生成日: {TODAY.strftime("%Y/%m/%d")}）')
c.font = Font(name=FONT, size=13, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_DARK)
c.alignment = center()
ws1.row_dimensions[row].height = 28
row += 1

# KPIブロック
row = s1_section(ws1, row, '■ パフォーマンス指標（期間合計）')
apply_header_row(ws1, row,
    [('指標', 30), ('数値（期間合計）', 15), ('内訳・補足', 15),
     ('', 15), ('', 15), ('', 15), ('', 13)], C_DARK)
row += 1

kpi_rows = [
    ('🔵 Total Impressions（表示回数）', f'{total_imps_t:,.0f}',  '期間合計（週別分割不可）', False),
    ('  ├ Promoted Listings',           f'{total_pl_imps:,.0f}', f'{total_pl_imps/total_imps_t*100:.1f}% of total', True),
    ('  ├ Promoted Offsite',            f'{total_off_imps:,.0f}',f'{total_off_imps/total_imps_t*100:.1f}% of total', True),
    ('  └ Organic',                     f'{total_org_imps:,.0f}',f'{total_org_imps/total_imps_t*100:.1f}% of total', True),
    ('👁 Listing Views（閲覧数）',      f'{total_pv_t:,.0f}',    '期間合計', False),
    ('🛒 Quantity Sold（販売数）',      f'{total_sold_t:.0f} 件','期間合計', False),
    ('📈 Click-through Rate（CTR）',    f'{overall_ctr:.3f}%',   '期間平均（PV÷Impressions）', False),
    ('💰 Sales Conversion Rate（CVR）', f'{overall_cvr:.3f}%',   '期間平均（Sold÷PV）', False),
    ('📦 総出品数',                     f'{total_items:,} 件',   f'売上ゼロ: {zero_sold_cnt:,}件 ({zero_sold_cnt/total_items*100:.1f}%)', False),
]
for label, val, note, is_sub in kpi_rows:
    bg = C_GRAY_BG if is_sub else None
    apply_body_row(ws1, row, [label, val, note, '', '', '', ''], bg_color=bg)
    # C列〜G列をマージして内訳・補足テキストが隠れないようにする
    ws1.merge_cells(f'C{row}:G{row}')
    ws1.cell(row=row, column=3).alignment = left()
    row += 1

row += 1

# 週別売上トレンド
row = s1_section(ws1, row, '■ 週別売上トレンド（Transaction Report USD取引のみ）')
apply_header_row(ws1, row,
    [('指標', 30),
     ('W1\n2/13-19\n(7日)', 15), ('W2\n2/20-26\n(7日)', 15),
     ('W3\n2/27-3/5\n(7日)', 15), ('W4\n3/6-15\n(10日)', 15),
     ('合計', 15), ('W3→W4\n変化', 13)],
    C_DARK, row_height=42)
row += 1

net_after_plp = total_gross - total_fvf - total_intl - total_plg_fee_tx - total_off_fee - PLP_FEE_TOTAL

w_data = [
    ('注文数',
     [f'{ws["orders"]}件' for ws in week_stats], f'{total_orders}件',
     pct_change(week_stats[3]['orders'], week_stats[2]['orders']), True, False),
    ('販売数（個）',
     [f'{ws["qty"]}個' for ws in week_stats], f'{total_qty}個',
     pct_change(week_stats[3]['qty'], week_stats[2]['qty']), False, False),
    ('売上 Gross（USD）',
     [f'${ws["gross"]:,.0f}' for ws in week_stats], f'${total_gross:,.0f}',
     pct_change(week_stats[3]['gross'], week_stats[2]['gross']), True, False),
    ('FVF（最終価値手数料）',
     [f'-${ws["fvf"]:,.0f}' for ws in week_stats], f'-${total_fvf:,.0f}',
     pct_change(week_stats[3]['fvf'], week_stats[2]['fvf']), False, True),
    ('International fee',
     [f'-${ws["intl"]:,.0f}' for ws in week_stats], f'-${total_intl:,.0f}',
     pct_change(week_stats[3]['intl'], week_stats[2]['intl']), False, True),
    ('PLG広告費',
     [f'-${ws["plg_fee"]:,.0f}' for ws in week_stats], f'-${total_plg_fee_tx:,.0f}',
     pct_change(week_stats[3]['plg_fee'], week_stats[2]['plg_fee']), False, True),
    ('Offsite広告費',
     [f'-${ws["off_fee"]:,.0f}' for ws in week_stats], f'-${total_off_fee:,.0f}',
     pct_change(week_stats[3]['off_fee'], week_stats[2]['off_fee']), False, True),
    ('手取り推定',
     [f'${ws["net"]:,.0f}' for ws in week_stats],
     f'${net_after_plp:,.0f}  ※PLP-${PLP_FEE_TOTAL:.0f}控除後',
     pct_change(week_stats[3]['net'], week_stats[2]['net']), True, False),
]

for label, week_vals, total_str, change_str, bold, is_cost in w_data:
    bg = 'FFF3E0' if is_cost else ('E3F2FD' if bold else None)
    apply_body_row(ws1, row, [label] + week_vals + [total_str, change_str],
                   bg_color=bg, bold=bold)
    row += 1

row += 1

# 収支サマリー
row = s1_section(ws1, row, '■ 収支サマリー（期間全体）')
apply_header_row(ws1, row,
    [('項目', 30), ('金額（USD）', 15), ('備考・実質率', 25),
     ('', 15), ('', 15), ('', 15), ('', 13)], C_DARK)
row += 1

finance_rows = [
    ('売上合計（Gross）',             f'${total_gross:,.2f}',          f'USD注文 {total_orders}件', True),
    ('FVF（最終価値手数料）',         f'-${total_fvf:,.2f}',           f'実質率: {total_fvf/total_gross*100:.1f}%', False),
    ('International fee',             f'-${total_intl:,.2f}',           '', False),
    ('PLG広告費（Promoted General）', f'-${total_plg_fee_tx:,.2f}',
     f'ROAS: {total_gross/total_plg_fee_tx:.1f}x' if total_plg_fee_tx>0 else '', False),
    ('Offsite広告費',                 f'-${total_off_fee:,.2f}',        'Transaction Report記載（CPC課金）', False),
    ('PLP広告費（手入力・推定）',     f'-${PLP_FEE_TOTAL:,.2f}',       '⚠ Transaction Reportに含まれないため手入力', False),
    ('eBay控除後 手取り（推定）',     f'${net_after_plp:,.2f}',        f'収益率: {net_after_plp/total_gross*100:.1f}%', True),
]
for label, amount, note, bold in finance_rows:
    if '手取り' in label:  bg = 'E8F5E9'
    elif amount.startswith('-'): bg = 'FFF3E0'
    else: bg = None
    apply_body_row(ws1, row, [label, amount, note, '', '', '', ''], bg_color=bg, bold=bold)
    # C列〜G列をマージして備考テキストが隠れないようにする
    ws1.merge_cells(f'C{row}:G{row}')
    ws1.cell(row=row, column=3).alignment = left()
    row += 1

row += 1

# アクション欄
row = s1_section(ws1, row, '■ 今週のアクション（スタッフ記入欄）', color='FFF9C4')
apply_header_row(ws1, row,
    [('カテゴリ', 30), ('今週やること（記入）', 40), ('担当', 12),
     ('', 15), ('', 15), ('', 15), ('', 13)], '455A64')
row += 1
for action, default_text in [
    ('🔥 売れ筋強化',      '在庫確認・補充予定'),
    ('🚨 コア落ち対応',    'コア落ちシートを確認→原因を調査・対処'),
    ('⭐ 準売れ筋チェック','価格・タイトル改善対象を選定'),
    ('🌱 育成候補対応',    '価格・競合・在庫を確認（PV多い順に対処）'),
    ('⚠️ 要調査商品対応',  '在庫ツール・仕入先URL確認'),
    ('🗑 削除実施',        '即削除リストを確認・取り下げ実行'),
]:
    apply_body_row(ws1, row, [action, default_text, '', '', '', '', ''], bg_color='FAFAFA')
    row += 1

row += 1

# 💡 AIからの自動提案セクション
row = s1_section(ws1, row, '■ 💡 AIからの自動提案', color='E8F5E9')
apply_header_row(ws1, row,
    [('提案内容', 90), ('', 15), ('', 15), ('', 15), ('', 15), ('', 15), ('', 13)],
    '2E7D32')
row += 1
for sug in suggestions:
    ws1.merge_cells(f'A{row}:G{row}')
    c = ws1.cell(row=row, column=1, value=sug)
    c.font = Font(name=FONT, size=9, color='1B5E20')
    c.fill = body_fill('F1F8E9')
    c.alignment = left()
    c.border = border_thin()
    ws1.row_dimensions[row].height = 28
    row += 1

# ─────────────────────────────────────────────────────────────
# Sheet 2: 🔥 コア売れ筋TOP15（先週比・先週カテゴリ追加）
# ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('🔥 コア売れ筋TOP15')
ws2.sheet_view.showGridLines = False
# A:順位 B:タイトル C:ItemID D:販売数 E:先週比Δ F:CVR G:PV H:インプ I:在庫 J:広告売上 K:掲載日数 L:先週カテゴリ
for col, w in zip('ABCDEFGHIJKL', [W_RANK, W_TITLE, W_ITEM_ID, W_SOLD, W_SCORE,
                                    W_CVR, W_PV, W_IMPS, W_QTY, W_AD_SALES,
                                    W_DAYS, W_PREV_CAT]):
    ws2.column_dimensions[col].width = w

ws2.merge_cells('A1:L1')
c = ws2['A1']
c.value = f'🔥 コア売れ筋 TOP15（販売数ランキング）― 前回期間：{LAST_PERIOD}'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_GREEN_HDR)
c.alignment = center()
ws2.row_dimensions[1].height = 24

hdr2 = [
    ('順位', 4), ('商品タイトル', 52), ('Item ID', 14),
    ('販売数', 8), ('前週比\nΔ販売数', 9), ('CVR', 8), ('PV', 9),
    ('インプ', 11), ('在庫数', 8), ('広告売上', 11), ('掲載日数', 8), ('前週カテゴリ', 14),
]
apply_header_row(ws2, 2, hdr2, C_GREEN_HDR)

for rank, item in enumerate(top15, 1):
    ad_s     = f'${item["ad_sales"]:,.0f}' if item['ad_sales'] > 0 else '-'
    delta    = fmt_delta_sold(item['sold'], item['id'])
    prev_cat = get_prev_category(item['id'])
    bg       = 'C8E6C9' if rank <= 3 else C_GREEN_BG
    apply_body_row(ws2, rank+2, [
        rank, item['title'], item['id'],
        int(item['sold']), delta, f'{item["cvr"]:.1f}%', int(item['pv']),
        f'{item["imps"]:,.0f}', int(item['qty']), ad_s,
        f'{item["days"]}日', prev_cat,
    ], bg_color=bg, height=20)

# ─────────────────────────────────────────────────────────────
# Sheet 3: 🚨 コア落ちアラート（新シート）
# ─────────────────────────────────────────────────────────────
ws_alert = wb.create_sheet('🚨 コア落ち')
ws_alert.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHIJ', [W_TITLE, W_ITEM_ID, W_SOLD, W_SOLD, W_SCORE,
                                  W_CVR, W_QTY, W_DAYS, W_PREV_CAT, W_REASON]):
    ws_alert.column_dimensions[col].width = w

ws_alert.merge_cells('A1:J1')
c = ws_alert['A1']
c.value = f'🚨 コア落ちアラート ― 先週TOP15から外れた商品（{len(コア落ち)}件）前回：{LAST_PERIOD}'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_ALERT_HDR)
c.alignment = center()
ws_alert.row_dimensions[1].height = 24

if not last_data:
    # 初回実行の場合はメッセージのみ
    ws_alert.merge_cells('A2:J2')
    msg = ws_alert['A2']
    msg.value = '初回実行のため前週データがありません。来週から自動的にコア落ち商品が表示されます。'
    msg.font = Font(name=FONT, size=10, color='AD1457')
    msg.fill = body_fill('FCE4EC')
    msg.alignment = left()
    ws_alert.row_dimensions[2].height = 24
elif not コア落ち:
    ws_alert.merge_cells('A2:J2')
    msg = ws_alert['A2']
    msg.value = '✅ 今週はコア落ちなし。全員が先週と同様にTOP15をキープしています。'
    msg.font = Font(name=FONT, size=10, color='1B5E20')
    msg.fill = body_fill('E8F5E9')
    msg.alignment = left()
    ws_alert.row_dimensions[2].height = 24
else:
    # 説明行
    ws_alert.merge_cells('A2:J2')
    note_a = ws_alert['A2']
    note_a.value = ('→ 先週のコア売れ筋（TOP15）から外れた商品。在庫切れ・ツールバグ・競合激化などが原因。'
                    '必ず今週中に原因を調査して対処すること。')
    note_a.font = Font(name=FONT, size=9, color='AD1457')
    note_a.fill = body_fill('FCE4EC')
    note_a.alignment = left()
    ws_alert.row_dimensions[2].height = 20

    hdr_a = [
        ('商品タイトル', 52), ('Item ID', 14),
        ('前週sold', 9), ('今週sold', 9), ('Δ販売数', 9),
        ('今週CVR', 8), ('在庫数', 8), ('今週カテゴリ', 8),
        ('今週カテゴリ', 14), ('考えられる原因（自動推測）', 30),
    ]
    apply_header_row(ws_alert, 3, hdr_a, C_ALERT_HDR)

    for i, ko in enumerate(コア落ち, 4):
        delta = ko['delta_sold']
        delta_str = f'▼{delta}' if delta < 0 else f'▲+{delta}' if delta > 0 else '±0'
        apply_body_row(ws_alert, i, [
            ko['title'], ko['id'],
            int(ko['prev_sold']), int(ko['curr_sold']), delta_str,
            f'{ko["curr_cvr"]:.1f}%', ko['qty'],
            ko['curr_cat'], ko['curr_cat'], ko['reason'],
        ], bg_color=C_ALERT_BG, height=20)

# ─────────────────────────────────────────────────────────────
# Sheet 4: ⭐ 準売れ筋（ポテンシャルスコア・優先度追加）
# ─────────────────────────────────────────────────────────────
ws3 = wb.create_sheet('⭐ 準売れ筋')
ws3.sheet_view.showGridLines = False
# 列: 優先度 / タイトル / ItemID / ポテンシャル / 販売数 / CVR / PV / インプ / 在庫 / 掲載日数 / 前週カテゴリ
for col, w in zip('ABCDEFGHIJK', [W_RANK, W_TITLE, W_ITEM_ID, W_SCORE,
                                   W_SOLD, W_CVR, W_PV, W_IMPS,
                                   W_QTY, W_DAYS, W_PREV_CAT]):
    ws3.column_dimensions[col].width = w

ws3.merge_cells('A1:K1')
c = ws3['A1']
c.value = f'⭐ 準売れ筋（{len(準売れ筋)}件）― ポテンシャルスコア順（CVR×PV）― S=最重点10件 / A=重点20件 / B=普通50件 / C=後回し'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_BLUE_HDR)
c.alignment = center()
ws3.row_dimensions[1].height = 24

ws3.merge_cells('A2:K2')
note3 = ws3['A2']
note3.value = ('【優先度の見方】S🔥→今週必ず対処 / A⭐→今週中に確認 / B🌱→来週以降 / C—→後回し  '
               '【ポテンシャル】CVR×PV：高いほど「少し手を加えれば大きく伸びる」商品')
note3.font = Font(name=FONT, size=9, color='0D47A1')
note3.fill = body_fill('E3F2FD')
note3.alignment = left()
ws3.row_dimensions[2].height = 20

hdr3 = [
    ('優先度', W_RANK), ('商品タイトル', W_TITLE), ('Item ID', W_ITEM_ID),
    ('ポテンシャル\n(CVR×PV)', W_SCORE), ('販売数', W_SOLD), ('CVR', W_CVR),
    ('PV', W_PV), ('インプ', W_IMPS), ('在庫数', W_QTY),
    ('掲載日数', W_DAYS), ('前週カテゴリ', W_PREV_CAT),
]
apply_header_row(ws3, 3, hdr3, C_BLUE_HDR)

# 優先度ごとに背景色を変える
PRIORITY_BG = {'S 🔥': 'C8E6C9', 'A ⭐': C_BLUE_BG, 'B 🌱': C_YELLOW_BG, 'C —': C_GRAY_BG}

for i, item in enumerate(準売れ筋, 4):
    prev_cat = get_prev_category(item['id'])
    bg       = PRIORITY_BG.get(item['priority'], C_BLUE_BG)
    bold     = item['priority'] == 'S 🔥'   # S優先度は太字で強調
    apply_body_row(ws3, i, [
        item['priority'], item['title'], item['id'],
        item['pot_score'], int(item['sold']), f'{item["cvr"]:.1f}%',
        int(item['pv']), f'{item["imps"]:,.0f}', int(item['qty']),
        f'{item["days"]}日', prev_cat,
    ], bg_color=bg, bold=bold)

# ─────────────────────────────────────────────────────────────
# Sheet 5: 🌱 育成候補（v3改訂：PV50以上・要調査除外・上位20件）
# ─────────────────────────────────────────────────────────────
ws4 = wb.create_sheet('🌱 育成候補')
ws4.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHIJ', [W_TITLE, W_ITEM_ID, W_PV, W_IMPS, W_CVR,
                                  W_QTY, W_DAYS, W_PROMO, W_ORGANIC, W_PREV_CAT]):
    ws4.column_dimensions[col].width = w

ws4.merge_cells('A1:J1')
c = ws4['A1']
c.value = (f'🌱 育成候補（{len(育成)}件）'
           f'― PV{IKUSEI_PV_THRESHOLD}以上・売上ゼロ・要調査除外 '
           f'― クリックされているが購入に繋がっていない（価格・ページ・送料が原因）')
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_YELLOW_HDR)
c.alignment = center()
ws4.row_dimensions[1].height = 24

ws4.merge_cells('A2:J2')
note4 = ws4['A2']
note4.value = ('【v3改訂】要調査（在庫ツールバグ疑い）は除外済。このリストはページが見られているのに買われていない商品。'
               '確認項目：① 価格（競合比較）② サイズ・在庫 ③ タイトル・写真の質 ④ 送料設定')
note4.font = Font(name=FONT, size=9, color='E65100')
note4.fill = body_fill('FFF9C4')
note4.alignment = left()
ws4.row_dimensions[2].height = 20

hdr4 = [
    ('商品タイトル', 52), ('Item ID', 14),
    ('PV', 9), ('インプ', 11), ('CTR', 8),
    ('在庫数', 8), ('掲載日数', 8), ('広告状態', 13), ('Organic比', 11), ('前週カテゴリ', 14),
]
apply_header_row(ws4, 3, hdr4, C_YELLOW_HDR, txt_color='FFFFFF')

for i, item in enumerate(育成, 4):
    org_ratio = f'{item["org_imps"]/item["imps"]*100:.0f}%' if item['imps'] > 0 else '-'
    prev_cat  = get_prev_category(item['id'])
    apply_body_row(ws4, i, [
        item['title'], item['id'],
        int(item['pv']), f'{item["imps"]:,.0f}', f'{item["ctr"]:.2f}%',
        int(item['qty']), f'{item["days"]}日', item['promo'], org_ratio, prev_cat,
    ], bg_color=C_YELLOW_BG)

# ─────────────────────────────────────────────────────────────
# Sheet 6: ⚠️ 要調査（先週カテゴリ追加）
# ─────────────────────────────────────────────────────────────
ws5 = wb.create_sheet('⚠️ 要調査')
ws5.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHIJK', [W_TITLE, W_ITEM_ID, W_IMPS, W_PV, W_QTY,
                                   W_DAYS, W_PROMO, W_ORGANIC, W_PREV_CAT,
                                   W_MEMO, W_MEMO]):
    ws5.column_dimensions[col].width = w

ws5.merge_cells('A1:K1')
c = ws5['A1']
c.value = (f'⚠️ 要調査 TOP50（全{len(要調査_all)}件中）'
           f'― インプ500以上・売上ゼロ・掲載90日以上 ― 在庫ツールバグ or 仕入先URL切れ疑い')
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_ORANGE_HDR)
c.alignment = center()
ws5.row_dimensions[1].height = 24

ws5.merge_cells('A2:K2')
note5 = ws5['A2']
note5.value = (f'⚠ 全{len(要調査_all)}件/{total_items}件({len(要調査_all)/total_items*100:.0f}%)がインプあり売上ゼロ'
               f' — 前週カテゴリで「⚠️ 要調査→今週も要調査」の商品は優先して調査すること'
               f' | 確認①在庫ツールURLが有効か ②仕入先に在庫があるか ③価格・競合確認')
note5.font = Font(name=FONT, size=9, color='BF360C')
note5.fill = body_fill('FFF3E0')
note5.alignment = left()
ws5.row_dimensions[2].height = 20

hdr5 = [
    ('商品タイトル', 52), ('Item ID', 14),
    ('インプ', 11), ('PV', 9), ('在庫数', 8),
    ('掲載日数', 8), ('広告状態', 13), ('Organic比', 11),
    ('前週カテゴリ', 14), ('調査メモ（記入）', 20), ('確認状況（記入）', 18),
]
apply_header_row(ws5, 3, hdr5, C_ORANGE_HDR)

for i, item in enumerate(要調査, 4):
    org_ratio = f'{item["org_imps"]/item["imps"]*100:.0f}%' if item['imps'] > 0 else '-'
    prev_cat  = get_prev_category(item['id'])
    # 前週も要調査だったら強調（オレンジ濃い）
    bg = 'FFCC80' if prev_cat == '⚠️ 要調査' else C_ORANGE_BG
    apply_body_row(ws5, i, [
        item['title'], item['id'],
        f'{item["imps"]:,.0f}', int(item['pv']), int(item['qty']),
        f'{item["days"]}日', item['promo'], org_ratio,
        prev_cat, '', '',
    ], bg_color=bg)

# ─────────────────────────────────────────────────────────────
# Sheet 7: 🗑 削除候補
# ─────────────────────────────────────────────────────────────
ws6 = wb.create_sheet('🗑 削除候補')
ws6.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHI', [W_TITLE, W_ITEM_ID, W_QTY, W_DAYS, W_PROMO,
                                 W_MEMO, W_MEMO, W_PREV_CAT, W_RANK]):
    ws6.column_dimensions[col].width = w

ws6.merge_cells('A1:I1')
c = ws6['A1']
c.value = f'🗑 削除候補 — L1即削除: {len(削除L1)}件 ／ L2要確認削除: {len(削除L2)}件'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_RED_HDR)
c.alignment = center()
ws6.row_dimensions[1].height = 24

ws6.merge_cells('A2:I2')
note6 = ws6['A2']
note6.value = '⚠️ 削除前に必ずeBayで確認：① ウォッチ数（関心度）② 生涯販売数（過去実績） → 両方ゼロに近い場合のみ削除推奨'
note6.font = Font(name=FONT, size=9, color='B71C1C', bold=True)
note6.fill = body_fill('FFEBEE')
note6.alignment = left()
ws6.row_dimensions[2].height = 20

apply_header_row(ws6, 3,
    [('【L1：即削除】インプ・PV・売上すべてゼロ（{0}件）'.format(len(削除L1)), 52),
     ('Item ID', 14), ('在庫数', 8), ('掲載日数', 8), ('広告状態', 13),
     ('ウォッチ数\n(eBayで確認)', 14), ('生涯販売数\n(eBayで確認)', 14),
     ('前週カテゴリ', 14), ('削除済✓', 8)],
    C_RED_HDR)

r = 4
for item in 削除L1:
    prev_cat = get_prev_category(item['id'])
    apply_body_row(ws6, r, [
        item['title'], item['id'],
        int(item['qty']), f'{item["days"]}日', item['promo'],
        '←eBayで確認', '←eBayで確認', prev_cat, '',
    ], bg_color='FFCDD2')
    r += 1

r += 1

ws6.merge_cells(f'A{r}:I{r}')
c2 = ws6.cell(row=r, column=1,
    value=f'【L2：要確認削除】掲載180日以上・インプ50未満・売上ゼロ（{len(削除L2)}件）― 削除前に上記の確認必須')
c2.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
c2.fill = hdr_fill(C_RED_HDR)
c2.alignment = left()
ws6.row_dimensions[r].height = 20
r += 1

apply_header_row(ws6, r,
    [('商品タイトル', 52), ('Item ID', 14), ('インプ', 11), ('PV', 8),
     ('在庫数', 8), ('掲載日数', 8), ('ウォッチ数\n(eBayで確認)', 14),
     ('前週カテゴリ', 14), ('削除済✓', 8)],
    '7B1FA2')
r += 1

for item in 削除L2:
    prev_cat = get_prev_category(item['id'])
    apply_body_row(ws6, r, [
        item['title'], item['id'],
        f'{item["imps"]:,.0f}', int(item['pv']),
        int(item['qty']), f'{item["days"]}日', '←eBayで確認',
        prev_cat, '',
    ], bg_color='F3E5F5')
    r += 1

# ─────────────────────────────────────────────────────────────
# Sheet 8: 📋 週次履歴（蓄積）
# ─────────────────────────────────────────────────────────────
ws7 = wb.create_sheet('📋 週次履歴（蓄積）')
ws7.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHIJ', [14, 22, 8, 13, 13, 8, 8, 8, 8, 10]):
    ws7.column_dimensions[col].width = w

ws7.merge_cells('A1:J1')
c = ws7['A1']
c.value = '📋 週次履歴 ― レポート実行ごとに自動蓄積（前週比較の基礎データ）'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_DARK)
c.alignment = center()
ws7.row_dimensions[1].height = 24

ws7.merge_cells('A2:J2')
note7 = ws7['A2']
note7.value = '→ このシートは毎週レポート実行時に自動更新されます（weekly_history.jsonに保存）。削除しないこと。'
note7.font = Font(name=FONT, size=9, color=C_DARK)
note7.fill = body_fill('E8EAF6')
note7.alignment = left()
ws7.row_dimensions[2].height = 18

hdr7 = [
    ('レポート日', 14), ('期間', 22), ('注文数', 8),
    ('売上Gross', 13), ('手取り推定', 13), ('CVR', 8), ('CTR', 8),
    ('出品数', 8), ('コア落ち', 8), ('要調査件数', 10),
]
apply_header_row(ws7, 3, hdr7, C_DARK)

# 過去の履歴データを表示（新しい順）
history_row = 4
for date_key in sorted(history.keys(), reverse=True):
    hd = history[date_key]
    apply_body_row(ws7, history_row, [
        date_key,
        hd.get('period', '—'),
        hd.get('total_orders', '—'),
        hd.get('total_gross', '—'),
        hd.get('net_after_plp', '—'),
        hd.get('overall_cvr', '—'),
        hd.get('overall_ctr', '—'),
        hd.get('total_items', '—'),
        '—',
        hd.get('要調査_count', '—'),
    ], bg_color='E3F2FD', bold=False)
    history_row += 1

# 今週のデータを追加（まだ保存前なので直接表示）
apply_body_row(ws7, history_row, [
    TODAY.strftime('%Y/%m/%d') + '（今週）',
    PERIOD_LABEL,
    total_orders,
    f'${total_gross:,.0f}',
    f'${net_after_plp:,.0f}',
    f'{overall_cvr:.3f}%',
    f'{overall_ctr:.3f}%',
    total_items,
    len(コア落ち),
    len(要調査_all),
], bg_color='C8E6C9', bold=True)

# ===== 保存 =====
wb.save(OUTPUT)

# ===== 週次履歴を保存（JSONファイルに書き出し）=====
current_week_data = {
    'period':        PERIOD_LABEL,
    'total_orders':  total_orders,
    'total_gross':   f'${total_gross:,.0f}',
    'net_after_plp': f'${net_after_plp:,.0f}',
    'overall_cvr':   f'{overall_cvr:.3f}%',
    'overall_ctr':   f'{overall_ctr:.3f}%',
    'total_items':   total_items,
    '要調査_count':  len(要調査_all),
    'top15_ids':     list(top15_ids),
    '準売れ筋_ids':  [i['id'] for i in 準売れ筋],
    '育成_ids':      [i['id'] for i in 育成],
    '要調査_ids':    list(要調査_ids),
    '削除L1_ids':    [i['id'] for i in 削除L1],
    '削除L2_ids':    [i['id'] for i in 削除L2],
    'all_ids':       [i['id'] for i in items],
    'per_item': {
        i['id']: {
            'sold':  i['sold'],
            'cvr':   i['cvr'],
            'pv':    i['pv'],
            'imps':  i['imps'],
            'title': i['title'][:60],
        }
        for i in items
    }
}
history[HISTORY_DATE_KEY] = current_week_data
save_history(history)

# ===== 完了メッセージ =====
print(f'✅ レポート生成完了: {OUTPUT}')
print(f'   💾 週次履歴保存: {HISTORY_FILE}')
print(f'   シート構成（8シート）:')
print(f'     📊 サマリー / 🔥 コア売れ筋TOP15({len(top15)}件) / 🚨 コア落ち({len(コア落ち)}件)')
print(f'     ⭐ 準売れ筋({len(準売れ筋)}件) / 🌱 育成候補({len(育成)}件, PV{IKUSEI_PV_THRESHOLD}+)')
print(f'     ⚠️ 要調査({len(要調査)}件/全{len(要調査_all)}件) / 🗑 削除候補(L1:{len(削除L1)}+L2:{len(削除L2)}件)')
print(f'     📋 週次履歴')
print(f'')
print(f'   💡 AIからの提案（{len(suggestions)}件）:')
for s in suggestions:
    print(f'     {s}')
