import csv, io, sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

sys.stdout.reconfigure(encoding='utf-8')

# ===== データ読み込み =====
traffic_file = r"C:\Users\KEISUKE SHIMOMOTO\Downloads\eBay-ListingsTrafficReport-Mar-15-2026-21_56_24-0700-13288691549.csv"
ads_file     = r"C:\Users\KEISUKE SHIMOMOTO\Downloads\rioxxrinaxjapan_advertising_sales_report_20260315.csv"
OUTPUT       = r"C:\Users\KEISUKE SHIMOMOTO\Desktop\reffort\commerce\ebay\analytics\eBay週次レポート_20260315.xlsx"

def to_f(s):
    try: return float(str(s).replace(',','').replace('%','').strip())
    except: return 0.0

def clean_id(s):
    return s.strip().replace('=','').replace('"','')

# Traffic Report
with open(traffic_file, encoding='utf-8-sig') as f:
    lines = f.readlines()
data = io.StringIO(''.join(lines[5:]))
rows = list(csv.DictReader(data))

items = []
for row in rows:
    title = row.get('Listing title','').strip()
    if not title: continue
    items.append({
        'title': title,
        'id': clean_id(row.get('eBay item ID','')),
        'imps': to_f(row.get('Total impressions',0)),
        'pv': to_f(row.get('Total page views',0)),
        'sold': to_f(row.get('Quantity sold',0)),
        'cvr': to_f(row.get('Sales conversion rate = Quantity sold/Total page views',0)),
        'qty': to_f(row.get('Quantity available',0)),
        'promo': row.get('Current promoted listings status','').strip(),
    })

# Advertising Report
with open(ads_file, encoding='utf-8-sig') as f:
    ads_lines = f.readlines()
ads_data = io.StringIO(''.join(ads_lines[2:]))
ad_by_item = defaultdict(lambda: {'plg_s':0,'plg_fee':0,'plp_s':0,'off_s':0})

def to_f2(s):
    try: return float(str(s).replace('$','').replace(',','').strip())
    except: return 0.0

for row in csv.DictReader(ads_data):
    iid = row.get('Item ID','').strip()
    ad = row.get('Ad type',''); st = row.get('Campaign strategy','')
    sale_type = row.get('Sale type','')
    s = to_f2(row.get('Sales','0'))
    fee = to_f2(row.get('Ad fees (General)','-'))
    if 'Offsite' in ad:
        ad_by_item[iid]['off_s'] += s
    elif 'Priority' in st:
        ad_by_item[iid]['plp_s'] += s
    elif 'General' in st and 'Halo' not in sale_type:
        ad_by_item[iid]['plg_s'] += s
        ad_by_item[iid]['plg_fee'] += fee

for item in items:
    ad = ad_by_item.get(item['id'], {})
    item['plg_s']   = ad.get('plg_s', 0)
    item['plg_fee'] = ad.get('plg_fee', 0)
    item['plp_s']   = ad.get('plp_s', 0)
    item['off_s']   = ad.get('off_s', 0)
    item['ad_sales']= item['plg_s'] + item['plp_s'] + item['off_s']

# 集計
total = len(items)
total_imps = sum(i['imps'] for i in items)
total_pv   = sum(i['pv'] for i in items)
total_sold = sum(i['sold'] for i in items)
zero_sold  = sum(1 for i in items if i['sold'] == 0)
zero_pv    = sum(1 for i in items if i['pv'] == 0)
overall_ctr = total_pv / total_imps * 100 if total_imps > 0 else 0
overall_cvr = total_sold / total_pv * 100 if total_pv > 0 else 0

top15  = sorted(items, key=lambda x: -x['sold'])[:15]
育成30  = sorted([i for i in items if i['sold'] == 0 and i['pv'] >= 20], key=lambda x: -x['pv'])[:30]
dead   = [i for i in items if i['imps'] == 0 and i['pv'] == 0 and i['sold'] == 0]
sold_items = sorted([i for i in items if i['sold'] > 0], key=lambda x: -x['sold'])

# ===== スタイル定義 =====
FONT = 'Arial'

def hdr_font(size=11, bold=True, color='FFFFFF'):
    return Font(name=FONT, size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color='000000'):
    return Font(name=FONT, size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=False)

def left_wrap():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

# カラーパレット
C_NAVY    = '1F3864'  # ヘッダー背景
C_ORANGE  = 'E84B37'  # アクセント
C_GREEN   = '1E7F5E'  # 売れ筋
C_YELLOW  = 'FFF2CC'  # 育成候補
C_RED_BG  = 'FFE0E0'  # 削除候補
C_GRAY    = 'F5F5F5'  # 偶数行背景
C_WHITE   = 'FFFFFF'
C_DARK_GREEN = '134F31'

# ===== Workbook作成 =====
wb = Workbook()

# ===== Sheet 1: サマリー =====
ws = wb.active
ws.title = '📊 サマリー'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 22
ws.row_dimensions[1].height = 10

# タイトル
ws.merge_cells('A2:D2')
ws['A2'] = 'eBay 週次パフォーマンスレポート'
ws['A2'].font = Font(name=FONT, size=18, bold=True, color=C_NAVY)
ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[2].height = 36

ws.merge_cells('A3:D3')
ws['A3'] = '期間: 2026年2月13日 〜 2026年3月15日'
ws['A3'].font = Font(name=FONT, size=11, color='666666')
ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[3].height = 22

ws.row_dimensions[4].height = 10

# KPIボックス（行5〜10）
kpi_headers = ['出品数', '売上ゼロ商品', '全体CTR', '全体CVR']
kpi_vals    = [
    f"{total:,}件",
    f"{zero_sold:,}件 ({zero_sold/total*100:.1f}%)",
    f"{overall_ctr:.3f}%",
    f"{overall_cvr:.3f}%",
]
kpi_sub = [
    '現在アクティブな出品',
    f'うち完全死蔵: {len(dead)}件',
    '表示 → 閲覧率',
    '閲覧 → 購入率',
]
kpi_cols = ['A','B','C','D']
kpi_colors = [C_NAVY, C_ORANGE, '2E75B6', C_GREEN]

for col_letter, hdr, val, sub, color in zip(kpi_cols, kpi_headers, kpi_vals, kpi_sub, kpi_colors):
    ws[f'{col_letter}5'] = hdr
    ws[f'{col_letter}5'].font = hdr_font(size=10, color=color)
    ws[f'{col_letter}5'].alignment = center()
    ws.row_dimensions[5].height = 18

    ws[f'{col_letter}6'] = val
    ws[f'{col_letter}6'].font = Font(name=FONT, size=20, bold=True, color=color)
    ws[f'{col_letter}6'].alignment = center()
    ws.row_dimensions[6].height = 38

    ws[f'{col_letter}7'] = sub
    ws[f'{col_letter}7'].font = Font(name=FONT, size=9, color='888888')
    ws[f'{col_letter}7'].alignment = center()
    ws.row_dimensions[7].height = 16

ws.row_dimensions[8].height = 12

# 広告費サマリー
ad_section_hdr = [['広告種別','売上（帰属）','広告費','費用率 / ROAS'],
                  ['PLG（成果報酬型）', None, None, None],
                  ['PLP（クリック課金）', None, None, None],
                  ['Offsite（クリック課金）', None, None, None],
                  ['合計', None, None, None]]

plg_sales_total = sum(i['plg_s'] for i in items)
plg_fee_total   = sum(i['plg_fee'] for i in items)
plp_sales_total = sum(i['plp_s'] for i in items)
off_sales_total = sum(i['off_s'] for i in items)
total_ad_sales  = plg_sales_total + plp_sales_total + off_sales_total
plg_fee_actual  = 3446.17
plp_fee_actual  = 490.97
off_fee_actual  = 2399.37

ad_rows = [
    ['広告種別', '帰属売上', '広告費（実費）', '費用率'],
    ['PLG（成果報酬型）',   f'${plg_sales_total:,.0f}', f'${plg_fee_actual:,.0f}', f'{plg_fee_actual/plg_sales_total*100:.1f}%' if plg_sales_total else '-'],
    ['PLP（クリック課金）', f'${plp_sales_total:,.0f}', f'${plp_fee_actual:,.0f}', f'{plp_fee_actual/plp_sales_total*100:.1f}%' if plp_sales_total else '-'],
    ['Offsite（クリック課金）', f'${off_sales_total:,.0f}', f'${off_fee_actual:,.0f}', f'{off_fee_actual/off_sales_total*100:.1f}%' if off_sales_total else '-'],
    ['合計', f'${total_ad_sales:,.0f}', f'${plg_fee_actual+plp_fee_actual+off_fee_actual:,.0f}',
     f'{(plg_fee_actual+plp_fee_actual+off_fee_actual)/total_ad_sales*100:.1f}%' if total_ad_sales else '-'],
]

ws.merge_cells('A9:D9')
ws['A9'] = '広告パフォーマンス（2/13〜3/15）'
ws['A9'].font = Font(name=FONT, size=12, bold=True, color=C_NAVY)
ws['A9'].fill = fill('EBF3FB')
ws['A9'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[9].height = 24

for r_idx, row_data in enumerate(ad_rows, start=10):
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border()
        cell.alignment = center()
        if r_idx == 10:  # ヘッダー行
            cell.font = hdr_font(size=10, color='FFFFFF')
            cell.fill = fill(C_NAVY)
        elif r_idx == 14:  # 合計行
            cell.font = Font(name=FONT, size=10, bold=True, color=C_NAVY)
            cell.fill = fill('DDEEFF')
        else:
            cell.font = body_font()
            cell.fill = fill(C_GRAY if r_idx % 2 == 0 else C_WHITE)
    ws.row_dimensions[r_idx].height = 20

ws.row_dimensions[15].height = 12

# 収支サマリー（簡易）
ws.merge_cells('A16:D16')
ws['A16'] = '収支概要（eBay控除後）'
ws['A16'].font = Font(name=FONT, size=12, bold=True, color=C_NAVY)
ws['A16'].fill = fill('EBF3FB')
ws['A16'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[16].height = 24

finance_rows = [
    ['項目', '金額', '備考', ''],
    ['商品売上', '$56,540', 'USD注文341件', ''],
    ['送料収入', '$8,060', '', ''],
    ['総収入', '$64,600', '', ''],
    ['FVF（手数料）', '-$5,455', '商品売上の9.6%', ''],
    ['International fee', '-$525', '0.8%', ''],
    ['広告費合計（PLG+PLP+Offsite）', '-$6,336', '収入の9.8%', ''],
    ['eBay控除後手取り', '$52,284', '仕入・送料・人件費は別途', '★'],
]

for r_idx, row_data in enumerate(finance_rows, start=17):
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border()
        cell.alignment = center() if c_idx > 1 else Alignment(horizontal='left', vertical='center', indent=1)
        if r_idx == 17:
            cell.font = hdr_font(size=10, color='FFFFFF')
            cell.fill = fill(C_NAVY)
        elif r_idx == 24:
            cell.font = Font(name=FONT, size=10, bold=True, color=C_GREEN)
            cell.fill = fill('D9F0E8')
        elif val and str(val).startswith('-'):
            cell.font = Font(name=FONT, size=10, color='CC0000')
            cell.fill = fill(C_GRAY if r_idx % 2 == 0 else C_WHITE)
        else:
            cell.font = body_font()
            cell.fill = fill(C_GRAY if r_idx % 2 == 0 else C_WHITE)
    ws.row_dimensions[r_idx].height = 20

ws.row_dimensions[25].height = 12

# アクションサマリー
ws.merge_cells('A26:D26')
ws['A26'] = '今週のアクション（優先順位順）'
ws['A26'].font = Font(name=FONT, size=12, bold=True, color=C_NAVY)
ws['A26'].fill = fill('EBF3FB')
ws['A26'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[26].height = 24

actions = [
    ['優先度', 'アクション', '対象', '担当'],
    ['① 最優先', '完全死蔵14件を即削除', '🗑 削除リストシート参照', '佐藤'],
    ['② 高', 'NIKE Air More Uptempo（在庫1）確認→取り下げ判断', 'ID: 357786855641', '佐藤'],
    ['③ 高', '育成候補TOP5の価格・競合チェック', '🌱 育成候補シート参照', '佐藤・須藤'],
    ['④ 中', 'PUMA Speed Cat 残りサイズの扱いを決定', 'ID: 358022227865', '社長判断'],
    ['⑤ 中', 'Mexico 66系 PLP継続可否の検討', '売れ筋TOP1〜4', '社長判断'],
]

for r_idx, row_data in enumerate(actions, start=27):
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border()
        if r_idx == 27:
            cell.font = hdr_font(size=10, color='FFFFFF')
            cell.fill = fill(C_NAVY)
            cell.alignment = center()
        else:
            cell.font = body_font()
            cell.fill = fill(C_GRAY if r_idx % 2 == 0 else C_WHITE)
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    ws.row_dimensions[r_idx].height = 28

# ===== Sheet 2: 売れ筋 =====
ws2 = wb.create_sheet('🔥 売れ筋TOP15')
ws2.sheet_view.showGridLines = False

cols2 = ['#', '商品名（フルタイトル）', 'Item ID', '販売数', 'CVR', 'ページビュー',
         'インプレッション', '広告帰属売上', 'PLG費用', '在庫数', '広告ステータス']
widths2 = [5, 65, 16, 10, 10, 14, 16, 16, 12, 10, 18]

for c, (h, w) in enumerate(zip(cols2, widths2), start=1):
    col = get_column_letter(c)
    ws2.column_dimensions[col].width = w
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hdr_font(size=10, color='FFFFFF')
    cell.fill = fill(C_DARK_GREEN)
    cell.alignment = center()
    cell.border = border()
ws2.row_dimensions[1].height = 28

for r_idx, item in enumerate(top15, start=2):
    row_fill = fill(C_GRAY) if r_idx % 2 == 0 else fill(C_WHITE)
    vals = [
        r_idx - 1,
        item['title'],
        item['id'],
        int(item['sold']),
        item['cvr'] / 100,
        int(item['pv']),
        int(item['imps']),
        item['ad_sales'] if item['ad_sales'] > 0 else None,
        item['plg_fee'] if item['plg_fee'] > 0 else None,
        int(item['qty']),
        item['promo'] if item['promo'] else '未設定',
    ]
    for c_idx, val in enumerate(vals, start=1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border()
        cell.font = body_font()
        cell.fill = row_fill
        if c_idx == 1:
            cell.alignment = center()
        elif c_idx == 2:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        elif c_idx == 5:
            cell.number_format = '0.0%'
            cell.alignment = center()
        elif c_idx in [8, 9]:
            cell.number_format = '$#,##0.00'
            cell.alignment = center()
        else:
            cell.alignment = center()
    ws2.row_dimensions[r_idx].height = 36

# CVRが低い行を警告色に（0.5%未満）
for r_idx in range(2, len(top15)+2):
    cvr_cell = ws2.cell(row=r_idx, column=5)
    if cvr_cell.value and cvr_cell.value < 0.005:
        cvr_cell.fill = fill('FFE0E0')
        cvr_cell.font = Font(name=FONT, size=10, color='CC0000', bold=True)

# ===== Sheet 3: 全売上商品 =====
ws3 = wb.create_sheet('⭐ 売上あり商品（全件）')
ws3.sheet_view.showGridLines = False

cols3 = ['#', '商品名（フルタイトル）', 'Item ID', '販売数', 'CVR', 'ページビュー',
         'インプレッション', '在庫数', '広告帰属売上', '広告ステータス']
widths3 = [5, 65, 16, 10, 10, 14, 16, 10, 16, 18]

for c, (h, w) in enumerate(zip(cols3, widths3), start=1):
    col = get_column_letter(c)
    ws3.column_dimensions[col].width = w
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font = hdr_font(size=10, color='FFFFFF')
    cell.fill = fill('2E75B6')
    cell.alignment = center()
    cell.border = border()
ws3.row_dimensions[1].height = 28

for r_idx, item in enumerate(sold_items, start=2):
    row_fill = fill(C_GRAY) if r_idx % 2 == 0 else fill(C_WHITE)
    vals = [
        r_idx - 1,
        item['title'],
        item['id'],
        int(item['sold']),
        item['cvr'] / 100,
        int(item['pv']),
        int(item['imps']),
        int(item['qty']),
        item['ad_sales'] if item['ad_sales'] > 0 else None,
        item['promo'] if item['promo'] else '未設定',
    ]
    for c_idx, val in enumerate(vals, start=1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border()
        cell.font = body_font()
        cell.fill = row_fill
        if c_idx == 1: cell.alignment = center()
        elif c_idx == 2: cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        elif c_idx == 5:
            cell.number_format = '0.0%'
            cell.alignment = center()
        elif c_idx == 9:
            cell.number_format = '$#,##0.00'
            cell.alignment = center()
        else: cell.alignment = center()
    ws3.row_dimensions[r_idx].height = 36

# ===== Sheet 4: 育成候補 =====
ws4 = wb.create_sheet('🌱 育成候補（PVあり・売上ゼロ）')
ws4.sheet_view.showGridLines = False

ws4.merge_cells('A1:G1')
ws4['A1'] = '需要はある（閲覧されている）のに売れていない商品 — 価格・競合・サイズを確認してください'
ws4['A1'].font = Font(name=FONT, size=11, bold=True, color=C_ORANGE)
ws4['A1'].fill = fill('FFF2CC')
ws4['A1'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws4.row_dimensions[1].height = 28

cols4 = ['#', '商品名（フルタイトル）', 'Item ID', 'ページビュー', 'インプレッション', '在庫数', '確認メモ']
widths4 = [5, 65, 16, 14, 16, 10, 30]

for c, (h, w) in enumerate(zip(cols4, widths4), start=1):
    col = get_column_letter(c)
    ws4.column_dimensions[col].width = w
    cell = ws4.cell(row=2, column=c, value=h)
    cell.font = hdr_font(size=10, color='FFFFFF')
    cell.fill = fill('BF8F00')
    cell.alignment = center()
    cell.border = border()
ws4.row_dimensions[2].height = 28

for r_idx, item in enumerate(育成30, start=3):
    row_fill = fill(C_YELLOW) if r_idx % 2 == 0 else fill(C_WHITE)
    # 在庫1かつPV多い = 特に要注意
    is_urgent = item['qty'] <= 1 and item['pv'] >= 500
    vals = [
        r_idx - 2,
        item['title'],
        item['id'],
        int(item['pv']),
        int(item['imps']),
        int(item['qty']),
        '⚠ 在庫1・要確認' if is_urgent else '',
    ]
    for c_idx, val in enumerate(vals, start=1):
        cell = ws4.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border()
        cell.fill = fill('FFE0B2') if is_urgent else row_fill
        cell.font = Font(name=FONT, size=10, bold=is_urgent)
        if c_idx == 1: cell.alignment = center()
        elif c_idx == 2: cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        else: cell.alignment = center()
    ws4.row_dimensions[r_idx].height = 36

# ===== Sheet 5: 削除リスト =====
ws5 = wb.create_sheet('🗑 削除リスト（完全死蔵）')
ws5.sheet_view.showGridLines = False

ws5.merge_cells('A1:E1')
ws5['A1'] = f'インプレッション・ページビュー・売上すべてゼロ → 即削除推奨（{len(dead)}件）'
ws5['A1'].font = Font(name=FONT, size=11, bold=True, color='CC0000')
ws5['A1'].fill = fill(C_RED_BG)
ws5['A1'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws5.row_dimensions[1].height = 28

cols5 = ['#', '商品名（フルタイトル）', 'Item ID', '在庫数', '削除済み✓']
widths5 = [5, 70, 16, 10, 14]

for c, (h, w) in enumerate(zip(cols5, widths5), start=1):
    col = get_column_letter(c)
    ws5.column_dimensions[col].width = w
    cell = ws5.cell(row=2, column=c, value=h)
    cell.font = hdr_font(size=10, color='FFFFFF')
    cell.fill = fill('C00000')
    cell.alignment = center()
    cell.border = border()
ws5.row_dimensions[2].height = 28

for r_idx, item in enumerate(dead, start=3):
    row_fill = fill(C_RED_BG) if r_idx % 2 == 0 else fill(C_WHITE)
    vals = [r_idx - 2, item['title'], item['id'], int(item['qty']), '']
    for c_idx, val in enumerate(vals, start=1):
        cell = ws5.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border()
        cell.font = body_font()
        cell.fill = row_fill
        if c_idx == 1: cell.alignment = center()
        elif c_idx == 2: cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        else: cell.alignment = center()
    ws5.row_dimensions[r_idx].height = 36

# ===== 保存 =====
wb.save(OUTPUT)
print(f'✅ 保存完了: {OUTPUT}')
print(f'   シート1: サマリー')
print(f'   シート2: 売れ筋TOP15 ({len(top15)}件)')
print(f'   シート3: 売上あり全件 ({len(sold_items)}件)')
print(f'   シート4: 育成候補 ({len(育成30)}件)')
print(f'   シート5: 削除リスト ({len(dead)}件)')
