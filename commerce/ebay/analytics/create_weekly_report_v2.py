import csv, io, sys
from collections import defaultdict
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

# ===== ファイルパス（毎週更新） =====
TRAFFIC_FILE = r"C:\Users\KEISUKE SHIMOMOTO\Downloads\eBay-ListingsTrafficReport-Mar-15-2026-21_56_24-0700-13288691549.csv"
ADS_FILE     = r"C:\Users\KEISUKE SHIMOMOTO\Downloads\rioxxrinaxjapan_advertising_sales_report_20260315.csv"
TRANS_FILE   = r"C:\Users\KEISUKE SHIMOMOTO\Downloads\Transaction-Mar-16-2026-01_18_10-0700-13288723972.csv"
OUTPUT       = r"C:\Users\KEISUKE SHIMOMOTO\Desktop\reffort\commerce\ebay\analytics\eBay週次レポート_v2_20260315.xlsx"

# ===== 期間設定 =====
PERIOD_LABEL = "2/13〜3/15（31日間）"
# 週定義: (表示ラベル, 開始日, 終了日, 日数)
WEEKS = [
    ('W1\n2/13-19', date(2026,2,13), date(2026,2,19), 7),
    ('W2\n2/20-26', date(2026,2,20), date(2026,2,26), 7),
    ('W3\n2/27-3/5',date(2026,2,27), date(2026,3,5),  7),
    ('W4\n3/6-15',  date(2026,3,6),  date(2026,3,15), 10),
]
# PLP費用（Transaction Reportに含まれないため手入力）
PLP_FEE_TOTAL = 491.0

TODAY = date.today()

# ===== ユーティリティ =====
def to_f(s):
    try: return float(str(s).replace(',','').replace('%','').replace('$','').strip())
    except: return 0.0

def clean_id(s):
    return s.strip().replace('=','').replace('"','')

def parse_tx_date(s):
    for fmt in ('%b %d, %Y', '%b  %d, %Y'):
        try: return datetime.strptime(s.strip(), fmt).date()
        except: pass
    return None

def days_listed(s):
    try:
        d = datetime.strptime(s.strip(), '%Y-%m-%d').date()
        return (TODAY - d).days
    except: return 0

def pct_change(new, old):
    if old == 0: return '—'
    c = (new - old) / old * 100
    return f'▲{c:.1f}%' if c >= 0 else f'▼{abs(c):.1f}%'

# ===== スタイル定数 =====
FONT = 'メイリオ'
C_GREEN_BG  = 'E8F5E9'  # 売れ筋
C_BLUE_BG   = 'E3F2FD'  # 準売れ筋
C_YELLOW_BG = 'FFFDE7'  # 育成候補
C_ORANGE_BG = 'FFF3E0'  # 要調査
C_RED_BG    = 'FFEBEE'  # 削除候補
C_GRAY_BG   = 'F5F5F5'  # サマリー背景
C_DARK      = '1A237E'  # ヘッダー（紺）
C_GREEN_HDR = '1B5E20'
C_BLUE_HDR  = '0D47A1'
C_YELLOW_HDR= 'E65100'
C_ORANGE_HDR= 'BF360C'
C_RED_HDR   = 'B71C1C'

def hdr_font(color='FFFFFF', size=10, bold=True):
    return Font(name=FONT, size=size, bold=bold, color=color)

def body_font(size=9, bold=False, color='000000'):
    return Font(name=FONT, size=size, bold=bold, color=color)

def hdr_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def body_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def border_thin():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def apply_header_row(ws, row_num, headers, bg_color, txt_color='FFFFFF', row_height=22):
    ws.row_dimensions[row_num].height = row_height
    for col_idx, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=h)
        cell.font = hdr_font(color=txt_color)
        cell.fill = hdr_fill(bg_color)
        cell.alignment = center()
        cell.border = border_thin()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

def apply_body_row(ws, row_num, values, bg_color=None, bold=False, height=20):
    ws.row_dimensions[row_num].height = height
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=v)
        cell.font = body_font(bold=bold)
        if bg_color:
            cell.fill = body_fill(bg_color)
        cell.border = border_thin()
        # 数値は右寄せ、テキストは左寄せ
        if isinstance(v, (int, float)):
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = left()

# ===== Traffic Report 読み込み =====
with open(TRAFFIC_FILE, encoding='utf-8-sig') as f:
    lines = f.readlines()
data = io.StringIO(''.join(lines[5:]))

items = []
for row in csv.DictReader(data):
    title = row.get('Listing title','').strip()
    if not title: continue
    start_str = row.get('Item Start Date','').strip()
    total_imps = to_f(row.get('Total impressions', 0))
    pl_imps    = to_f(row.get('Total Promoted Listings impressions (applies to eBay site only)', 0))
    off_imps   = to_f(row.get('Total Promoted Offsite impressions (applies to off-eBay only)', 0))
    org_imps   = to_f(row.get('Total organic impressions on eBay site', 0))
    items.append({
        'title':     title,
        'id':        clean_id(row.get('eBay item ID','')),
        'start':     start_str,
        'days':      days_listed(start_str),
        'imps':      total_imps,
        'pl_imps':   pl_imps,
        'off_imps':  off_imps,
        'org_imps':  org_imps,
        'pv':        to_f(row.get('Total page views', 0)),
        'sold':      to_f(row.get('Quantity sold', 0)),
        'cvr':       to_f(row.get('Sales conversion rate = Quantity sold/Total page views', 0)),
        'ctr':       to_f(row.get('Click-through rate = Page views from eBay site/Total impressions', 0)),
        'qty':       to_f(row.get('Quantity available', 0)),
        'promo':     row.get('Current promoted listings status','').strip(),
    })

# ===== Advertising Report 読み込み =====
with open(ADS_FILE, encoding='utf-8-sig') as f:
    ads_lines = f.readlines()
ads_data = io.StringIO(''.join(ads_lines[2:]))
ad_by_item = defaultdict(lambda: {'plg_s':0,'plg_fee':0,'plp_s':0,'off_s':0})

for row in csv.DictReader(ads_data):
    iid = row.get('Item ID','').strip()
    ad = row.get('Ad type','')
    st = row.get('Campaign strategy','')
    sale_type = row.get('Sale type','')
    s   = to_f(row.get('Sales','0'))
    fee = to_f(row.get('Ad fees (General)','-'))
    if 'Offsite' in ad:
        ad_by_item[iid]['off_s'] += s
    elif 'Priority' in st:
        ad_by_item[iid]['plp_s'] += s
    elif 'General' in st and 'Halo' not in sale_type:
        ad_by_item[iid]['plg_s'] += s
        ad_by_item[iid]['plg_fee'] += fee

for item in items:
    ad = ad_by_item.get(item['id'], {})
    item['plg_s']    = ad.get('plg_s', 0)
    item['plg_fee']  = ad.get('plg_fee', 0)
    item['plp_s']    = ad.get('plp_s', 0)
    item['off_s']    = ad.get('off_s', 0)
    item['ad_sales'] = item['plg_s'] + item['plp_s'] + item['off_s']

# ===== Transaction Report 読み込み（週別集計）=====
with open(TRANS_FILE, encoding='utf-8-sig') as f:
    trans_lines = f.readlines()
trans_data = io.StringIO(''.join(trans_lines[11:]))

week_stats = []
for wlabel, wstart, wend, wdays in WEEKS:
    week_stats.append({
        'label': wlabel, 'start': wstart, 'end': wend, 'days': wdays,
        'orders':0, 'gross':0.0, 'fvf':0.0, 'intl':0.0,
        'plg_fee':0.0, 'off_fee':0.0, 'qty':0,
    })

total_orders = 0
total_gross = total_fvf = total_intl = total_plg_fee_tx = total_off_fee = 0.0
total_qty = 0

for row in csv.DictReader(trans_data):
    cur = row.get('Transaction currency','').strip()
    if cur != 'USD': continue
    t     = row.get('Type','').strip()
    d_str = row.get('Transaction creation date','').strip()
    d     = parse_tx_date(d_str)
    if not d: continue
    desc  = row.get('Description','').strip()

    if t == 'Order':
        gross = to_f(row.get('Gross transaction amount', 0))
        fvf   = to_f(row.get('Final Value Fee - fixed', 0)) + to_f(row.get('Final Value Fee - variable', 0))
        intl  = to_f(row.get('International fee', 0))
        qty   = int(to_f(row.get('Quantity', 0)))
        total_orders += 1; total_gross += gross
        total_fvf += fvf; total_intl += intl; total_qty += qty
        for ws in week_stats:
            if ws['start'] <= d <= ws['end']:
                ws['orders'] += 1; ws['gross'] += gross
                ws['fvf'] += fvf; ws['intl'] += intl; ws['qty'] += qty
                break
    elif t == 'Other fee' and 'Promoted Listings - General fee' in desc:
        fee = abs(to_f(row.get('Net amount', 0)))
        total_plg_fee_tx += fee
        for ws in week_stats:
            if ws['start'] <= d <= ws['end']:
                ws['plg_fee'] += fee; break
    elif t == 'Other fee' and 'Promoted Offsite fee' in desc:
        fee = abs(to_f(row.get('Net amount', 0)))
        total_off_fee += fee
        for ws in week_stats:
            if ws['start'] <= d <= ws['end']:
                ws['off_fee'] += fee; break

# 週別手取り
for ws in week_stats:
    ws['net'] = ws['gross'] - ws['fvf'] - ws['intl'] - ws['plg_fee'] - ws['off_fee']

# ===== Traffic集計 =====
total_items   = len(items)
total_imps_t  = sum(i['imps'] for i in items)
total_pv_t    = sum(i['pv'] for i in items)
total_sold_t  = sum(i['sold'] for i in items)
total_pl_imps = sum(i['pl_imps'] for i in items)
total_off_imps= sum(i['off_imps'] for i in items)
total_org_imps= sum(i['org_imps'] for i in items)
zero_sold_cnt = sum(1 for i in items if i['sold'] == 0)
overall_ctr   = total_pv_t / total_imps_t * 100 if total_imps_t > 0 else 0
overall_cvr   = total_sold_t / total_pv_t * 100 if total_pv_t > 0 else 0

# ===== 商品分類 =====
top15      = sorted(items, key=lambda x: -x['sold'])[:15]
top15_ids  = {i['id'] for i in top15}

# ⭐ 準売れ筋: sold>0, not top15（伸びしろ候補）
準売れ筋   = sorted([i for i in items if i['sold']>0 and i['id'] not in top15_ids],
                   key=lambda x: (-x['sold'], -x['cvr']))

# 🌱 育成候補: sold=0, pv>=20（需要あり、購入障壁あり）
育成       = sorted([i for i in items if i['sold']==0 and i['pv']>=20], key=lambda x: -x['pv'])[:30]

# ⚠️ 要調査: sold=0, imps>=500, days>=90（在庫ツールバグ or 仕入先URL切れ疑い）
# 全件数を記録しTOP50のみ表示（全件は在庫ツール問題の深刻さを示す指標として報告）
要調査_all  = [i for i in items if i['sold']==0 and i['imps']>=500 and i['days']>=90]
要調査     = sorted(要調査_all, key=lambda x: -x['imps'])[:50]

# 🗑 削除候補L1: 完全死蔵（即削除）
削除L1     = [i for i in items if i['imps']==0 and i['pv']==0 and i['sold']==0]

# 🗑 削除候補L2: 高齢・ほぼ非表示・売上ゼロ（要確認削除）
#   → 掲載180日以上 & インプ50未満 & PV5未満 & 売上ゼロ（かつL1外）
削除L1_ids = {i['id'] for i in 削除L1}
削除L2     = sorted(
    [i for i in items if i['id'] not in 削除L1_ids
     and i['imps'] < 50 and i['pv'] < 5 and i['sold'] == 0 and i['days'] >= 180],
    key=lambda x: -x['days']
)

# ===== Excel ワークブック作成 =====
wb = Workbook()

# ─────────────────────────────────────────────────────────────
# Sheet 1: 📊 サマリー
# ─────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = '📊 サマリー'
ws1.sheet_view.showGridLines = False
# 列幅：A=指標名, B=W1, C=W2, D=W3, E=W4, F=合計, G=W3→W4変化
ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 15
ws1.column_dimensions['F'].width = 15
ws1.column_dimensions['G'].width = 13

def s1_section(ws, row_num, title, color='E8EAF6'):
    ws.merge_cells(f'A{row_num}:G{row_num}')
    c = ws.cell(row=row_num, column=1, value=title)
    c.font = Font(name=FONT, size=10, bold=True, color=C_DARK)
    c.fill = body_fill(color)
    c.alignment = left()
    ws.row_dimensions[row_num].height = 20
    return row_num + 1

row = 1

# ─ タイトル ─
ws1.merge_cells(f'A{row}:G{row}')
c = ws1.cell(row=row, column=1,
    value=f'📊 eBay 週次レポート  {PERIOD_LABEL}  （生成日: {TODAY.strftime("%Y/%m/%d")}）')
c.font = Font(name=FONT, size=13, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_DARK)
c.alignment = center()
ws1.row_dimensions[row].height = 28
row += 1

# ─ KPIブロック ─
row = s1_section(ws1, row, '■ パフォーマンス指標（期間合計）')
apply_header_row(ws1, row,
    [('指標', 30), ('数値（期間合計）', 15), ('内訳・補足', 15),
     ('', 15), ('', 15), ('', 15), ('', 13)], C_DARK)
row += 1

kpi_rows = [
    ('🔵 Total Impressions（表示回数）', f'{total_imps_t:,.0f}',  '期間合計（週別分割不可）', False),
    ('  ├ Promoted Listings',           f'{total_pl_imps:,.0f}', f'{total_pl_imps/total_imps_t*100:.1f}% of total', True),
    ('  ├ Promoted Offsite',            f'{total_off_imps:,.0f}',f'{total_off_imps/total_imps_t*100:.1f}% of total', True),
    ('  └ Organic',                     f'{total_org_imps:,.0f}',f'{total_org_imps/total_imps_t*100:.1f}% of total', True),
    ('👁 Listing Views（閲覧数）',      f'{total_pv_t:,.0f}',    '期間合計', False),
    ('🛒 Quantity Sold（販売数）',      f'{total_sold_t:.0f} 件','期間合計', False),
    ('📈 Click-through Rate（CTR）',    f'{overall_ctr:.3f}%',   '期間平均（PV÷Impressions）', False),
    ('💰 Sales Conversion Rate（CVR）', f'{overall_cvr:.3f}%',   '期間平均（Sold÷PV）', False),
    ('📦 総出品数',                     f'{total_items:,} 件',   f'売上ゼロ: {zero_sold_cnt:,}件 ({zero_sold_cnt/total_items*100:.1f}%)', False),
]
for label, val, note, is_sub in kpi_rows:
    bg = C_GRAY_BG if is_sub else None
    apply_body_row(ws1, row, [label, val, note, '', '', '', ''], bg_color=bg)
    row += 1

row += 1

# ─ 週別売上トレンド ─
row = s1_section(ws1, row, '■ 週別売上トレンド（Transaction Report USD取引のみ）')
apply_header_row(ws1, row,
    [('指標', 30),
     ('W1\n2/13-19\n(7日)', 15),
     ('W2\n2/20-26\n(7日)', 15),
     ('W3\n2/27-3/5\n(7日)', 15),
     ('W4\n3/6-15\n(10日)', 15),
     ('合計', 15),
     ('W3→W4\n変化', 13)],
    C_DARK, row_height=42)
row += 1

net_after_plp = total_gross - total_fvf - total_intl - total_plg_fee_tx - total_off_fee - PLP_FEE_TOTAL

# (ラベル, W1, W2, W3, W4, 合計テキスト, W3→W4変化キー or None, 強調, 費用フラグ)
w_data = [
    ('注文数',
     [f'{ws["orders"]}件' for ws in week_stats],
     f'{total_orders}件',
     pct_change(week_stats[3]['orders'], week_stats[2]['orders']),
     True, False),
    ('販売数（個）',
     [f'{ws["qty"]}個' for ws in week_stats],
     f'{total_qty}個',
     pct_change(week_stats[3]['qty'], week_stats[2]['qty']),
     False, False),
    ('売上 Gross（USD）',
     [f'${ws["gross"]:,.0f}' for ws in week_stats],
     f'${total_gross:,.0f}',
     pct_change(week_stats[3]['gross'], week_stats[2]['gross']),
     True, False),
    ('FVF（最終価値手数料）',
     [f'-${ws["fvf"]:,.0f}' for ws in week_stats],
     f'-${total_fvf:,.0f}',
     pct_change(week_stats[3]['fvf'], week_stats[2]['fvf']),
     False, True),
    ('International fee',
     [f'-${ws["intl"]:,.0f}' for ws in week_stats],
     f'-${total_intl:,.0f}',
     pct_change(week_stats[3]['intl'], week_stats[2]['intl']),
     False, True),
    ('PLG広告費',
     [f'-${ws["plg_fee"]:,.0f}' for ws in week_stats],
     f'-${total_plg_fee_tx:,.0f}',
     pct_change(week_stats[3]['plg_fee'], week_stats[2]['plg_fee']),
     False, True),
    ('Offsite広告費',
     [f'-${ws["off_fee"]:,.0f}' for ws in week_stats],
     f'-${total_off_fee:,.0f}',
     pct_change(week_stats[3]['off_fee'], week_stats[2]['off_fee']),
     False, True),
    ('手取り推定',
     [f'${ws["net"]:,.0f}' for ws in week_stats],
     f'${net_after_plp:,.0f}  ※PLP-${PLP_FEE_TOTAL:.0f}控除後',
     pct_change(week_stats[3]['net'], week_stats[2]['net']),
     True, False),
]

for label, week_vals, total_str, change_str, bold, is_cost in w_data:
    if is_cost:
        bg = 'FFF3E0'
    elif bold:
        bg = 'E3F2FD'
    else:
        bg = None
    apply_body_row(ws1, row,
        [label] + week_vals + [total_str, change_str],
        bg_color=bg, bold=bold)
    row += 1

row += 1

# ─ 収支サマリー ─
row = s1_section(ws1, row, '■ 収支サマリー（期間全体）')
apply_header_row(ws1, row,
    [('項目', 30), ('金額（USD）', 15), ('備考・実質率', 25),
     ('', 15), ('', 15), ('', 15), ('', 13)], C_DARK)
row += 1

finance_rows = [
    ('売上合計（Gross）',         f'${total_gross:,.2f}',         f'USD注文 {total_orders}件',                            True),
    ('FVF（最終価値手数料）',     f'-${total_fvf:,.2f}',          f'実質率: {total_fvf/total_gross*100:.1f}%',            False),
    ('International fee',         f'-${total_intl:,.2f}',          '',                                                     False),
    ('PLG広告費（Promoted General）', f'-${total_plg_fee_tx:,.2f}',
     f'ROAS: {total_gross/total_plg_fee_tx:.1f}x' if total_plg_fee_tx > 0 else '',                                        False),
    ('Offsite広告費',              f'-${total_off_fee:,.2f}',      'Transaction Report記載（CPC課金）',                    False),
    ('PLP広告費（手入力・推定）',  f'-${PLP_FEE_TOTAL:,.2f}',     '⚠ Transaction Reportに含まれないため手入力',          False),
    ('eBay控除後 手取り（推定）',  f'${net_after_plp:,.2f}',      f'収益率: {net_after_plp/total_gross*100:.1f}%',        True),
]
for label, amount, note, bold in finance_rows:
    if '手取り' in label:
        bg = 'E8F5E9'
    elif amount.startswith('-'):
        bg = 'FFF3E0'
    else:
        bg = None
    apply_body_row(ws1, row, [label, amount, note, '', '', '', ''], bg_color=bg, bold=bold)
    row += 1

row += 1

# ─ アクション欄 ─
row = s1_section(ws1, row, '■ 今週のアクション（スタッフ記入欄）', color='FFF9C4')
apply_header_row(ws1, row,
    [('カテゴリ', 30), ('今週やること（記入）', 40), ('担当', 12),
     ('', 15), ('', 15), ('', 15), ('', 13)], '455A64')
row += 1
for action, default_text in [
    ('🔥 売れ筋強化',     '在庫確認・補充予定'),
    ('⭐ 準売れ筋チェック','価格・タイトル改善対象を選定'),
    ('🌱 育成候補対応',   '購入障壁を調査（価格・サイズ・競合）'),
    ('⚠️ 要調査商品対応', '在庫ツール・仕入先URL確認'),
    ('🗑 削除実施',       '即削除リストを確認・取り下げ実行'),
]:
    apply_body_row(ws1, row, [action, default_text, '', '', '', '', ''], bg_color='FAFAFA')
    row += 1

# ─────────────────────────────────────────────────────────────
# Sheet 2: 🔥 コア売れ筋
# ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('🔥 コア売れ筋TOP15')
ws2.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHIJ', [4, 52, 14, 8, 8, 9, 11, 8, 11, 8]):
    ws2.column_dimensions[col].width = w

ws2.merge_cells('A1:J1')
c = ws2['A1']
c.value = '🔥 コア売れ筋 TOP15（販売数ランキング）― 在庫確保・広告維持・集中強化'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_GREEN_HDR)
c.alignment = center()
ws2.row_dimensions[1].height = 24

hdr2 = [
    ('順位', 4), ('商品タイトル', 52), ('Item ID', 14),
    ('販売数', 8), ('CVR', 8), ('PV', 9),
    ('インプ', 11), ('在庫数', 8), ('広告売上', 11), ('掲載日数', 8),
]
apply_header_row(ws2, 2, hdr2, C_GREEN_HDR)

for rank, item in enumerate(top15, 1):
    ad_s = f'${item["ad_sales"]:,.0f}' if item['ad_sales'] > 0 else '-'
    bg = 'C8E6C9' if rank <= 3 else C_GREEN_BG
    apply_body_row(ws2, rank+2, [
        rank, item['title'], item['id'],
        int(item['sold']), f'{item["cvr"]:.1f}%', int(item['pv']),
        f'{item["imps"]:,.0f}', int(item['qty']),
        ad_s, f'{item["days"]}日',
    ], bg_color=bg, height=20)

# ─────────────────────────────────────────────────────────────
# Sheet 3: ⭐ 準売れ筋
# ─────────────────────────────────────────────────────────────
ws3 = wb.create_sheet('⭐ 準売れ筋')
ws3.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHI', [52, 14, 8, 8, 9, 11, 8, 11, 8]):
    ws3.column_dimensions[col].width = w

ws3.merge_cells('A1:I1')
c = ws3['A1']
c.value = f'⭐ 準売れ筋（{len(準売れ筋)}件）― 売上あり・TOP15外 ― 価格微調整でコア売れ筋に育つ可能性あり'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_BLUE_HDR)
c.alignment = center()
ws3.row_dimensions[1].height = 24

hdr3 = [
    ('商品タイトル', 52), ('Item ID', 14),
    ('販売数', 8), ('CVR', 8), ('PV', 9),
    ('インプ', 11), ('在庫数', 8), ('Organic比', 11), ('掲載日数', 8),
]
apply_header_row(ws3, 2, hdr3, C_BLUE_HDR)

for i, item in enumerate(準売れ筋, 3):
    org_ratio = f'{item["org_imps"]/item["imps"]*100:.0f}%' if item['imps'] > 0 else '-'
    apply_body_row(ws3, i, [
        item['title'], item['id'],
        int(item['sold']), f'{item["cvr"]:.1f}%', int(item['pv']),
        f'{item["imps"]:,.0f}', int(item['qty']),
        org_ratio, f'{item["days"]}日',
    ], bg_color=C_BLUE_BG)

# ─────────────────────────────────────────────────────────────
# Sheet 4: 🌱 育成候補
# ─────────────────────────────────────────────────────────────
ws4 = wb.create_sheet('🌱 育成候補')
ws4.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHI', [52, 14, 9, 11, 8, 8, 8, 13, 11]):
    ws4.column_dimensions[col].width = w

ws4.merge_cells('A1:I1')
c = ws4['A1']
c.value = f'🌱 育成候補（{len(育成)}件）― PV20以上・売上ゼロ ― 需要あり・何かが購入を妨げている'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_YELLOW_HDR)
c.alignment = center()
ws4.row_dimensions[1].height = 24

ws4.merge_cells('A2:I2')
note4 = ws4['A2']
note4.value = '→ 確認項目：価格（競合比較）/ サイズ・カラー在庫 / タイトル・写真の品質 / 送料設定'
note4.font = Font(name=FONT, size=9, color='E65100')
note4.fill = body_fill('FFF9C4')
note4.alignment = left()
ws4.row_dimensions[2].height = 18

hdr4 = [
    ('商品タイトル', 52), ('Item ID', 14),
    ('PV', 9), ('インプ', 11), ('CTR', 8),
    ('在庫数', 8), ('掲載日数', 8), ('広告状態', 13), ('Organic比', 11),
]
apply_header_row(ws4, 3, hdr4, C_YELLOW_HDR, txt_color='FFFFFF')

for i, item in enumerate(育成, 4):
    org_ratio = f'{item["org_imps"]/item["imps"]*100:.0f}%' if item['imps'] > 0 else '-'
    apply_body_row(ws4, i, [
        item['title'], item['id'],
        int(item['pv']), f'{item["imps"]:,.0f}', f'{item["ctr"]:.2f}%',
        int(item['qty']), f'{item["days"]}日', item['promo'], org_ratio,
    ], bg_color=C_YELLOW_BG)

# ─────────────────────────────────────────────────────────────
# Sheet 5: ⚠️ 要調査
# ─────────────────────────────────────────────────────────────
ws5 = wb.create_sheet('⚠️ 要調査')
ws5.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHI', [52, 14, 11, 9, 8, 8, 13, 11, 18]):
    ws5.column_dimensions[col].width = w

ws5.merge_cells('A1:I1')
c = ws5['A1']
c.value = f'⚠️ 要調査 TOP50（全{len(要調査_all)}件中）― インプ500以上・売上ゼロ・掲載90日以上 ― 在庫ツールバグ or 仕入先URL切れ疑い'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_ORANGE_HDR)
c.alignment = center()
ws5.row_dimensions[1].height = 24

ws5.merge_cells('A2:I2')
note5 = ws5['A2']
note5.value = (f'⚠ 全{len(要調査_all)}件/{total_items}件({len(要調査_all)/total_items*100:.0f}%)がインプあり売上ゼロ'
               f' — 在庫ツールのバグが広範囲に影響している可能性あり'
               f' | 確認①在庫ツールURLが有効か ②仕入先に在庫があるか ③他仕入先に切替可能か')
note5.font = Font(name=FONT, size=9, color='BF360C')
note5.fill = body_fill('FFF3E0')
note5.alignment = left()
ws5.row_dimensions[2].height = 20

hdr5 = [
    ('商品タイトル', 52), ('Item ID', 14),
    ('インプ', 11), ('PV', 9), ('在庫数', 8),
    ('掲載日数', 8), ('広告状態', 13), ('Organic比', 11), ('確認状況（記入）', 18),
]
apply_header_row(ws5, 3, hdr5, C_ORANGE_HDR)

for i, item in enumerate(要調査, 4):
    org_ratio = f'{item["org_imps"]/item["imps"]*100:.0f}%' if item['imps'] > 0 else '-'
    apply_body_row(ws5, i, [
        item['title'], item['id'],
        f'{item["imps"]:,.0f}', int(item['pv']), int(item['qty']),
        f'{item["days"]}日', item['promo'], org_ratio, '',
    ], bg_color=C_ORANGE_BG)

# ─────────────────────────────────────────────────────────────
# Sheet 6: 🗑 削除候補
# ─────────────────────────────────────────────────────────────
ws6 = wb.create_sheet('🗑 削除候補')
ws6.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGH', [52, 14, 8, 8, 13, 14, 14, 8]):
    ws6.column_dimensions[col].width = w

ws6.merge_cells('A1:H1')
c = ws6['A1']
c.value = f'🗑 削除候補 — L1即削除: {len(削除L1)}件 ／ L2要確認削除: {len(削除L2)}件'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_RED_HDR)
c.alignment = center()
ws6.row_dimensions[1].height = 24

ws6.merge_cells('A2:H2')
note6 = ws6['A2']
note6.value = '⚠️ 削除前に必ずeBayで確認：① ウォッチ数（関心度）② 生涯販売数（過去実績） → 両方ゼロに近い場合のみ削除推奨'
note6.font = Font(name=FONT, size=9, color='B71C1C', bold=True)
note6.fill = body_fill('FFEBEE')
note6.alignment = left()
ws6.row_dimensions[2].height = 20

# L1ヘッダー
apply_header_row(ws6, 3,
    [('【L1：即削除】インプ・PV・売上すべてゼロ（{0}件）'.format(len(削除L1)), 52),
     ('Item ID', 14), ('在庫数', 8), ('掲載日数', 8), ('広告状態', 13),
     ('ウォッチ数\n(eBayで確認)', 14), ('生涯販売数\n(eBayで確認)', 14), ('削除済✓', 8)],
    C_RED_HDR)

r = 4
for item in 削除L1:
    apply_body_row(ws6, r, [
        item['title'], item['id'],
        int(item['qty']), f'{item["days"]}日', item['promo'],
        '←eBayで確認', '←eBayで確認', '',
    ], bg_color='FFCDD2')
    r += 1

r += 1

# L2ヘッダー
ws6.merge_cells(f'A{r}:H{r}')
c2 = ws6.cell(row=r, column=1, value=f'【L2：要確認削除】掲載180日以上・インプ50未満・売上ゼロ（{len(削除L2)}件）― 削除前に上記の確認必須')
c2.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
c2.fill = hdr_fill(C_RED_HDR)
c2.alignment = left()
ws6.row_dimensions[r].height = 20
r += 1

apply_header_row(ws6, r,
    [('商品タイトル', 52), ('Item ID', 14), ('インプ', 11), ('PV', 8),
     ('在庫数', 8), ('掲載日数', 8), ('ウォッチ数\n(eBayで確認)', 14), ('削除済✓', 8)],
    '7B1FA2')
r += 1

for item in 削除L2:
    apply_body_row(ws6, r, [
        item['title'], item['id'],
        f'{item["imps"]:,.0f}', int(item['pv']),
        int(item['qty']), f'{item["days"]}日', '←eBayで確認', '',
    ], bg_color='F3E5F5')
    r += 1

# ─────────────────────────────────────────────────────────────
# Sheet 7: 📋 履歴（マスター蓄積用・将来比較用）
# ─────────────────────────────────────────────────────────────
ws7 = wb.create_sheet('📋 週次履歴（蓄積）')
ws7.sheet_view.showGridLines = False
for col, w in zip('ABCDEFG', [14, 22, 8, 13, 13, 8, 8]):
    ws7.column_dimensions[col].width = w

ws7.merge_cells('A1:G1')
c = ws7['A1']
c.value = '📋 週次履歴 ― 毎週レポート実行後にサマリー数値を追記していく（前週比較の基礎データ）'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_DARK)
c.alignment = center()
ws7.row_dimensions[1].height = 24

ws7.merge_cells('A2:G2')
note7 = ws7['A2']
note7.value = '→ 毎週月曜日にこのシートに1行追記すること。来週から前週比が自動的に計算できるようになる。'
note7.font = Font(name=FONT, size=9, color=C_DARK)
note7.fill = body_fill('E8EAF6')
note7.alignment = left()
ws7.row_dimensions[2].height = 18

hdr7 = [
    ('レポート日', 14), ('期間', 22), ('注文数', 8),
    ('売上Gross', 13), ('手取り推定', 13), ('CVR', 8), ('CTR', 8),
]
apply_header_row(ws7, 3, hdr7, C_DARK)

apply_body_row(ws7, 4, [
    TODAY.strftime('%Y/%m/%d'),
    PERIOD_LABEL,
    total_orders,
    f'${total_gross:,.0f}',
    f'${net_after_plp:,.0f}',
    f'{overall_cvr:.3f}%',
    f'{overall_ctr:.3f}%',
], bg_color='E3F2FD', bold=True)

for empty_row in range(5, 20):
    apply_body_row(ws7, empty_row,
                   ['（来週以降ここに追記）', '', '', '', '', '', ''],
                   bg_color='FAFAFA')
    break

# ===== 保存 =====
wb.save(OUTPUT)
print(f'✅ レポート生成完了: {OUTPUT}')
print(f'   シート構成: サマリー / コア売れ筋TOP15 / 準売れ筋({len(準売れ筋)}件) / 育成候補({len(育成)}件) / 要調査({len(要調査)}件) / 削除候補(L1:{len(削除L1)}+L2:{len(削除L2)}件) / 週次履歴')
