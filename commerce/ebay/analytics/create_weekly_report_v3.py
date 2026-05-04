import csv, io, sys, json, os, re, time
import urllib.request
from collections import defaultdict
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# .envファイルから機密情報を読み込む
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))

sys.stdout.reconfigure(encoding='utf-8')

# ===== ファイルパス（毎週更新） =====
# Traffic Report: 週別に4ファイルをSeller Hubからダウンロード（各週月〜日で期間指定）
# ファイルがない週は空リスト扱い（エラーにならない）
TRAFFIC_FILES = [
    r"",  # W1 Traffic CSV
    r"",  # W2 Traffic CSV
    r"",  # W3 Traffic CSV
    r"",  # W4 Traffic CSV
]
# 全期間まとめて1ファイルの場合はこちら（週別ファイルがない場合のフォールバック）
TRAFFIC_FILE_COMBINED = r"C:\Users\KEISUKE SHIMOMOTO\Downloads\eBay-ListingsTrafficReport-Mar-15-2026-21_56_24-0700-13288691549.csv"
ADS_FILE     = r"C:\Users\KEISUKE SHIMOMOTO\Downloads\rioxxrinaxjapan_advertising_sales_report_20260315.csv"
TRANS_FILE   = r"C:\Users\KEISUKE SHIMOMOTO\Downloads\Transaction-Mar-16-2026-01_18_10-0700-13288723972.csv"

# 出力ファイル名は日付を自動挿入（手動更新不要）
# 相対パス化：このスクリプトと同じフォルダ(commerce/ebay/analytics/)を基準にする
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TODAY        = date.today()
OUTPUT       = os.path.join(_SCRIPT_DIR, f"eBay週次レポート_v3_{TODAY.strftime('%Y%m%d')}_{datetime.now().strftime('%H%M')}.xlsx")

# 週次履歴ファイル（前週比較のためのJSONデータ）
HISTORY_FILE = os.path.join(_SCRIPT_DIR, "weekly_history.json")

# ===== eBay API設定（削除候補のウォッチ数・生涯販売数を自動取得）=====
# FalseにするとAPIをスキップ（CSVファイルがない時や手動確認したい時）
FETCH_EBAY_API   = True
EBAY_APP_ID      = os.getenv('EBAY_APP_ID')
EBAY_DEV_ID      = os.getenv('EBAY_DEV_ID')
EBAY_CERT_ID     = os.getenv('EBAY_CERT_ID')
EBAY_USER_TOKEN  = os.getenv('EBAY_USER_TOKEN')
EBAY_SELLER_CACHE = os.path.join(_SCRIPT_DIR, "ebay_seller_cache.json")
EBAY_CACHE_DAYS   = 7     # キャッシュ有効日数（7日以内なら再取得しない）
EBAY_API_DELAY    = 1.0   # バルクAPIのページ間待機秒数

# ===== 期間設定（今日を基準に直近4週を自動計算・手動設定不要）=====
def _calc_weeks():
    """直近4つの完了週（月〜日）を自動生成。月曜実行→前週がW4
    例: 3/24(月)実行 → W4=3/17-3/23, W3=3/10-3/16, W2=3/3-3/9, W1=2/24-3/2"""
    # 直近の日曜日を探す（月曜なら昨日、火曜なら2日前…日曜なら今日）
    days_since_sunday = (TODAY.weekday() + 1) % 7
    last_sunday = TODAY - timedelta(days=days_since_sunday) if days_since_sunday else TODAY
    weeks = []
    for i in range(3, -1, -1):
        w_end   = last_sunday - timedelta(weeks=i)
        w_start = w_end - timedelta(days=6)
        n       = 4 - i
        label   = f'W{n}\n{w_start.month}/{w_start.day}-{w_end.month}/{w_end.day}'
        weeks.append((label, w_start, w_end, 7))
    return weeks

WEEKS        = _calc_weeks()
PERIOD_LABEL = (f'{WEEKS[0][1].month}/{WEEKS[0][1].day}〜'
                f'{WEEKS[-1][2].month}/{WEEKS[-1][2].day}'
                f'（{sum(w[3] for w in WEEKS)}日間）')
# PLP費用（Transaction Reportに含まれないため手入力）
PLP_FEE_TOTAL = 491.0

# ===== トランザクション取得方法 =====
# True = eBay GetOrders APIから自動取得（CSV不要・最新データ反映）
# False = Transaction Report CSV（従来方式・手動ダウンロード必要）
USE_TRANSACTION_API = True

# ===== Traffic Report 取得方法 =====
# True = eBay Sell Analytics API (OAuth 2.0) で自動取得（CSV不要）
# False = Traffic Report CSV（従来方式・手動ダウンロード必要）
USE_TRAFFIC_API = True

# 手数料推定比率（APIで取得できないfeeの推定用・過去データ基準）
# International fee ≈ 総売上の0.8%, PLG ≈ 5.3%, Offsite ≈ 3.7%
EST_INTL_RATE    = 0.008   # International fee / gross
EST_PLG_RATE     = 0.053   # PLG広告費 / gross
EST_OFFSITE_RATE = 0.037   # Offsite広告費 / gross

# 為替レート（USD換算用・概算値。Transaction Reportの実レートで更新推奨）
USD_RATES = {
    'USD': 1.0,
    'GBP': 1.27,   # 英ポンド
    'EUR': 1.08,   # ユーロ
    'AUD': 0.65,   # 豪ドル
    'CAD': 0.73,   # カナダドル
}

# 通貨→サイト名マッピング（ebaymagの出品先は通貨で判別）
# USD = USサイト上の購入（Switzerland/Spain等の海外バイヤー含む）
CURRENCY_TO_SITE = {
    'USD': 'US',
    'GBP': 'UK',
    'AUD': 'Australia',
    'CAD': 'Canada',
    'EUR': 'EU (DE/IT)',
}
# サイト別表示順
SITE_ORDER = ['US', 'UK', 'EU (DE/IT)', 'Australia', 'Canada']

def to_usd(amount, currency):
    """金額をUSDに換算する"""
    rate = USD_RATES.get(currency, 1.0)
    return amount * rate

def currency_site(currency):
    """通貨からサイト名を返す"""
    return CURRENCY_TO_SITE.get(currency, 'US')

# 育成候補の閾値（出品数が減るにつれて下げる。AIが自動提案する）
IKUSEI_PV_THRESHOLD = 50   # PV閾値（現在50）
IKUSEI_MAX          = 20   # 最大表示件数

# ===== eBay API: GetMyeBaySelling で全商品のSKU・ウォッチ数・販売数・価格を一括取得 =====
def load_seller_cache():
    """セラーキャッシュを読み込む（なければ空を返す）"""
    if os.path.exists(EBAY_SELLER_CACHE):
        with open(EBAY_SELLER_CACHE, encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_seller_cache(cache):
    """セラーキャッシュをJSONに保存"""
    with open(EBAY_SELLER_CACHE, 'w', encoding='utf-8') as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

def fetch_all_seller_data():
    """GetMyeBaySelling で全出品商品を一括取得（200件/ページ・全ページ）
    戻り値: {item_id: {sku, watch, lifetime_sold, price, fetched}} の辞書"""
    url    = 'https://api.ebay.com/ws/api.dll'
    result = {}
    page   = 1
    total_pages = 1
    while page <= total_pages:
        xml_body = f'''<?xml version="1.0" encoding="utf-8"?>
<GetMyeBaySelling xmlns="urn:ebay:apis:eBLBaseComponents">
  <RequesterCredentials><eBayAuthToken>{EBAY_USER_TOKEN}</eBayAuthToken></RequesterCredentials>
  <ActiveList>
    <Include>true</Include>
    <IncludeWatchCount>true</IncludeWatchCount>
    <Pagination><EntriesPerPage>200</EntriesPerPage><PageNumber>{page}</PageNumber></Pagination>
  </ActiveList>
  <DetailLevel>ReturnSummary</DetailLevel>
</GetMyeBaySelling>'''
        req = urllib.request.Request(url, data=xml_body.encode('utf-8'), method='POST')
        req.add_header('X-EBAY-API-COMPATIBILITY-LEVEL', '967')
        req.add_header('X-EBAY-API-DEV-NAME',  EBAY_DEV_ID)
        req.add_header('X-EBAY-API-APP-NAME',  EBAY_APP_ID)
        req.add_header('X-EBAY-API-CERT-NAME', EBAY_CERT_ID)
        req.add_header('X-EBAY-API-CALL-NAME', 'GetMyeBaySelling')
        req.add_header('X-EBAY-API-SITEID',    '0')
        req.add_header('Content-Type',         'text/xml')
        try:
            with urllib.request.urlopen(req, timeout=30) as res:
                body = res.read().decode('utf-8')
            tp = re.search(r'<TotalNumberOfPages>(\d+)</TotalNumberOfPages>', body)
            if tp:
                total_pages = int(tp.group(1))
            blocks = re.findall(r'<Item>(.*?)</Item>', body, re.DOTALL)
            for block in blocks:
                iid_m = re.search(r'<ItemID>(\d+)</ItemID>', block)
                if not iid_m:
                    continue
                # SKUは<SKU>または<CustomLabel>タグに入る
                sku_m = re.search(r'<SKU>(.*?)</SKU>', block)
                if not sku_m:
                    sku_m = re.search(r'<CustomLabel>(.*?)</CustomLabel>', block)
                title_m = re.search(r'<Title>(.*?)</Title>', block)
                watch_m = re.search(r'<WatchCount>(\d+)</WatchCount>',       block)
                sold_m  = re.search(r'<QuantitySold>(\d+)</QuantitySold>',   block)
                price_m = re.search(r'<CurrentPrice[^>]*>([\d.]+)</CurrentPrice>', block)
                qty_m   = re.search(r'<QuantityAvailable>(\d+)</QuantityAvailable>', block)
                start_m = re.search(r'<StartTime>(.*?)</StartTime>', block)
                # 出品開始日をYYYY-MM-DD形式に変換
                start_date = ''
                if start_m:
                    try:
                        start_date = start_m.group(1)[:10]  # "2025-06-15T..." → "2025-06-15"
                    except:
                        pass
                result[iid_m.group(1)] = {
                    'sku':           sku_m.group(1).strip() if sku_m   else '',
                    'title':         title_m.group(1).strip() if title_m else '',
                    'watch':         int(watch_m.group(1))  if watch_m else 0,
                    'lifetime_sold': int(sold_m.group(1))   if sold_m  else 0,
                    'price':         float(price_m.group(1)) if price_m else 0.0,
                    'qty':           int(qty_m.group(1))    if qty_m   else 0,
                    'start_date':    start_date,
                    'fetched':       TODAY.strftime('%Y-%m-%d'),
                }
            print(f'  📄 ページ {page}/{total_pages}（{len(blocks)}件 / 累計{len(result)}件）')
        except Exception as e:
            print(f'  ⚠️ ページ{page}エラー: {str(e)[:100]}')
        page += 1
        if page <= total_pages:
            time.sleep(EBAY_API_DELAY)   # 過剰なAPIコールを防ぐ
    return result

def get_item_api_data(item_id):
    """seller_cacheから1件のデータを返す。
    戻り値: (sku, watch, lifetime_sold, price, judge) のタプル
      judge: ✅ 削除OK / ⚠️ 要確認 / 🚫 削除NG"""
    d            = seller_cache.get(item_id, {})
    sku          = d.get('sku',           '')
    watch        = d.get('watch',         None)
    lifetime_sold= d.get('lifetime_sold', None)
    price        = d.get('price',         None)
    if watch is None:
        return '', '?', '?', '?', '—'   # キャッシュなし
    # 削除判定：ウォッチ数or生涯販売数が多いほど慎重に
    if watch > 50 or lifetime_sold > 3:
        judge = '🚫 削除NG'     # 人気あり・実績あり → 削除NG
    elif watch > 10 or lifetime_sold > 0:
        judge = '⚠️ 要確認'    # 少し関心あり → 確認必須
    else:
        judge = '✅ 削除OK'    # 関心なし実績なし → 削除推奨
    return sku, watch, lifetime_sold, price, judge

# ===== eBay API: GetOrders でトランザクションデータを自動取得 =====
def fetch_orders_api(period_start, period_end):
    """GetOrders APIで指定期間の全注文を取得（ページネーション対応）
    戻り値: [{order_id, date, gross, fvf, qty, currency}, ...] のリスト"""
    url = 'https://api.ebay.com/ws/api.dll'
    # eBay APIはUTC。JSTの00:00はUTC前日15:00だが、日付単位で十分
    from_iso = f'{period_start.isoformat()}T00:00:00.000Z'
    to_iso   = f'{period_end.isoformat()}T23:59:59.000Z'

    all_orders = []
    page = 1
    total_pages = 1

    while page <= total_pages:
        xml_body = f'''<?xml version="1.0" encoding="utf-8"?>
<GetOrdersRequest xmlns="urn:ebay:apis:eBLBaseComponents">
  <RequesterCredentials><eBayAuthToken>{EBAY_USER_TOKEN}</eBayAuthToken></RequesterCredentials>
  <CreateTimeFrom>{from_iso}</CreateTimeFrom>
  <CreateTimeTo>{to_iso}</CreateTimeTo>
  <OrderRole>Seller</OrderRole>
  <OrderStatus>Completed</OrderStatus>
  <IncludeFinalValueFee>true</IncludeFinalValueFee>
  <Pagination>
    <EntriesPerPage>100</EntriesPerPage>
    <PageNumber>{page}</PageNumber>
  </Pagination>
</GetOrdersRequest>'''
        req = urllib.request.Request(url, data=xml_body.encode('utf-8'), method='POST')
        req.add_header('X-EBAY-API-COMPATIBILITY-LEVEL', '967')
        req.add_header('X-EBAY-API-DEV-NAME',  EBAY_DEV_ID)
        req.add_header('X-EBAY-API-APP-NAME',  EBAY_APP_ID)
        req.add_header('X-EBAY-API-CERT-NAME', EBAY_CERT_ID)
        req.add_header('X-EBAY-API-CALL-NAME', 'GetOrders')
        req.add_header('X-EBAY-API-SITEID',    '0')
        req.add_header('Content-Type',          'text/xml')

        try:
            with urllib.request.urlopen(req, timeout=30) as res:
                body = res.read().decode('utf-8')

            # エラーチェック
            ack = re.search(r'<Ack>(.*?)</Ack>', body)
            if ack and ack.group(1) == 'Failure':
                err = re.search(r'<ShortMessage>(.*?)</ShortMessage>', body)
                print(f'  ⚠️ GetOrders エラー: {err.group(1) if err else "不明"}')
                break

            # ページ情報
            tp = re.search(r'<TotalNumberOfPages>(\d+)</TotalNumberOfPages>', body)
            if tp:
                total_pages = int(tp.group(1))

            # 注文パース（OrderArrayの中のOrder要素を取得）
            orders = re.findall(r'<Order>(.*?)</Order>', body, re.DOTALL)
            for order_xml in orders:
                # 通貨取得（AmountPaidのcurrencyID属性）
                cur_m = re.search(r'<AmountPaid currencyID="(\w+)"', order_xml)
                currency = cur_m.group(1) if cur_m else ''

                # サイト取得（TransactionSiteID）
                site_m = re.search(r'<TransactionSiteID>(.*?)</TransactionSiteID>', order_xml)
                site = site_m.group(1) if site_m else 'US'

                # 注文日
                created_m = re.search(r'<CreatedTime>(.*?)</CreatedTime>', order_xml)
                if not created_m:
                    continue
                try:
                    order_date = datetime.strptime(created_m.group(1)[:10], '%Y-%m-%d').date()
                except:
                    continue

                # 売上金額（Total - 通貨を問わず取得）
                total_m = re.search(r'<Total currencyID="\w+">([\d.]+)</Total>', order_xml)
                gross_native = float(total_m.group(1)) if total_m else 0.0

                # Totalが取れない場合はAmountPaidで補完
                if gross_native == 0:
                    paid_m = re.search(r'<AmountPaid currencyID="\w+">([\d.]+)</AmountPaid>', order_xml)
                    gross_native = float(paid_m.group(1)) if paid_m else 0.0

                # USD換算
                gross = to_usd(gross_native, currency)

                # FVF（Transaction内のFinalValueFee合計。通貨を問わず取得）
                fvf_native = 0.0
                txn_blocks = re.findall(r'<Transaction>(.*?)</Transaction>', order_xml, re.DOTALL)
                for txn in txn_blocks:
                    fvf_m = re.search(r'<FinalValueFee[^>]*>([\d.]+)</FinalValueFee>', txn)
                    if fvf_m:
                        fvf_native += float(fvf_m.group(1))
                fvf = to_usd(fvf_native, currency)

                # 数量（TransactionArray内の各TransactionのQuantityPurchased合計）
                qty = 0
                qty_matches = re.findall(r'<QuantityPurchased>(\d+)</QuantityPurchased>', order_xml)
                for q in qty_matches:
                    qty += int(q)

                # 注文に含まれるItemIDを取得（週別売上ゼロSKU追跡用）
                item_ids_in_order = re.findall(r'<ItemID>(\d+)</ItemID>', order_xml)

                all_orders.append({
                    'date':     order_date,
                    'gross':    gross,
                    'gross_native': gross_native,
                    'fvf':      fvf,
                    'qty':      qty,
                    'currency': currency,
                    'site':     site,
                    'item_ids': item_ids_in_order,
                })

            print(f'  📄 GetOrders ページ {page}/{total_pages}（{len(orders)}件 / 累計{len(all_orders)}件USD）')
        except Exception as e:
            print(f'  ⚠️ GetOrders ページ{page}エラー: {str(e)[:100]}')
            break

        page += 1
        if page <= total_pages:
            time.sleep(0.5)  # API負荷対策

    return all_orders

# ===== eBay Sell Analytics API: Traffic Report（OAuth 2.0）=====
def fetch_traffic_api_daily(period_start, period_end):
    """Traffic Report APIで日別サマリーを取得（dimension=DAY）
    戻り値: [{date, imps, pv, sold, ctr, cvr}, ...]"""
    from ebay_oauth import get_access_token
    token = get_access_token()

    date_from = period_start.strftime('%Y%m%d')
    date_to   = period_end.strftime('%Y%m%d')
    filter_str = f'date_range:[{date_from}..{date_to}],marketplace_ids:{{EBAY_US}}'
    # eBay APIはmetricパラメータを複数回指定する必要がある（カンマ区切り不可）
    params = urllib.parse.urlencode([
        ('dimension', 'DAY'),
        ('metric', 'LISTING_IMPRESSION_TOTAL'),
        ('metric', 'LISTING_VIEWS_TOTAL'),
        ('metric', 'TRANSACTION'),
        ('metric', 'CLICK_THROUGH_RATE'),
        ('metric', 'SALES_CONVERSION_RATE'),
        ('filter', filter_str),
    ])
    url = f'https://api.ebay.com/sell/analytics/v1/traffic_report?{params}'
    req = urllib.request.Request(url)
    req.add_header('Authorization', f'Bearer {token}')
    req.add_header('Content-Type', 'application/json')
    req.add_header('X-EBAY-C-MARKETPLACE-ID', 'EBAY_US')

    with urllib.request.urlopen(req, timeout=30) as res:
        data = json.loads(res.read().decode())

    # レスポンスをパース
    # メトリックはリクエスト順に返る: LISTING_IMPRESSION_TOTAL, LISTING_VIEWS_TOTAL, TRANSACTION, CTR, CVR
    METRIC_ORDER = ['LISTING_IMPRESSION_TOTAL', 'LISTING_VIEWS_TOTAL', 'TRANSACTION',
                    'CLICK_THROUGH_RATE', 'SALES_CONVERSION_RATE']
    results = []
    for record in data.get('records', []):
        # dimensionValues[0]はdict: {"value": "20260317", "applicable": true}
        day_str = record['dimensionValues'][0]['value']
        d = datetime.strptime(day_str, '%Y%m%d').date()
        vals = {}
        for mk, mv in zip(METRIC_ORDER, record['metricValues']):
            vals[mk] = float(mv.get('value', 0)) if mv.get('applicable') else 0
        results.append({
            'date': d,
            'imps': vals.get('LISTING_IMPRESSION_TOTAL', 0),
            'pv':   vals.get('LISTING_VIEWS_TOTAL', 0),
            'sold': vals.get('TRANSACTION', 0),
            'ctr':  vals.get('CLICK_THROUGH_RATE', 0),
            'cvr':  vals.get('SALES_CONVERSION_RATE', 0),
        })
    return results


def fetch_traffic_api_listings(period_start, period_end, listing_ids=None, sort_metric=None):
    """Traffic Report APIで商品別データを取得（dimension=LISTING）
    listing_ids: 指定しない場合はTOP200を返す
    sort_metric: ソート基準（例: '-TRANSACTION'で販売数降順）
    戻り値: [{listing_id, title, imps, pv, sold, ctr, cvr}, ...]"""
    from ebay_oauth import get_access_token
    token = get_access_token()

    date_from = period_start.strftime('%Y%m%d')
    date_to   = period_end.strftime('%Y%m%d')
    filter_str = f'date_range:[{date_from}..{date_to}],marketplace_ids:{{EBAY_US}}'
    if listing_ids:
        ids_str = '|'.join(listing_ids)
        filter_str += f',listing_ids:{{{ids_str}}}'

    # eBay APIはmetricパラメータを複数回指定する必要がある
    param_list = [
        ('dimension', 'LISTING'),
        ('metric', 'LISTING_IMPRESSION_TOTAL'),
        ('metric', 'LISTING_VIEWS_TOTAL'),
        ('metric', 'TRANSACTION'),
        ('metric', 'CLICK_THROUGH_RATE'),
        ('metric', 'SALES_CONVERSION_RATE'),
        ('filter', filter_str),
    ]
    if sort_metric:
        param_list.append(('sort', sort_metric))
    url = f'https://api.ebay.com/sell/analytics/v1/traffic_report?{urllib.parse.urlencode(param_list)}'

    req = urllib.request.Request(url)
    req.add_header('Authorization', f'Bearer {token}')
    req.add_header('Content-Type', 'application/json')
    req.add_header('X-EBAY-C-MARKETPLACE-ID', 'EBAY_US')

    with urllib.request.urlopen(req, timeout=30) as res:
        data = json.loads(res.read().decode())

    # 商品タイトル辞書（dimensionMetadataから取得）
    titles = {}
    for dm_group in data.get('dimensionMetadata', []):
        for mr in dm_group.get('metadataRecords', []):
            lid = str(mr.get('value', {}).get('value', ''))
            mvs = mr.get('metadataValues', [])
            if mvs:
                titles[lid] = mvs[0].get('value', '')

    # レスポンスをパース
    METRIC_ORDER = ['LISTING_IMPRESSION_TOTAL', 'LISTING_VIEWS_TOTAL', 'TRANSACTION',
                    'CLICK_THROUGH_RATE', 'SALES_CONVERSION_RATE']
    results = []
    for record in data.get('records', []):
        # dimensionValues[0]はdict: {"value": "listing_id", "applicable": true}
        lid = record['dimensionValues'][0]['value']
        vals = {}
        for mk, mv in zip(METRIC_ORDER, record['metricValues']):
            vals[mk] = float(mv.get('value', 0)) if mv.get('applicable') else 0
        results.append({
            'listing_id': lid,
            'title':      titles.get(lid, ''),
            'imps':       vals.get('LISTING_IMPRESSION_TOTAL', 0),
            'pv':         vals.get('LISTING_VIEWS_TOTAL', 0),
            'sold':       vals.get('TRANSACTION', 0),
            'ctr':        vals.get('CLICK_THROUGH_RATE', 0),
            'cvr':        vals.get('SALES_CONVERSION_RATE', 0),
        })
    return results


def fetch_traffic_all_listings(period_start, period_end, all_item_ids):
    """全商品のTrafficデータをバッチ取得（200件ずつ・429リトライ対応）
    all_item_ids: 全商品IDのリスト
    戻り値: {item_id: {imps, pv, sold, ctr, cvr, title}} の辞書"""
    result = {}
    batch_size = 200
    total_batches = (len(all_item_ids) + batch_size - 1) // batch_size

    for batch_idx in range(total_batches):
        start = batch_idx * batch_size
        batch_ids = all_item_ids[start:start + batch_size]
        # 429エラー時は最大3回リトライ（指数バックオフ）
        for retry in range(3):
            try:
                listings = fetch_traffic_api_listings(period_start, period_end, listing_ids=batch_ids)
                for item in listings:
                    result[item['listing_id']] = item
                break  # 成功したらリトライループを抜ける
            except urllib.error.HTTPError as e:
                if e.code == 429 and retry < 2:
                    wait = (retry + 1) * 5  # 5秒, 10秒
                    if retry == 0:
                        print(f'  ⏳ Rate limit到達。{wait}秒待機...')
                    time.sleep(wait)
                else:
                    print(f'  ⚠️ Traffic APIバッチ{batch_idx+1}エラー: {str(e)[:100]}')
                    break
            except Exception as e:
                print(f'  ⚠️ Traffic APIバッチ{batch_idx+1}エラー: {str(e)[:100]}')
                break
        if (batch_idx + 1) % 10 == 0 or batch_idx == total_batches - 1:
            print(f'  📄 Traffic API バッチ {batch_idx+1}/{total_batches}（累計{len(result)}件）')
        if batch_idx < total_batches - 1:
            time.sleep(0.5)  # API負荷対策（0.3→0.5に増加）

    return result


# ===== 週次履歴の読み書き =====
def load_history():
    """前週データを読み込む（ファイルがなければ空を返す）"""
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_history(history):
    """週次データをJSONに保存（最大12週分保持）"""
    if len(history) > 12:
        oldest_key = sorted(history.keys())[0]
        del history[oldest_key]
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

def get_last_week_data(history, current_key):
    """前週のデータを返す（同じ週のテスト実行データは除外する）
    current_keyはW4終了日（日曜日）。前週=それより7日以上前のキーのみ対象。"""
    from datetime import datetime as _dt
    try:
        current_date = _dt.strptime(current_key, '%Y-%m-%d').date()
    except ValueError:
        return None
    # 現在のW4終了日より6日以上前のキーのみ「前週」として扱う
    # （同じ週内のテスト実行を排除するため）
    past_keys = sorted([
        k for k in history.keys()
        if k < current_key and (_dt.strptime(k, '%Y-%m-%d').date() <= current_date - timedelta(days=6))
    ])
    if not past_keys:
        return None
    return history[past_keys[-1]]

# 履歴を読み込み
history = load_history()
# キーをW4終了日にする → 同じ週に何度実行しても上書き（テスト実行のゴミが溜まらない）
HISTORY_DATE_KEY = WEEKS[-1][2].strftime('%Y-%m-%d')  # W4の日曜日
last_data = get_last_week_data(history, HISTORY_DATE_KEY)
LAST_PERIOD = last_data.get('period', '—') if last_data else '—'

print(f"📅 前週データ: {'あり（' + LAST_PERIOD + '）' if last_data else 'なし（初回実行）'}")

# ===== ユーティリティ =====
def to_f(s):
    try: return float(str(s).replace(',','').replace('%','').replace('$','').strip())
    except: return 0.0

def clean_id(s):
    return s.strip().replace('=','').replace('"','')

def parse_tx_date(s):
    for fmt in ('%b %d, %Y', '%b  %d, %Y'):
        try: return datetime.strptime(s.strip(), fmt).date()
        except: pass
    return None

def days_listed(s):
    try:
        d = datetime.strptime(s.strip(), '%Y-%m-%d').date()
        return (TODAY - d).days
    except: return 0

def pct_change(new, old):
    """W3→W4の変化を「前週比 XX%」形式で表示（100%=横ばい、110%=10%増、90%=10%減）"""
    if old == 0: return '—'
    ratio = new / old * 100
    return f'{ratio:.0f}%'

# ===== 前週データ参照ヘルパー =====
def get_prev_category(item_id):
    """前週のカテゴリを返す"""
    if not last_data:
        return '—'
    cat_map = [
        ('🔥 コア売れ筋', 'top15_ids'),
        ('⭐ 準売れ筋',   '準売れ筋_ids'),
        ('🌱 育成候補',   '育成_ids'),
        ('⚠️ 要調査',     '要調査_ids'),
        ('🗑 削除L1',     '削除L1_ids'),
        ('🗑 削除L2',     '削除L2_ids'),
    ]
    for cat, key in cat_map:
        if item_id in last_data.get(key, []):
            return cat
    if item_id in last_data.get('all_ids', []):
        return '（圏外）'
    return '🆕 新規'

def get_prev_sold(item_id):
    """前週の販売数を返す（なければNone）"""
    if not last_data:
        return None
    return last_data.get('per_item', {}).get(item_id, {}).get('sold', None)

def get_prev_cvr(item_id):
    """前週のCVRを返す（なければNone）"""
    if not last_data:
        return None
    return last_data.get('per_item', {}).get(item_id, {}).get('cvr', None)

def fmt_delta_sold(current_sold, item_id):
    """前週比販売数の変化を文字列で返す（+3 / -2 / 0 / —）"""
    prev = get_prev_sold(item_id)
    if prev is None:
        return '—'
    delta = int(current_sold) - int(prev)
    if delta > 0:  return f'+{delta}'
    elif delta < 0: return str(delta)
    return '0'

# ===== 標準列幅（全シート共通・統一感のために定数化）=====
W_RANK     = 6    # 順位・優先度
W_TITLE    = 52   # 商品タイトル
W_ITEM_ID  = 14   # Item ID
W_SCORE    = 11   # スコア・ポテンシャル
W_SOLD     = 8    # 販売数
W_CVR      = 8    # CVR
W_PV       = 9    # PV
W_IMPS     = 12   # インプレッション
W_QTY      = 8    # 在庫数
W_DAYS     = 8    # 掲載日数
W_PROMO    = 13   # 広告状態
W_ORGANIC  = 10   # Organic比
W_PREV_CAT = 14   # 前週カテゴリ
W_AD_SALES = 12   # 広告売上
W_MEMO     = 20   # メモ・確認欄
W_CHECK    =  8   # チェックボックス欄（✅入力のみ）
W_REASON   = 28   # 原因・推測
W_SKU      = 12   # SKU（商品管理番号）
W_WATCH    = 10   # ウォッチ数（API取得）
W_PRICE    = 11   # 現在価格USD（API取得）

# ===== スタイル定数 =====
FONT = 'メイリオ'
C_GREEN_BG  = 'E8F5E9'
C_BLUE_BG   = 'E3F2FD'
C_YELLOW_BG = 'FFFDE7'
C_ORANGE_BG = 'FFF3E0'
C_RED_BG    = 'FFEBEE'
C_ALERT_BG  = 'FCE4EC'   # コア落ちアラート
C_GRAY_BG   = 'F5F5F5'
C_DARK      = '1A237E'
C_GREEN_HDR = '1B5E20'
C_BLUE_HDR  = '0D47A1'
C_YELLOW_HDR= 'E65100'
C_ORANGE_HDR= 'BF360C'
C_RED_HDR   = 'B71C1C'
C_ALERT_HDR = 'AD1457'   # コア落ち

def hdr_font(color='FFFFFF', size=10, bold=True):
    return Font(name=FONT, size=size, bold=bold, color=color)

def body_font(size=9, bold=False, color='000000'):
    return Font(name=FONT, size=size, bold=bold, color=color)

def hdr_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def body_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def border_thin():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    """データセル用（折り返しなし・1行表示）"""
    return Alignment(horizontal='left', vertical='center', wrap_text=False)

def left_wrap():
    """説明文・ノートセル用（折り返しあり・複数行OK）"""
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def apply_header_row(ws, row_num, headers, bg_color, txt_color='FFFFFF',
                     row_height=22, set_col_widths=True):
    """ヘッダー行を描画。set_col_widths=Falseで列幅を変更しない"""
    ws.row_dimensions[row_num].height = row_height
    for col_idx, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=h)
        cell.font = hdr_font(color=txt_color)
        cell.fill = hdr_fill(bg_color)
        cell.alignment = center()
        cell.border = border_thin()
        if set_col_widths:
            ws.column_dimensions[get_column_letter(col_idx)].width = w

def apply_body_row(ws, row_num, values, bg_color=None, bold=False, height=26):
    """データ行を描画。デフォルト行高さ26（2行分を確保）"""
    ws.row_dimensions[row_num].height = height
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=v)
        cell.font = body_font(bold=bold)
        if bg_color:
            cell.fill = body_fill(bg_color)
        cell.border = border_thin()
        if isinstance(v, (int, float)):
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = left()   # wrap_text=False：1行表示

# ===== eBay商品ページへのハイパーリンク設定 =====
def add_ebay_link(ws, row, col, item_id):
    """Item IDセルにeBay商品ページへのハイパーリンクを設定（クリックで商品ページが開く）"""
    cell = ws.cell(row=row, column=col)
    cell.hyperlink = f'https://www.ebay.com/itm/{item_id}'
    cell.font = Font(name=FONT, size=9, color='0563C1', underline='single')

# ===== Traffic Report 読み込み（API自動 or CSV手動）=====
def parse_traffic_csv(filepath):
    """Traffic Report CSVを解析して商品リストを返す"""
    with open(filepath, encoding='utf-8-sig') as f:
        lines = f.readlines()
    data = io.StringIO(''.join(lines[5:]))
    result = []
    for row in csv.DictReader(data):
        title = row.get('Listing title','').strip()
        if not title: continue
        start_str = row.get('Item Start Date','').strip()
        total_imps = to_f(row.get('Total impressions', 0))
        pl_imps    = to_f(row.get('Total Promoted Listings impressions (applies to eBay site only)', 0))
        off_imps   = to_f(row.get('Total Promoted Offsite impressions (applies to off-eBay only)', 0))
        org_imps   = to_f(row.get('Total organic impressions on eBay site', 0))
        result.append({
            'title':    title,
            'id':       clean_id(row.get('eBay item ID','')),
            'start':    start_str,
            'days':     days_listed(start_str),
            'imps':     total_imps,
            'pl_imps':  pl_imps,
            'off_imps': off_imps,
            'org_imps': org_imps,
            'pv':       to_f(row.get('Total page views', 0)),
            'sold':     to_f(row.get('Quantity sold', 0)),
            'cvr':      to_f(row.get('Sales conversion rate = Quantity sold/Total page views', 0)),
            'ctr':      to_f(row.get('Click-through rate = Page views from eBay site/Total impressions', 0)),
            'qty':      to_f(row.get('Quantity available', 0)),
            'promo':    row.get('Current promoted listings status','').strip(),
        })
    return result

USE_WEEKLY_TRAFFIC = False   # API or 週別CSVが使えればTrueにする
weekly_traffic_totals = [{'imps':0, 'pv':0, 'sold':0, 'pl_imps':0, 'off_imps':0, 'org_imps':0} for _ in range(4)]

if USE_TRAFFIC_API:
    # ===== Traffic Report API で全自動取得 =====
    from ebay_oauth import has_valid_tokens
    if has_valid_tokens():
        print('📡 Sell Analytics API（Traffic Report）で自動取得中...')
        try:
            # 1) 日別サマリー → 週別集計（429リトライ対応）
            daily_data = None
            for _retry in range(3):
                try:
                    daily_data = fetch_traffic_api_daily(WEEKS[0][1], WEEKS[-1][2])
                    break
                except urllib.error.HTTPError as e:
                    if e.code == 429 and _retry < 2:
                        wait = (_retry + 1) * 10
                        print(f'  ⏳ Rate limit。{wait}秒待機中...')
                        time.sleep(wait)
                    else:
                        raise
            if not daily_data:
                raise Exception('Traffic API日別データ取得に失敗')
            print(f'  ✅ 日別データ: {len(daily_data)}日分')
            weekly_traffic_totals = []
            for wlabel, wstart, wend, wdays in WEEKS:
                wt = {'imps':0, 'pv':0, 'sold':0, 'pl_imps':0, 'off_imps':0, 'org_imps':0}
                for dd in daily_data:
                    if wstart <= dd['date'] <= wend:
                        wt['imps'] += dd['imps']
                        wt['pv']   += dd['pv']
                        wt['sold'] += dd['sold']
                weekly_traffic_totals.append(wt)
                print(f'  {wlabel.split(chr(10))[0]}: Imps={wt["imps"]:,.0f} PV={wt["pv"]:,.0f} Sold={wt["sold"]:.0f}')
            USE_WEEKLY_TRAFFIC = True

            # 2) 商品別Trafficデータ（TOP200×3回 + W4/W3 = 合計5コールで取得）
            # ※全14,413件バッチ取得はレート制限に引っかかるため廃止
            _sc = load_seller_cache()
            time.sleep(1)  # API負荷対策

            # 全期間: 売れた商品TOP200（コア売れ筋・準売れ筋用）
            print(f'  📡 全期間 売上TOP200取得中...')
            sold_top = fetch_traffic_api_listings(WEEKS[0][1], WEEKS[-1][2], sort_metric='-TRANSACTION')
            time.sleep(1)

            # 全期間: PV TOP200（育成候補用）
            print(f'  📡 全期間 PV TOP200取得中...')
            pv_top = fetch_traffic_api_listings(WEEKS[0][1], WEEKS[-1][2], sort_metric='-LISTING_VIEWS_TOTAL')
            time.sleep(1)

            # 全期間: インプレッションTOP200（要調査用）
            print(f'  📡 全期間 Impressions TOP200取得中...')
            imps_top = fetch_traffic_api_listings(WEEKS[0][1], WEEKS[-1][2], sort_metric='-LISTING_IMPRESSION_TOTAL')
            time.sleep(1)

            # W4（前週）売上TOP200（コア売れ筋判定用）
            w4_start, w4_end = WEEKS[3][1], WEEKS[3][2]
            print(f'  📡 W4 売上TOP200取得中...')
            w4_top_list = fetch_traffic_api_listings(w4_start, w4_end, sort_metric='-TRANSACTION')
            w4_traffic = {item['listing_id']: item for item in w4_top_list}
            print(f'  ✅ W4 TOP: {len(w4_traffic)}件')
            time.sleep(1)

            # W3 売上TOP200（コア落ち検知用）
            w3_start, w3_end = WEEKS[2][1], WEEKS[2][2]
            print(f'  📡 W3 売上TOP200取得中...')
            w3_top_list = fetch_traffic_api_listings(w3_start, w3_end, sort_metric='-TRANSACTION')
            w3_traffic = {item['listing_id']: item for item in w3_top_list}
            print(f'  ✅ W3 TOP: {len(w3_traffic)}件')

            # 3つのTOP200を統合して重複排除 → APIで取れた商品のTrafficデータ辞書
            api_traffic = {}  # {item_id: {imps, pv, sold, ctr, cvr, title}}
            for item_list in [sold_top, pv_top, imps_top]:
                for item in item_list:
                    lid = item['listing_id']
                    if lid not in api_traffic:
                        api_traffic[lid] = item
                    else:
                        # 複数リストに出現する場合は値が大きい方を採用
                        existing = api_traffic[lid]
                        if item.get('sold', 0) > existing.get('sold', 0):
                            api_traffic[lid] = item
            print(f'  ✅ 統合結果: {len(api_traffic)}件のTrafficデータ取得')

            # seller_cacheからqty/days/start_dateを取得するヘルパー
            def _sc_qty(iid):
                v = _sc.get(iid, {}).get('qty', 0)
                return int(v) if isinstance(v, (int, float)) else 0
            def _sc_days(iid):
                sd = _sc.get(iid, {}).get('start_date', '')
                if sd:
                    try: return (TODAY - datetime.strptime(sd, '%Y-%m-%d').date()).days
                    except: pass
                return 0
            def _sc_start(iid):
                return _sc.get(iid, {}).get('start_date', '')

            # itemsリスト構築（APIデータ + セラーキャッシュ）
            items = []
            for iid, td in api_traffic.items():
                sc_data = _sc.get(iid, {})
                w4d = w4_traffic.get(iid, {})
                w3d = w3_traffic.get(iid, {})
                # CVR: Traffic APIはratio(0.00x)で返す → %に変換（CSVと単位を合わせる）
                cvr_all = td.get('cvr', 0) * 100 if td.get('cvr', 0) < 1 else td.get('cvr', 0)
                cvr_w3  = w3d.get('cvr', 0) * 100 if w3d.get('cvr', 0) < 1 else w3d.get('cvr', 0)
                cvr_w4  = w4d.get('cvr', 0) * 100 if w4d.get('cvr', 0) < 1 else w4d.get('cvr', 0)
                items.append({
                    'title':    td.get('title', sc_data.get('title', '')),
                    'id':       iid,
                    'start':    _sc_start(iid),
                    'days':     _sc_days(iid),
                    'imps':     td.get('imps', 0),
                    'pl_imps':  0,
                    'off_imps': 0,
                    'org_imps': 0,
                    'pv':       td.get('pv', 0),
                    'sold':     td.get('sold', 0),
                    'cvr':      cvr_all,
                    'ctr':      td.get('ctr', 0) * 100 if td.get('ctr', 0) < 1 else td.get('ctr', 0),
                    'qty':      _sc_qty(iid),
                    'promo':    '',
                    'weekly_sold': [0, 0, w3d.get('sold', 0), w4d.get('sold', 0)],
                    'weekly_pv':   [0, 0, w3d.get('pv', 0),   w4d.get('pv', 0)],
                    'weekly_imps': [0, 0, w3d.get('imps', 0), w4d.get('imps', 0)],
                    'weekly_cvr':  [0, 0, cvr_w3, cvr_w4],
                })
            # APIに出てこない商品もセラーキャッシュから追加（削除候補判定用）
            api_ids = set(api_traffic.keys())
            for iid, sc_data in _sc.items():
                if iid not in api_ids:
                    items.append({
                        'title': sc_data.get('title', ''),
                        'id': iid, 'start': _sc_start(iid), 'days': _sc_days(iid),
                        'imps': 0, 'pl_imps': 0, 'off_imps': 0, 'org_imps': 0,
                        'pv': 0, 'sold': 0, 'cvr': 0, 'ctr': 0,
                        'qty': _sc_qty(iid),
                        'promo': '',
                        'weekly_sold': [0]*4, 'weekly_pv': [0]*4,
                        'weekly_imps': [0]*4, 'weekly_cvr': [0.0]*4,
                    })
            print(f'  ✅ Traffic API完了: {len(items)}件（APIデータあり{len(api_traffic)}件 / 全{len(_sc)}件）')

        except Exception as e:
            print(f'  ❌ Traffic API エラー: {e}')
            print(f'  📋 CSVフォールバックに切り替えます...')
            USE_TRAFFIC_API_ACTIVE = False
            items = parse_traffic_csv(TRAFFIC_FILE_COMBINED)
            for item in items:
                item['weekly_sold'] = [0]*4
                item['weekly_pv'] = [0]*4
                item['weekly_imps'] = [0]*4
                item['weekly_cvr'] = [0.0]*4
    else:
        print('⚠️ OAuth未設定。CSVフォールバック。初回認証: python ebay_oauth.py')
        items = parse_traffic_csv(TRAFFIC_FILE_COMBINED)
        for item in items:
            item['weekly_sold'] = [0]*4
            item['weekly_pv'] = [0]*4
            item['weekly_imps'] = [0]*4
            item['weekly_cvr'] = [0.0]*4
else:
    # CSVモード
    print('📊 Traffic Report CSV読み込み中...')
    items = parse_traffic_csv(TRAFFIC_FILE_COMBINED)
    for item in items:
        item['weekly_sold'] = [0]*4
        item['weekly_pv']   = [0]*4
        item['weekly_imps'] = [0]*4
        item['weekly_cvr']  = [0.0]*4
    print(f'  全期間: {len(items)}件')

# ===== Advertising Report 読み込み（CSVがなければスキップ）=====
ad_by_item = defaultdict(lambda: {'plg_s':0,'plg_fee':0,'plp_s':0,'off_s':0})
if os.path.exists(ADS_FILE):
    with open(ADS_FILE, encoding='utf-8-sig') as f:
        ads_lines = f.readlines()
    ads_data = io.StringIO(''.join(ads_lines[2:]))
    for row in csv.DictReader(ads_data):
        iid  = row.get('Item ID','').strip()
        ad   = row.get('Ad type','')
        st   = row.get('Campaign strategy','')
        sale_type = row.get('Sale type','')
        s    = to_f(row.get('Sales','0'))
        fee  = to_f(row.get('Ad fees (General)','-'))
        if 'Offsite' in ad:
            ad_by_item[iid]['off_s'] += s
        elif 'Priority' in st:
            ad_by_item[iid]['plp_s'] += s
        elif 'General' in st and 'Halo' not in sale_type:
            ad_by_item[iid]['plg_s']  += s
            ad_by_item[iid]['plg_fee']+= fee
    print(f'📊 Advertising Report CSV: {len(ad_by_item)}件の広告データ読み込み')
else:
    print('⚠️ Advertising Report CSVなし。広告データは空で続行します。')

for item in items:
    ad = ad_by_item.get(item['id'], {})
    item['plg_s']    = ad.get('plg_s', 0)
    item['plg_fee']  = ad.get('plg_fee', 0)
    item['plp_s']    = ad.get('plp_s', 0)
    item['off_s']    = ad.get('off_s', 0)
    item['ad_sales'] = item['plg_s'] + item['plp_s'] + item['off_s']
    # API modeでpromoが空の場合、広告データから推定
    if not item.get('promo') and (ad.get('plg_s', 0) > 0 or ad.get('plp_s', 0) > 0 or ad.get('off_s', 0) > 0):
        parts = []
        if ad.get('plg_s', 0) > 0: parts.append('PLG')
        if ad.get('plp_s', 0) > 0: parts.append('PLP')
        if ad.get('off_s', 0) > 0: parts.append('Offsite')
        item['promo'] = '/'.join(parts)

# ===== トランザクションデータ取得（API or CSV）=====
week_stats = []
for wlabel, wstart, wend, wdays in WEEKS:
    week_stats.append({
        'label': wlabel, 'start': wstart, 'end': wend, 'days': wdays,
        'orders':0, 'gross':0.0, 'fvf':0.0, 'intl':0.0,
        'plg_fee':0.0, 'off_fee':0.0, 'qty':0,
    })

total_orders = 0
total_gross = total_fvf = total_intl = total_plg_fee_tx = total_off_fee = 0.0
total_qty = 0

# サイト別集計（通貨ベース: USD→US, GBP→UK, EUR→EU等）
site_stats = defaultdict(lambda: {'orders':0, 'gross_usd':0.0, 'gross_native':0.0,
                                   'currency':'', 'fvf_usd':0.0, 'qty':0})
# サイト×週別集計（サマリーの週別売上セクション用）
site_week_stats = defaultdict(lambda: [{'orders':0, 'gross':0.0, 'qty':0} for _ in range(4)])
# 週別売れたSKUセット（売上ゼロ比率の週別算出用・API使用時に上書き）
week_sold_skus = [set() for _ in range(4)]

if USE_TRANSACTION_API:
    # --- GetOrders APIで全期間の注文を取得（全サイト・全通貨）---
    print('📡 GetOrders APIからトランザクションデータを取得中...')
    api_period_start = WEEKS[0][1]   # W1開始日
    api_period_end   = WEEKS[-1][2]  # W4終了日（完了週の日曜日）
    api_orders = fetch_orders_api(api_period_start, api_period_end)
    print(f'  ✅ 取得完了: {len(api_orders)}件（全サイト）')

    # 週別に売れたSKUを追跡（売上ゼロ比率の週別算出用）
    _tmp_cache = load_seller_cache()
    week_sold_skus = [set() for _ in range(4)]  # W1〜W4ごとの売れたSKUセット

    # 週別に振り分け（USD換算済み）
    for order in api_orders:
        d = order['date']
        gross = order['gross']        # USD換算済み
        fvf   = order['fvf']          # USD換算済み
        qty   = order['qty']
        site  = order['site']
        currency = order['currency']
        # International fee はAPIに含まれない → 売上比率で推定
        intl  = gross * EST_INTL_RATE

        total_orders += 1
        total_gross  += gross
        total_fvf    += fvf
        total_intl   += intl
        total_qty    += qty

        # サイト別集計（通貨ベースで振り分け: USD=US, GBP=UK, EUR=EU等）
        site_name = currency_site(currency)
        ss = site_stats[site_name]
        ss['orders'] += 1
        ss['gross_usd'] += gross
        ss['gross_native'] += order['gross_native']
        ss['currency'] = currency
        ss['fvf_usd'] += fvf
        ss['qty'] += qty

        for wi, ws in enumerate(week_stats):
            if ws['start'] <= d <= ws['end']:
                ws['orders'] += 1
                ws['gross']  += gross
                ws['fvf']    += fvf
                ws['intl']   += intl
                ws['qty']    += qty
                # サイト×週別集計
                site_week_stats[site_name][wi]['orders'] += 1
                site_week_stats[site_name][wi]['gross']  += gross
                site_week_stats[site_name][wi]['qty']    += qty
                # 週別売れたSKU追跡（ItemID→SKUに変換）
                for iid in order.get('item_ids', []):
                    sku = _tmp_cache.get(iid, {}).get('sku', '')
                    if sku:
                        week_sold_skus[wi].add(sku)
                break

    # PLG・Offsite費用はAPIに含まれない → 売上比率で推定
    total_plg_fee_tx = total_gross * EST_PLG_RATE
    total_off_fee    = total_gross * EST_OFFSITE_RATE
    # 各週にも比率配分
    for ws in week_stats:
        ws['plg_fee'] = ws['gross'] * EST_PLG_RATE
        ws['off_fee'] = ws['gross'] * EST_OFFSITE_RATE

    # サイト別集計の表示（通貨ベース分類）
    print(f'  📊 サイト別内訳（通貨ベース分類）:')
    for sname in SITE_ORDER:
        ss = site_stats.get(sname)
        if not ss or ss['orders'] == 0:
            continue
        if sname == 'US':
            print(f'    {sname}: {ss["orders"]}件 ${ss["gross_usd"]:,.0f}')
        else:
            print(f'    {sname}: {ss["orders"]}件 {ss["currency"]} {ss["gross_native"]:,.0f} (≈${ss["gross_usd"]:,.0f})')
    print(f'  📊 手数料推定: FVF=${total_fvf:,.0f} / Intl≈${total_intl:,.0f} / PLG≈${total_plg_fee_tx:,.0f} / Offsite≈${total_off_fee:,.0f}')

else:
    # --- 従来方式: Transaction Report CSV ---
    with open(TRANS_FILE, encoding='utf-8-sig') as f:
        trans_lines = f.readlines()
    trans_data = io.StringIO(''.join(trans_lines[11:]))

    for row in csv.DictReader(trans_data):
        cur = row.get('Transaction currency','').strip()
        if cur != 'USD': continue
        t     = row.get('Type','').strip()
        d_str = row.get('Transaction creation date','').strip()
        d     = parse_tx_date(d_str)
        if not d: continue
        desc  = row.get('Description','').strip()

        if t == 'Order':
            gross = to_f(row.get('Gross transaction amount', 0))
            fvf   = to_f(row.get('Final Value Fee - fixed', 0)) + to_f(row.get('Final Value Fee - variable', 0))
            intl  = to_f(row.get('International fee', 0))
            qty   = int(to_f(row.get('Quantity', 0)))
            total_orders += 1; total_gross += gross
            total_fvf += fvf; total_intl += intl; total_qty += qty
            for ws in week_stats:
                if ws['start'] <= d <= ws['end']:
                    ws['orders'] += 1; ws['gross'] += gross
                    ws['fvf'] += fvf; ws['intl'] += intl; ws['qty'] += qty
                    break
        elif t == 'Other fee' and 'Promoted Listings - General fee' in desc:
            fee = abs(to_f(row.get('Net amount', 0)))
            total_plg_fee_tx += fee
            for ws in week_stats:
                if ws['start'] <= d <= ws['end']:
                    ws['plg_fee'] += fee; break
        elif t == 'Other fee' and 'Promoted Offsite fee' in desc:
            fee = abs(to_f(row.get('Net amount', 0)))
            total_off_fee += fee
            for ws in week_stats:
                if ws['start'] <= d <= ws['end']:
                    ws['off_fee'] += fee; break

for ws in week_stats:
    ws['net'] = ws['gross'] - ws['fvf'] - ws['intl'] - ws['plg_fee'] - ws['off_fee']

# ===== Traffic集計 =====
# 出品数: ユニークSKU数（同一SKUの重複リスティングを1商品としてカウント）
# seller_cacheはこの時点ではまだ未読込の場合があるので、ここで一時的に読み込む
_temp_cache = load_seller_cache()
def _sku_of(item):
    """アイテムのSKUを取得（seller_cache経由）"""
    return _temp_cache.get(item['id'], {}).get('sku', '')
_all_skus = set()
for i in items:
    sku = _sku_of(i)
    if sku:
        _all_skus.add(sku)
total_items   = len(_all_skus) if _all_skus else len(items)
total_listings = len(items)   # リスティング数（参考値）

def _count_unique_skus(item_list):
    """アイテムリストからユニークSKU数をカウント（商品数=SKU数の原則）"""
    skus = set()
    for i in item_list:
        sku = _sku_of(i)
        if sku:
            skus.add(sku)
    # SKUが取れない場合はリスティング数をフォールバック
    return len(skus) if skus else len(item_list)
total_imps_t  = sum(i['imps'] for i in items)
total_pv_t    = sum(i['pv'] for i in items)
total_sold_t  = sum(i['sold'] for i in items)
total_pl_imps = sum(i['pl_imps'] for i in items)
total_off_imps= sum(i['off_imps'] for i in items)
total_org_imps= sum(i['org_imps'] for i in items)
# 売上ゼロSKU数（ユニークSKU単位）
_zero_sold_skus = set()
for i in items:
    if i['sold'] == 0:
        sku = _sku_of(i)
        if sku:
            _zero_sold_skus.add(sku)
zero_sold_cnt = len(_zero_sold_skus) if _zero_sold_skus else sum(1 for i in items if i['sold'] == 0)
overall_ctr   = total_pv_t / total_imps_t * 100 if total_imps_t > 0 else 0
overall_cvr   = total_sold_t / total_pv_t * 100 if total_pv_t > 0 else 0

# ===== 商品分類 =====
# 🔥 コア売れ筋TOP15（週別データがある場合はW4=前週のみで判定）
if USE_WEEKLY_TRAFFIC:
    # W4（前週）の販売数でランキング
    top15     = sorted(items, key=lambda x: -x.get('weekly_sold', [0]*4)[3])[:15]
    # W4でsold=0のものは除外
    top15     = [i for i in top15 if i.get('weekly_sold', [0]*4)[3] > 0]
else:
    top15     = sorted(items, key=lambda x: -x['sold'])[:15]
top15_ids = {i['id'] for i in top15}

# ⭐ 準売れ筋: sold>0 かつ TOP15外
# ポテンシャルスコア = CVR × PV（高いほど「今すぐ伸ばせる」商品）
# 優先度: S=上位10件（最重点）/ A=11-30件（重点）/ B=31-80件（普通）/ C=それ以下
_準売れ筋_raw = [i for i in items if i['sold']>0 and i['id'] not in top15_ids]
for i in _準売れ筋_raw:
    i['pot_score'] = round(i['cvr'] * i['pv'], 1)   # ポテンシャルスコア計算
準売れ筋 = sorted(_準売れ筋_raw, key=lambda x: -x['pot_score'])  # スコア降順で並べ替え
for idx, i in enumerate(準売れ筋):                  # 優先度ラベルを付与
    if   idx < 10:  i['priority'] = 'S 🔥'
    elif idx < 30:  i['priority'] = 'A ⭐'
    elif idx < 80:  i['priority'] = 'B 🌱'
    else:           i['priority'] = 'C —'
準売れ筋_ids = {i['id'] for i in 準売れ筋}

# ⚠️ 要調査: sold=0 & インプ500以上（全件記録し上位50件表示）
# Traffic APIモードでは掲載日数が取れないため days>=90 条件を省略
要調査_all  = [i for i in items if i['sold']==0 and i['imps']>=500
               and (i['days']>=90 if i['days'] > 0 else True)]
要調査      = sorted(要調査_all, key=lambda x: -x['imps'])[:50]
要調査_ids  = {i['id'] for i in 要調査_all}   # 全件のIDセット（育成除外に使用）

# 🌱 育成候補（v3改訂）
#   条件: sold=0 & PV>=閾値 & 要調査に含まれない（ツールバグ疑いを除外）& 上位20件
#   【v3変更点】PV20→50に引き上げ、要調査との重複を除去、スタッフが対処可能な件数に絞る
育成        = sorted(
    [i for i in items if i['sold']==0 and i['pv']>=IKUSEI_PV_THRESHOLD
     and i['id'] not in 要調査_ids],
    key=lambda x: -x['pv']
)[:IKUSEI_MAX]
育成_ids    = {i['id'] for i in 育成}

# 🗑 削除候補L1: 完全死蔵（即削除）
削除L1      = [i for i in items if i['imps']==0 and i['pv']==0 and i['sold']==0]
削除L1_ids  = {i['id'] for i in 削除L1}

# 🗑 削除候補L2: 高齢・ほぼ非表示・売上ゼロ（要確認削除）
削除L2      = sorted(
    [i for i in items if i['id'] not in 削除L1_ids
     and i['imps']<50 and i['pv']<5 and i['sold']==0
     and (i['days']>=180 if i['days'] > 0 else True)],
    key=lambda x: -x['days']
)
削除L2_ids  = {i['id'] for i in 削除L2}

# 全カテゴリのSKU数（商品数=SKU数の原則・バリエーション数でカウントしない）
削除L1_sku_cnt = _count_unique_skus(削除L1)
削除L2_sku_cnt = _count_unique_skus(削除L2)
要調査_sku_cnt = _count_unique_skus(要調査_all)
育成_sku_cnt   = _count_unique_skus(育成)
準売れ筋_sku_cnt = _count_unique_skus(準売れ筋)

# ===== eBay APIデータ取得（GetMyeBaySelling で全商品のSKU・ウォッチ数・価格を一括取得）=====
seller_cache = load_seller_cache()
if FETCH_EBAY_API:
    # キャッシュ有効期限チェック（EBAY_CACHE_DAYS日以内なら再取得しない）
    cache_valid = False
    if seller_cache:
        sample = next(iter(seller_cache.values()), {}).get('fetched', '')
        if sample:
            try:
                age = (TODAY - datetime.strptime(sample, '%Y-%m-%d').date()).days
                # 週次レポート基準日（月曜）は前週新規出品がseller_cacheに反映されていないとTOP15等が空表示になるため強制再取得
                if TODAY.weekday() == 0 and age >= 1:
                    print(f'🔄 月曜のためseller_cache強制再取得（前回:{sample} / {age}日前 → 前週新規出品の取りこぼし防止）')
                elif age < EBAY_CACHE_DAYS:
                    cache_valid = True
                    print(f'✅ eBay APIキャッシュ利用（取得日:{sample} / {age}日前 / {len(seller_cache)}件）')
            except Exception:
                pass   # 日付パース失敗 → 再取得
    # キャッシュに必須フィールドが含まれていなければ強制更新
    if cache_valid and seller_cache:
        _sample_item = next(iter(seller_cache.values()), {})
        missing = [k for k in ('title', 'qty', 'start_date') if k not in _sample_item]
        if missing:
            cache_valid = False
            print(f'📡 キャッシュに{"/".join(missing)}がないため再取得します...')
    if not cache_valid:
        print('📡 GetMyeBaySelling 全件取得中...')
        print(f'   （約{len(items)}件 / 1ページ200件 / ページ間{EBAY_API_DELAY}秒待機）')
        seller_cache = fetch_all_seller_data()
        save_seller_cache(seller_cache)
        print(f'✅ 全件取得完了（{len(seller_cache)}件 → {EBAY_SELLER_CACHE}）')
else:
    print('⏭ eBay APIスキップ（FETCH_EBAY_API=False）')

# ===== コア落ちアラート計算 =====
# 前週TOP15にいたが今週TOP15から外れた商品を特定する
items_by_id = {i['id']: i for i in items}

def get_current_category(item_id):
    """現在のカテゴリを返す"""
    if item_id in top15_ids:    return '🔥 コア売れ筋'
    if item_id in 準売れ筋_ids: return '⭐ 準売れ筋'
    if item_id in 育成_ids:     return '🌱 育成候補'
    if item_id in 要調査_ids:   return '⚠️ 要調査'
    if item_id in 削除L1_ids:   return '🗑 削除L1'
    if item_id in 削除L2_ids:   return '🗑 削除L2'
    return '（その他）'

def auto_reason(item):
    """コア落ちの原因を自動推測する（W3→W4の変化を分析）"""
    if item is None:            return '🗑 出品リストから消えた（削除済？）'
    # W3→W4のデータ変化から原因を推測
    w3_imps = int(item.get('weekly_imps', [0]*4)[2])
    w4_imps = int(item.get('weekly_imps', [0]*4)[3])
    w3_pv   = int(item.get('weekly_pv',   [0]*4)[2])
    w4_pv   = int(item.get('weekly_pv',   [0]*4)[3])
    w3_sold = int(item.get('weekly_sold',  [0]*4)[2])
    w4_sold = int(item.get('weekly_sold',  [0]*4)[3])
    if w4_imps == 0:
        return '🔧 W4でImpressionゼロ（在庫ツールバグ or 出品停止？）'
    if w4_pv == 0 and w4_imps > 0:
        return '🔧 表示されているがクリックゼロ（タイトル・画像の問題？）'
    if w4_pv > 0 and w4_sold == 0:
        # PVがあるのに売れていない → 価格 or 在庫の問題
        if w3_pv > 0 and w4_pv < w3_pv * 0.7:
            return f'📉 PV激減（W3:{w3_pv}→W4:{w4_pv}）。検索順位低下 or 競合に流出の可能性'
        return f'💰 PV{w4_pv}あるが未購入。価格負け or 在庫切れサイズを確認'
    if w4_sold > 0 and w3_sold > 0 and w4_sold < w3_sold:
        # 売れてはいるがTOP15から脱落
        if w4_imps < w3_imps * 0.8:
            return f'📉 Imps減少（{w3_imps:,}→{w4_imps:,}）でsoldも減少。露出低下が原因'
        return f'📉 販売数減少（{w3_sold}→{w4_sold}件）。競合の価格・在庫を確認'
    return '📉 TOP15圏外に。他商品の販売数増加によるランク変動'

if USE_WEEKLY_TRAFFIC:
    # --- 週別データがある場合: W3 TOP15 → W4 TOP15 の比較 ---
    # W3(前週)のTOP15とW4(直近週)のTOP15を比較してコア落ちを検出
    w3_ranked = sorted(items, key=lambda x: -x.get('weekly_sold', [0]*4)[2])
    w3_top15  = [i for i in w3_ranked[:15] if i.get('weekly_sold', [0]*4)[2] > 0]
    w3_top15_ids = {i['id'] for i in w3_top15}
    w4_ranked = sorted(items, key=lambda x: -x.get('weekly_sold', [0]*4)[3])
    w4_top15  = [i for i in w4_ranked[:15] if i.get('weekly_sold', [0]*4)[3] > 0]
    w4_top15_ids = {i['id'] for i in w4_top15}
    コア落ち_ids_w = w3_top15_ids - w4_top15_ids
    コア落ち = []
    for iid in コア落ち_ids_w:
        item = items_by_id.get(iid)
        if not item: continue
        w3s = item.get('weekly_sold', [0]*4)[2]
        w4s = item.get('weekly_sold', [0]*4)[3]
        コア落ち.append({
            'title':      item['title'],
            'id':         iid,
            'prev_sold':  w3s,
            'curr_sold':  w4s,
            'delta_sold': int(w4s) - int(w3s),
            'prev_cvr':   item.get('weekly_cvr', [0.0]*4)[2],
            'curr_cvr':   item.get('weekly_cvr', [0.0]*4)[3],
            'qty':        int(item['qty']),
            'curr_cat':   get_current_category(iid),
            'reason':     auto_reason(item),
        })
    コア落ち.sort(key=lambda x: -x['prev_sold'])
    print(f'  🚨 コア落ち検出: W3 TOP15→W4 TOP15比較 → {len(コア落ち)}件脱落')
elif last_data:
    # --- フォールバック: 前レポートとの比較（weekly_history.json） ---
    prev_top15_ids = set(last_data.get('top15_ids', []))
    コア落ち_ids   = prev_top15_ids - top15_ids
    コア落ち = []
    for iid in コア落ち_ids:
        item      = items_by_id.get(iid)
        prev_info = last_data.get('per_item', {}).get(iid, {})
        prev_sold = prev_info.get('sold', 0)
        prev_cvr  = prev_info.get('cvr', 0.0)
        prev_title= prev_info.get('title', iid)
        curr_sold = item['sold'] if item else 0
        curr_cvr  = item['cvr']  if item else 0.0
        コア落ち.append({
            'title':      item['title'] if item else prev_title,
            'id':         iid,
            'prev_sold':  prev_sold,
            'curr_sold':  curr_sold,
            'delta_sold': int(curr_sold) - int(prev_sold),
            'prev_cvr':   prev_cvr,
            'curr_cvr':   curr_cvr,
            'qty':        int(item['qty']) if item else 0,
            'curr_cat':   get_current_category(iid),
            'reason':     auto_reason(item),
        })
    コア落ち.sort(key=lambda x: -x['prev_sold'])
else:
    コア落ち = []

# ===== AI総評（具体的な要因分析 + 具体的なアクション）を生成 =====
def _item_label(item):
    """【SKU】タイトル の形式で商品を識別する文字列を返す"""
    sku = seller_cache.get(item.get('id', ''), {}).get('sku', '')
    title = item.get('title', '')
    if sku:
        return f'【{sku}】{title}'
    return title

def _core_label(core_item):
    """コア落ちアイテム用（idでseller_cacheから引く）"""
    sku = seller_cache.get(core_item.get('id', ''), {}).get('sku', '')
    title = core_item.get('title', '')
    if sku:
        return f'【{sku}】{title}'
    return title

def generate_ai_review():
    """W3→W4のデータ変化を、具体的な商品名・数値・要因まで掘り下げて分析する"""
    lines = []

    if not USE_WEEKLY_TRAFFIC:
        return ['（週別Trafficデータなし。API取得が有効な場合のみ総評を生成します）']

    wt = weekly_traffic_totals

    # --- 1. 販売数の変化要因を具体的に分析 ---
    w3_sold = wt[2]['sold']
    w4_sold = wt[3]['sold']
    sold_diff = w4_sold - w3_sold

    if sold_diff != 0:
        direction = '増加' if sold_diff > 0 else '減少'
        lines.append(f'【販売数】W3: {w3_sold:.0f}件 → W4: {w4_sold:.0f}件（{sold_diff:+.0f}件 {direction}）')

        # W4でのTOP売れ筋を特定（W3と比較して伸びた商品・落ちた商品を名指し）
        gainers = []  # W3→W4で販売数が伸びた商品
        losers = []   # W3→W4で販売数が落ちた商品
        for item in items:
            ws3 = int(item.get('weekly_sold', [0]*4)[2])
            ws4 = int(item.get('weekly_sold', [0]*4)[3])
            if ws4 > ws3 and ws4 > 0:
                gainers.append((_item_label(item), ws3, ws4, ws4 - ws3))
            elif ws3 > ws4 and ws3 > 0:
                losers.append((_item_label(item), ws3, ws4, ws4 - ws3))
        gainers.sort(key=lambda x: -x[3])
        losers.sort(key=lambda x: x[3])

        if gainers:
            top_g = gainers[:3]
            g_text = ' / '.join(f'{t}（{s3}→{s4}件）' for t, s3, s4, _ in top_g)
            total_gain = sum(x[3] for x in gainers)
            lines.append(f'  伸びた商品（計+{total_gain}件）: {g_text}')
        if losers:
            top_l = losers[:3]
            l_text = ' / '.join(f'{t}（{s3}→{s4}件）' for t, s3, s4, _ in top_l)
            total_loss = sum(x[3] for x in losers)
            lines.append(f'  落ちた商品（計{total_loss}件）: {l_text}')

    # --- 2. CVR変化の要因分析 ---
    w3_cvr = wt[2]['sold'] / wt[2]['pv'] * 100 if wt[2]['pv'] > 0 else 0
    w4_cvr = wt[3]['sold'] / wt[3]['pv'] * 100 if wt[3]['pv'] > 0 else 0
    cvr_diff = w4_cvr - w3_cvr

    lines.append(f'【CVR】W3: {w3_cvr:.3f}% → W4: {w4_cvr:.3f}%（{cvr_diff:+.3f}%）')
    if abs(cvr_diff) > 0.01:
        # PVとSoldの伸び率を比較してCVR変化の要因を特定
        pv_ratio = wt[3]['pv'] / wt[2]['pv'] if wt[2]['pv'] > 0 else 1
        sold_ratio = wt[3]['sold'] / wt[2]['sold'] if wt[2]['sold'] > 0 else 1
        if cvr_diff > 0:
            if sold_ratio > pv_ratio:
                lines.append(f'  → PV増加率{pv_ratio*100:.0f}%に対し販売数増加率{sold_ratio*100:.0f}%が上回っている。閲覧者が購入に至る率が改善しており、価格設定や在庫充足が機能している。')
            else:
                lines.append(f'  → PVの伸び以上に販売数が伸びている。売れ筋の在庫が安定供給されている結果と考えられる。')
        else:
            if pv_ratio > sold_ratio:
                lines.append(f'  → PV増加率{pv_ratio*100:.0f}%に対し販売数増加率{sold_ratio*100:.0f}%が下回っている。見に来ているが買っていない＝価格負けか在庫切れの可能性。コア売れ筋の在庫・価格を確認。')
            else:
                lines.append(f'  → PV・販売数ともに減少。需要自体が落ちているか、競合に流れている可能性。eBay Product Researchで売れ筋の変化を確認。')

    # --- 3. Impressions変化の要因分析 ---
    w3_imp = wt[2]['imps']
    w4_imp = wt[3]['imps']
    imp_diff = w4_imp - w3_imp
    if w3_imp > 0 and abs(imp_diff / w3_imp) > 0.05:
        imp_pct = w4_imp / w3_imp * 100
        lines.append(f'【Impressions】W3: {w3_imp:,.0f} → W4: {w4_imp:,.0f}（前週比{imp_pct:.0f}%）')
        # Impsが大きく変動した商品を特定
        imp_changes = []
        for item in items:
            i3 = int(item.get('weekly_imps', [0]*4)[2])
            i4 = int(item.get('weekly_imps', [0]*4)[3])
            if abs(i4 - i3) > 1000:
                imp_changes.append((_item_label(item), i3, i4, i4 - i3))
        imp_changes.sort(key=lambda x: -abs(x[3]))
        if imp_changes[:3]:
            top_imp = imp_changes[:3]
            imp_text = ' / '.join(f'{t}（{i3:,}→{i4:,}）' for t, i3, i4, _ in top_imp)
            lines.append(f'  変動が大きい商品: {imp_text}')

    # --- 4. コア落ちの具体的な分析 ---
    if コア落ち:
        lines.append(f'【コア落ち】{len(コア落ち)}件がTOP15から脱落')
        for item in コア落ち[:3]:
            reason = item.get('reason', '不明')
            lines.append(f'  ・{_core_label(item)}（{item["prev_sold"]}→{item["curr_sold"]}件）: {reason}')

    # --- 5. サイト別トレンド ---
    site_notes = []
    for sname in SITE_ORDER:
        sw = site_week_stats.get(sname)
        if not sw:
            continue
        w3o = sw[2]['orders']
        w4o = sw[3]['orders']
        if w3o > 0 and w4o != w3o:
            diff = w4o - w3o
            site_notes.append(f'{sname}: {w3o}→{w4o}件（{diff:+d}）')
    if site_notes:
        lines.append(f'【サイト別注文数推移】{" / ".join(site_notes)}')

    # --- 6. 売上ゼロ比率の推移 ---
    if total_items > 0:
        w3_zero = total_items - len(week_sold_skus[2])
        w4_zero = total_items - len(week_sold_skus[3])
        w3_pct = w3_zero / total_items * 100
        w4_pct = w4_zero / total_items * 100
        pct_diff = w4_pct - w3_pct
        lines.append(f'【売上ゼロ比率】W3: {w3_pct:.1f}% → W4: {w4_pct:.1f}%（{pct_diff:+.1f}%）')
        if w4_pct > 95:
            lines.append(f'  → {total_items:,}SKU中、W4に1件でも売れたのは{len(week_sold_skus[3])}SKUのみ。出品の大半が機能していない状態。削除を加速し、売れるSKU密度を上げることが最重要課題。')

    # --- 7. 具体的な次のアクション ---
    lines.append('')
    lines.append('【今週やるべきこと】')
    action_num = 1

    if コア落ち:
        # コア落ちで在庫切れっぽいものを特定
        stockout = [c for c in コア落ち if '在庫' in c.get('reason', '')]
        price = [c for c in コア落ち if '価格' in c.get('reason', '') or '競合' in c.get('reason', '')]
        if stockout:
            names = '・'.join(_core_label(c) for c in stockout[:2])
            lines.append(f'  {action_num}. 在庫切れ疑い: {names} → 仕入先の在庫確認・即補充')
            action_num += 1
        if price:
            names = '・'.join(_core_label(c) for c in price[:2])
            lines.append(f'  {action_num}. 価格負け疑い: {names} → 競合セラーの価格を確認し調整')
            action_num += 1

    # 準売れ筋の中でCVR高い＆PLP未使用の商品を提案
    plp_candidates = [i for i in 準売れ筋[:20] if i.get('plp_s', 0) == 0 and i['cvr'] > 0.5]
    if plp_candidates:
        names = '・'.join(_item_label(i) for i in plp_candidates[:2])
        lines.append(f'  {action_num}. PLP広告候補（CVR高い＆PLP未使用）: {names}')
        action_num += 1

    if total_items > 0 and zero_sold_cnt / total_items > 0.9:
        lines.append(f'  {action_num}. 削除候補の精査を継続（現在{削除L1_sku_cnt}件SKUが即削除対象）')
        action_num += 1

    要調査_ratio_ai = 要調査_sku_cnt / total_items * 100 if total_items > 0 else 0
    if 要調査_ratio_ai > 5:
        lines.append(f'  {action_num}. 要調査{要調査_sku_cnt}件SKUのうち上位10件の在庫ツール・仕入先URLを確認')
        action_num += 1

    return lines

ai_review = generate_ai_review()

# ===== AIからの動的提案を生成 =====
suggestions = []

# 出品数の変化チェック
if last_data:
    last_total = last_data.get('total_items', 0)
    if last_total > 0:
        reduction = last_total - total_items
        if reduction >= 50:
            new_thresh = max(20, IKUSEI_PV_THRESHOLD - 10)
            suggestions.append(
                f'📉 出品数が前回比▼{reduction}件減少（{last_total}→{total_items}件）。'
                f'大掃除が順調に進んでいます。育成候補のPV閾値を{IKUSEI_PV_THRESHOLD}→{new_thresh}に'
                f'下げることを推奨します（次回のスクリプト設定変更で反映）。'
            )

# コア落ちチェック
if コア落ち:
    suggestions.append(
        f'🚨 先週のコア売れ筋から{len(コア落ち)}件が脱落しました。'
        f'「🚨 コア落ち」シートを確認してください。'
    )

# 育成候補の件数チェック
if 育成_sku_cnt < 5:
    suggestions.append(
        f'💡 育成候補が{育成_sku_cnt}件SKUと少なくなっています。'
        f'出品数削減が進んでいる証拠です。PV閾値を{IKUSEI_PV_THRESHOLD}→{max(20, IKUSEI_PV_THRESHOLD-10)}に'
        f'下げることを推奨します（IKUSEI_PV_THRESHOLDを変更してください）。'
    )
elif len(育成) >= IKUSEI_MAX:
    suggestions.append(
        f'💡 育成候補が上限{IKUSEI_MAX}件に達しています。スタッフが全件対応できない場合は'
        f'PV閾値を{IKUSEI_PV_THRESHOLD}→{IKUSEI_PV_THRESHOLD+10}に上げて絞ることを検討してください。'
    )

# 要調査の割合チェック
要調査_ratio = 要調査_sku_cnt / total_items * 100 if total_items > 0 else 0
if 要調査_ratio > 50:
    suggestions.append(
        f'🚨 要調査が全出品の{要調査_ratio:.0f}%（{要調査_sku_cnt}件SKU）を占めています。'
        f'在庫ツールのバグが広範囲に影響している可能性が極めて高い。'
        f'Cowatechへの修正依頼を最優先にしてください。'
    )

if not suggestions:
    suggestions.append('✅ 特に問題なし。引き続き現在の方針で運営を継続してください。')

# ===== Excel ワークブック作成 =====
wb = Workbook()

# ─────────────────────────────────────────────────────────────
# Sheet 0: 💬 AI総評（サマリーの前に配置）
# ─────────────────────────────────────────────────────────────
ws0 = wb.active
ws0.title = '💬 AI総評'
ws0.sheet_view.showGridLines = False
ws0.column_dimensions['A'].width = 120

_r0 = 1
ws0.merge_cells(f'A{_r0}:A{_r0}')
c = ws0.cell(row=_r0, column=1,
    value=f'💬 AI総評 ― {PERIOD_LABEL}  （生成日: {TODAY.strftime("%Y/%m/%d")}）')
c.font = Font(name=FONT, size=13, bold=True, color='FFFFFF')
c.fill = hdr_fill('2E7D32')
c.alignment = Alignment(horizontal='left', vertical='center')
ws0.row_dimensions[_r0].height = 30
_r0 += 1

for line in ai_review:
    c = ws0.cell(row=_r0, column=1, value=line)
    if line.startswith('【'):
        c.font = Font(name=FONT, size=10, bold=True, color='1B5E20')
        ws0.row_dimensions[_r0].height = 24
    elif line.startswith('  '):
        c.font = Font(name=FONT, size=9, color='333333')
        ws0.row_dimensions[_r0].height = max(18, min(80, len(line) * 0.35))
    elif line == '':
        ws0.row_dimensions[_r0].height = 8
    else:
        c.font = Font(name=FONT, size=9, color='1B5E20')
        ws0.row_dimensions[_r0].height = max(18, min(80, len(line) * 0.35))
    c.fill = body_fill('F1F8E9')
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    c.border = border_thin()
    _r0 += 1

# ─────────────────────────────────────────────────────────────
# Sheet 1: 📊 サマリー
# ─────────────────────────────────────────────────────────────
ws1 = wb.create_sheet('📊 サマリー')
ws1.sheet_view.showGridLines = False
# サマリー列幅：週別トレンドのB〜Eを均等幅にして見た目を揃える
# 内訳・補足はC:Gをマージするので各列幅は均等でOK
ws1.column_dimensions['A'].width = 32   # 指標名ラベル
ws1.column_dimensions['B'].width = 16   # 数値 / W1
ws1.column_dimensions['C'].width = 16   # 内訳(C:Gマージ) / W2
ws1.column_dimensions['D'].width = 16   # W3
ws1.column_dimensions['E'].width = 16   # W4
ws1.column_dimensions['F'].width = 20   # 合計（長い文字列に対応）
ws1.column_dimensions['G'].width = 13   # W3→W4変化

def s1_section(ws, row_num, title, color='E8EAF6'):
    ws.merge_cells(f'A{row_num}:G{row_num}')
    c = ws.cell(row=row_num, column=1, value=title)
    c.font = Font(name=FONT, size=10, bold=True, color=C_DARK)
    c.fill = body_fill(color)
    c.alignment = left()
    ws.row_dimensions[row_num].height = 20
    return row_num + 1

row = 1

# タイトル
ws1.merge_cells(f'A{row}:G{row}')
c = ws1.cell(row=row, column=1,
    value=f'📊 eBay 週次レポート  {PERIOD_LABEL}  （生成日: {TODAY.strftime("%Y/%m/%d")}）')
c.font = Font(name=FONT, size=13, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_DARK)
c.alignment = center()
ws1.row_dimensions[row].height = 28
row += 1

# KPIブロック（週別パフォーマンス指標）
if USE_WEEKLY_TRAFFIC:
    # 週別Traffic CSVがある場合 → 売上トレンドと同じ形式で週別表示
    row = s1_section(ws1, row, '■ パフォーマンス指標（週別）')
    apply_header_row(ws1, row,
        [('指標', 32)] +
        [(ws['label'], 16) for ws in week_stats] +
        [('合計', 20), ('W3→W4\n変化', 13)],
        C_DARK, row_height=28, set_col_widths=False)
    row += 1

    # 週別集計値
    wt = weekly_traffic_totals

    # 総出品数を週別に取得（weekly_history.jsonから各週末時点の値を参照）
    _weekly_items = []
    for _wi in range(4):
        _w_end = WEEKS[_wi][2]  # 各週の終了日（日曜）
        _w_key = _w_end.strftime('%Y-%m-%d')
        if _w_key in history:
            _weekly_items.append(history[_w_key].get('total_items', None))
        else:
            _weekly_items.append(None)
    _weekly_items[3] = total_items  # W4は今回の値
    for _wi in range(2, -1, -1):  # 履歴がない週は直近の既知値で埋める
        if _weekly_items[_wi] is None:
            _weekly_items[_wi] = _weekly_items[_wi + 1] if _wi + 1 < 4 and _weekly_items[_wi + 1] is not None else total_items
    _items_delta = _weekly_items[3] - _weekly_items[2]
    _items_delta_str = f'{_items_delta:+,}件' if _items_delta != 0 else '±0'

    kpi_weekly = [
        ('🔵 Impressions（表示回数）',
         [f'{wt[i]["imps"]:,.0f}' for i in range(4)], f'{total_imps_t:,.0f}',
         pct_change(wt[3]['imps'], wt[2]['imps']), True, False),
        ('👁 Page Views（閲覧数）',
         [f'{wt[i]["pv"]:,.0f}' for i in range(4)], f'{total_pv_t:,.0f}',
         pct_change(wt[3]['pv'], wt[2]['pv']), True, False),
        ('🛒 Sold（販売数）',
         [f'{wt[i]["sold"]:.0f}件' for i in range(4)], f'{total_sold_t:.0f}件',
         pct_change(wt[3]['sold'], wt[2]['sold']), True, False),
        ('📈 CTR（クリック率）',
         [f'{(wt[i]["pv"]/wt[i]["imps"]*100 if wt[i]["imps"]>0 else 0):.3f}%' for i in range(4)],
         f'{overall_ctr:.3f}%', '', False, False),
        ('💰 CVR（コンバージョン率）',
         [f'{(wt[i]["sold"]/wt[i]["pv"]*100 if wt[i]["pv"]>0 else 0):.3f}%' for i in range(4)],
         f'{overall_cvr:.3f}%', '', True, False),
        ('📦 総出品数（SKU）',
         [f'{_weekly_items[i]:,}件' for i in range(4)], '',
         _items_delta_str, False, False),
        ('📉 売上ゼロ比率',
         [f'{(total_items - len(week_sold_skus[i]))/total_items*100:.1f}%'
          if total_items > 0 else '—' for i in range(4)],
         '',
         f'{((total_items-len(week_sold_skus[3]))/total_items*100 - (total_items-len(week_sold_skus[2]))/total_items*100):+.1f}%'
         if total_items > 0 and len(week_sold_skus[2]) > 0 else '', True, False),
    ]
    for label, week_vals, total_str, change_str, bold, is_cost in kpi_weekly:
        bg = 'E3F2FD' if bold else None
        apply_body_row(ws1, row, [label] + week_vals + [total_str, change_str],
                       bg_color=bg, bold=bold)
        row += 1

else:
    # フォールバック: 全期間1ファイル（従来の期間合計表示）
    row = s1_section(ws1, row, '■ パフォーマンス指標（期間合計）※週別CSVで週別化可能')
    apply_header_row(ws1, row,
        [('指標', 32), ('数値（期間合計）', 16), ('内訳・補足（C〜G結合）', 16),
         ('', 16), ('', 16), ('', 20), ('', 13)], C_DARK, set_col_widths=False)
    row += 1
    kpi_rows = [
        ('🔵 Total Impressions（表示回数）', f'{total_imps_t:,.0f}',  '※週別CSVで週別表示可能', False),
        ('  ├ Promoted Listings',           f'{total_pl_imps:,.0f}' if total_pl_imps > 0 else '—', f'{total_pl_imps/total_imps_t*100:.1f}% of total' if total_imps_t and total_pl_imps > 0 else '（APIモードでは内訳なし）', True),
        ('  ├ Promoted Offsite',            f'{total_off_imps:,.0f}' if total_off_imps > 0 else '—', f'{total_off_imps/total_imps_t*100:.1f}% of total' if total_imps_t and total_off_imps > 0 else '', True),
        ('  └ Organic',                     f'{total_org_imps:,.0f}' if total_org_imps > 0 else '—', f'{total_org_imps/total_imps_t*100:.1f}% of total' if total_imps_t and total_org_imps > 0 else '', True),
        ('👁 Listing Views（閲覧数）',      f'{total_pv_t:,.0f}',    '期間合計', False),
        ('🛒 Quantity Sold（販売数）',      f'{total_sold_t:.0f} 件','期間合計', False),
        ('📈 CTR',                          f'{overall_ctr:.3f}%',   '期間平均（PV÷Impressions）', False),
        ('💰 CVR',                          f'{overall_cvr:.3f}%',   '期間平均（Sold÷PV）', False),
        ('📦 総出品数（SKU）',               f'{total_items:,} 件',   '', False),
        ('📉 売上ゼロ比率',                   f'{zero_sold_cnt:,} 件',  f'{zero_sold_cnt/total_items*100:.1f}%' if total_items > 0 else '—', False),
    ]
    for label, val, note, is_sub in kpi_rows:
        bg = C_GRAY_BG if is_sub else None
        apply_body_row(ws1, row, [label, val, note, '', '', '', ''], bg_color=bg)
        ws1.merge_cells(f'C{row}:G{row}')
        ws1.cell(row=row, column=3).alignment = left_wrap()
        row += 1

row += 1

# 週別売上トレンド（全サイト + サイト別）
row = s1_section(ws1, row, '■ 週別売上（全サイト合計 → サイト別内訳・USD換算）')
apply_header_row(ws1, row,
    [('指標', 32)] +
    [(ws['label'], 16) for ws in week_stats] +
    [('合計', 20), ('W3→W4\n変化', 13)],
    C_DARK, row_height=28, set_col_widths=False)
row += 1

net_after_plp = total_gross - total_fvf - total_intl - total_plg_fee_tx - total_off_fee - PLP_FEE_TOTAL

# --- 全サイト合計 + サイト別（注文数・売上を交互色で表示）---
# 色定義: 注文数行=薄青, 売上行=薄オレンジ（交互で見やすく）
BG_ORDERS = 'E3F2FD'   # 薄い青
BG_SALES  = 'FFF3E0'   # 薄いオレンジ

# サイト別ラベルマッピング（国旗絵文字付き）
SITE_LABEL = {'US': '【US】', 'UK': '【UK】',
              'EU (DE/IT)': '【EU】',
              'Australia': '【AU】', 'Canada': '【CA】'}

# 全サイト合計
apply_body_row(ws1, row, [
    '【全サイト】注文数',
    *[f'{ws["orders"]}件' for ws in week_stats],
    f'{total_orders}件',
    pct_change(week_stats[3]['orders'], week_stats[2]['orders']),
], bg_color=BG_ORDERS, bold=True)
row += 1
apply_body_row(ws1, row, [
    '【全サイト】売上（USD）',
    *[f'${ws["gross"]:,.0f}' for ws in week_stats],
    f'${total_gross:,.0f}',
    pct_change(week_stats[3]['gross'], week_stats[2]['gross']),
], bg_color=BG_SALES, bold=True)
row += 1

# サイト別内訳（US→UK→EU→AU→CA）
for sname in SITE_ORDER:
    ss = site_stats.get(sname)
    if not ss or ss['orders'] == 0:
        continue
    sw = site_week_stats.get(sname, [{'orders':0,'gross':0.0,'qty':0}]*4)
    pct = ss['gross_usd'] / total_gross * 100 if total_gross > 0 else 0
    slabel = SITE_LABEL.get(sname, sname)
    apply_body_row(ws1, row, [
        f'{slabel} 注文数',
        *[f'{sw[i]["orders"]}件' for i in range(4)],
        f'{ss["orders"]}件（{pct:.0f}%）',
        pct_change(sw[3]['orders'], sw[2]['orders']),
    ], bg_color=BG_ORDERS, bold=(sname == 'US'))
    row += 1
    apply_body_row(ws1, row, [
        f'{slabel} 売上',
        *[f'${sw[i]["gross"]:,.0f}' for i in range(4)],
        f'${ss["gross_usd"]:,.0f}',
        pct_change(sw[3]['gross'], sw[2]['gross']),
    ], bg_color=BG_SALES, bold=(sname == 'US'))
    row += 1

row += 1

# --- 手数料・収支トレンド ---
row = s1_section(ws1, row, '■ 週別手数料・手取り推定（全サイト・USD）')
apply_header_row(ws1, row,
    [('指標', 32)] +
    [(ws['label'], 16) for ws in week_stats] +
    [('合計', 20), ('W3→W4\n変化', 13)],
    C_DARK, row_height=28, set_col_widths=False)
row += 1

w_data_cost = [
    ('FVF（最終価値手数料）',
     [f'-${ws["fvf"]:,.0f}' for ws in week_stats], f'-${total_fvf:,.0f}',
     pct_change(week_stats[3]['fvf'], week_stats[2]['fvf']), False, True),
    ('International fee',
     [f'-${ws["intl"]:,.0f}' for ws in week_stats], f'-${total_intl:,.0f}',
     pct_change(week_stats[3]['intl'], week_stats[2]['intl']), False, True),
    ('PLG広告費',
     [f'-${ws["plg_fee"]:,.0f}' for ws in week_stats], f'-${total_plg_fee_tx:,.0f}',
     pct_change(week_stats[3]['plg_fee'], week_stats[2]['plg_fee']), False, True),
    ('Offsite広告費',
     [f'-${ws["off_fee"]:,.0f}' for ws in week_stats], f'-${total_off_fee:,.0f}',
     pct_change(week_stats[3]['off_fee'], week_stats[2]['off_fee']), False, True),
    ('手取り推定',
     [f'${ws["net"]:,.0f}' for ws in week_stats],
     f'${net_after_plp:,.0f}  ※PLP-${PLP_FEE_TOTAL:.0f}控除後',
     pct_change(week_stats[3]['net'], week_stats[2]['net']), True, False),
]

for label, week_vals, total_str, change_str, bold, is_cost in w_data_cost:
    bg = 'FFF3E0' if is_cost else ('E3F2FD' if bold else None)
    apply_body_row(ws1, row, [label] + week_vals + [total_str, change_str],
                   bg_color=bg, bold=bold)
    row += 1

row += 1

# 収支サマリー
row = s1_section(ws1, row, '■ 収支サマリー（期間全体）')
apply_header_row(ws1, row,
    [('項目', 32), ('金額（USD）', 16), ('備考・実質率（C〜G結合）', 16),
     ('', 16), ('', 16), ('', 20), ('', 13)], C_DARK, set_col_widths=False)
row += 1

finance_rows = [
    ('売上合計（Gross）',             f'${total_gross:,.2f}',          f'USD注文 {total_orders}件', True),
    ('FVF（最終価値手数料）',         f'-${total_fvf:,.2f}',           f'実質率: {total_fvf/total_gross*100:.1f}%', False),
    ('International fee',             f'-${total_intl:,.2f}',           '', False),
    ('PLG広告費（Promoted General）', f'-${total_plg_fee_tx:,.2f}',
     f'ROAS: {total_gross/total_plg_fee_tx:.1f}x' if total_plg_fee_tx>0 else '', False),
    ('Offsite広告費',                 f'-${total_off_fee:,.2f}',        'Transaction Report記載（CPC課金）', False),
    ('PLP広告費（Promoted Priority）', f'-${PLP_FEE_TOTAL:,.2f}',       '⚠ APIで取得不可。Seller Hub広告管理画面の実績値を設定（PLP_FEE_TOTAL変数）', False),
    ('eBay控除後 手取り（推定）',     f'${net_after_plp:,.2f}',        f'収益率: {net_after_plp/total_gross*100:.1f}%', True),
]
for label, amount, note, bold in finance_rows:
    if '手取り' in label:  bg = 'E8F5E9'
    elif amount.startswith('-'): bg = 'FFF3E0'
    else: bg = None
    apply_body_row(ws1, row, [label, amount, note, '', '', '', ''], bg_color=bg, bold=bold)
    # C列〜G列をマージして備考テキストが隠れないようにする
    ws1.merge_cells(f'C{row}:G{row}')
    ws1.cell(row=row, column=3).alignment = left_wrap()
    row += 1

row += 1

# （アクション欄は削除済み — 担当は口頭で決定）

# ─────────────────────────────────────────────────────────────
# Sheet 2: 🔥 コア売れ筋TOP15（前週=W4の販売数でランキング）
# ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('🔥 コア売れ筋TOP15')
ws2.sheet_view.showGridLines = False

# 週ラベル取得（改行を含まない短縮形）
def short_week_label(w_idx):
    lbl = WEEKS[w_idx][0]   # "W1\n2/24-3/2" → "W1"
    return lbl.split('\n')[0]

# W4の期間ラベル
w4_period = WEEKS[3][0].replace('\n', ' ')  # "W4 3/17-3/23"

if USE_WEEKLY_TRAFFIC:
    # 前週(W4)のデータで表示: 順位/タイトル/ItemID/SKU/W4sold/W3→W4Δ/W4CVR/W4PV/W4Imps/在庫/ウォッチ/生涯販売/前週カテゴリ
    core_cols = [W_RANK, W_TITLE, W_ITEM_ID, W_SKU,
                 W_SOLD, W_SCORE, W_CVR, W_PV, W_IMPS,
                 W_QTY, W_WATCH, W_SOLD, 8, W_PREV_CAT]
    for col_idx, w in enumerate(core_cols):
        ws2.column_dimensions[get_column_letter(col_idx+1)].width = w

    last_col = get_column_letter(len(core_cols))
    ws2.merge_cells(f'A1:{last_col}1')
    c = ws2['A1']
    c.value = f'🔥 コア売れ筋 TOP15 ― 前週 {w4_period} の販売数ランキング'
    c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
    c.fill = hdr_fill(C_GREEN_HDR)
    c.alignment = center()
    ws2.row_dimensions[1].height = 24

    hdr2 = [
        ('順位', W_RANK), ('商品タイトル', W_TITLE), ('Item ID', W_ITEM_ID),
        ('SKU', W_SKU),
        (f'{short_week_label(3)}\nsold', W_SOLD), (f'W3→W4\nΔ', W_SCORE),
        (f'{short_week_label(3)}\nCVR', W_CVR), (f'{short_week_label(3)}\nPV', W_PV),
        (f'{short_week_label(3)}\nインプ', W_IMPS), ('在庫数', W_QTY),
        ('ウォッチ\n(API)', W_WATCH), ('生涯販売\n(API)', W_SOLD),
        ('PLP\n経由', 8),
        ('前週\nカテゴリ', W_PREV_CAT),
    ]
    apply_header_row(ws2, 2, hdr2, C_GREEN_HDR, row_height=30)

    for rank, item in enumerate(top15, 1):
        sku, watch, lsold, _, _ = get_item_api_data(item['id'])
        w4s = int(item.get('weekly_sold', [0]*4)[3])
        w3s = int(item.get('weekly_sold', [0]*4)[2])
        w4_cvr = item.get('weekly_cvr', [0]*4)[3]
        w4_pv  = int(item.get('weekly_pv', [0]*4)[3])
        w4_imp = int(item.get('weekly_imps', [0]*4)[3])
        diff = w4s - w3s
        delta_w = f'+{diff}' if diff > 0 else str(diff) if diff < 0 else '0'
        prev_cat = get_prev_category(item['id'])
        plp_flag = '✅' if item.get('plp_s', 0) > 0 else ''
        bg       = 'C8E6C9' if rank <= 3 else C_GREEN_BG
        apply_body_row(ws2, rank+2, [
            rank, item['title'], item['id'],
            sku,
            w4s, delta_w, f'{w4_cvr:.1f}%', w4_pv,
            f'{w4_imp:,.0f}', int(item['qty']),
            watch if watch != '?' else '—', lsold if lsold != '?' else '—',
            plp_flag,
            prev_cat,
        ], bg_color=bg, height=20)
        add_ebay_link(ws2, rank+2, 3, item['id'])  # C列=Item ID
else:
    # CSVモード: 全期間の合計sold
    core_cols = [W_RANK, W_TITLE, W_ITEM_ID, W_SKU,
                 W_SOLD, W_SCORE, W_CVR, W_PV, W_IMPS,
                 W_QTY, W_AD_SALES, W_DAYS, W_WATCH, W_SOLD, W_PREV_CAT]
    for col_idx, w in enumerate(core_cols):
        ws2.column_dimensions[get_column_letter(col_idx+1)].width = w

    last_col = get_column_letter(len(core_cols))
    ws2.merge_cells(f'A1:{last_col}1')
    c = ws2['A1']
    c.value = f'🔥 コア売れ筋 TOP15（販売数ランキング）― 期間：{PERIOD_LABEL}'
    c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
    c.fill = hdr_fill(C_GREEN_HDR)
    c.alignment = center()
    ws2.row_dimensions[1].height = 24

    hdr2 = [
        ('順位', W_RANK), ('商品タイトル', W_TITLE), ('Item ID', W_ITEM_ID),
        ('SKU', W_SKU),
        ('販売数', W_SOLD), ('前週比\nΔ販売数', W_SCORE), ('CVR', W_CVR), ('PV', W_PV),
        ('インプ', W_IMPS), ('在庫数', W_QTY), ('広告売上', W_AD_SALES),
        ('掲載日数', W_DAYS),
        ('ウォッチ数\n(API)', W_WATCH), ('生涯販売数\n(API)', W_SOLD),
        ('前週カテゴリ', W_PREV_CAT),
    ]
    apply_header_row(ws2, 2, hdr2, C_GREEN_HDR, row_height=30)

    for rank, item in enumerate(top15, 1):
        sku, watch, lsold, _, _ = get_item_api_data(item['id'])
        ad_s     = f'${item["ad_sales"]:,.0f}' if item.get('ad_sales', 0) > 0 else '-'
        delta    = fmt_delta_sold(item['sold'], item['id'])
        prev_cat = get_prev_category(item['id'])
        bg       = 'C8E6C9' if rank <= 3 else C_GREEN_BG
        apply_body_row(ws2, rank+2, [
            rank, item['title'], item['id'],
            sku,
            int(item['sold']), delta, f'{item["cvr"]:.1f}%', int(item['pv']),
            f'{item["imps"]:,.0f}', int(item['qty']), ad_s,
            f'{item["days"]}日',
            watch if watch != '?' else '—', lsold if lsold != '?' else '—',
            prev_cat,
        ], bg_color=bg, height=20)
        add_ebay_link(ws2, rank+2, 3, item['id'])  # C列=Item ID

# ─────────────────────────────────────────────────────────────
# Sheet 3: 🚨 コア落ちアラート（新シート）
# ─────────────────────────────────────────────────────────────
ws_alert = wb.create_sheet('🚨 コア落ち')
ws_alert.sheet_view.showGridLines = False
# A:タイトル B:ItemID C:SKU D:前週sold E:今週sold F:Δ G:CVR H:在庫 I:前週カテゴリ J:今週カテゴリ K:原因
for col, w in zip('ABCDEFGHIJK', [W_TITLE, W_ITEM_ID, W_SKU, W_SOLD, W_SOLD, W_SCORE,
                                    W_CVR, W_QTY, W_PREV_CAT, W_PREV_CAT, W_REASON]):
    ws_alert.column_dimensions[col].width = w

ws_alert.merge_cells('A1:K1')
c = ws_alert['A1']
if USE_WEEKLY_TRAFFIC:
    alert_title = f'🚨 コア落ちアラート ― W3→W4でTOP15から外れた商品（{len(コア落ち)}件）'
else:
    alert_title = f'🚨 コア落ちアラート ― 先週TOP15から外れた商品（{len(コア落ち)}件）前回：{LAST_PERIOD}'
c.value = alert_title
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_ALERT_HDR)
c.alignment = center()
ws_alert.row_dimensions[1].height = 24

if not USE_WEEKLY_TRAFFIC and not last_data:
    # 初回実行の場合はメッセージのみ（週別データもなし）
    ws_alert.merge_cells('A2:K2')
    msg = ws_alert['A2']
    msg.value = '初回実行のため前週データがありません。週別Traffic CSVがあればW3→W4比較が可能です。'
    msg.font = Font(name=FONT, size=10, color='AD1457')
    msg.fill = body_fill('FCE4EC')
    msg.alignment = left_wrap()
    ws_alert.row_dimensions[2].height = 24
elif not コア落ち:
    ws_alert.merge_cells('A2:K2')
    msg = ws_alert['A2']
    msg.value = '✅ 今週はコア落ちなし。全員が先週と同様にTOP15をキープしています。'
    msg.font = Font(name=FONT, size=10, color='1B5E20')
    msg.fill = body_fill('E8F5E9')
    msg.alignment = left_wrap()
    ws_alert.row_dimensions[2].height = 24
else:
    # 説明行
    ws_alert.merge_cells('A2:K2')
    note_a = ws_alert['A2']
    if USE_WEEKLY_TRAFFIC:
        note_a.value = ('→ W3(前週)のTOP15からW4(直近週)で外れた商品。在庫切れ・ツールバグ・競合激化などが原因。'
                        '必ず今週中に原因を調査して対処すること。')
    else:
        note_a.value = ('→ 先週のコア売れ筋（TOP15）から外れた商品。在庫切れ・ツールバグ・競合激化などが原因。'
                        '必ず今週中に原因を調査して対処すること。')
    note_a.font = Font(name=FONT, size=9, color='AD1457')
    note_a.fill = body_fill('FCE4EC')
    note_a.alignment = left_wrap()
    ws_alert.row_dimensions[2].height = 30

    w3_label = short_week_label(2) if USE_WEEKLY_TRAFFIC else '前週'
    w4_label = short_week_label(3) if USE_WEEKLY_TRAFFIC else '今週'
    hdr_a = [
        ('商品タイトル', W_TITLE), ('Item ID', W_ITEM_ID),
        ('SKU', W_SKU),
        (f'{w3_label}\nsold', W_SOLD), (f'{w4_label}\nsold', W_SOLD), ('Δ販売数', W_SCORE),
        (f'{w4_label}\nCVR', W_CVR), ('在庫数', W_QTY),
        ('前週\nカテゴリ', W_PREV_CAT), ('今週\nカテゴリ', W_PREV_CAT), ('考えられる原因（自動推測）', W_REASON),
    ]
    apply_header_row(ws_alert, 3, hdr_a, C_ALERT_HDR, row_height=30)

    for i, ko in enumerate(コア落ち, 4):
        sku, _, _, _, _ = get_item_api_data(ko['id'])
        delta = ko['delta_sold']
        delta_str = str(delta) if delta < 0 else f'+{delta}' if delta > 0 else '0'
        prev_cat = get_prev_category(ko['id'])
        apply_body_row(ws_alert, i, [
            ko['title'], ko['id'],
            sku,
            int(ko['prev_sold']), int(ko['curr_sold']), delta_str,
            f'{ko["curr_cvr"]:.1f}%', ko['qty'],
            prev_cat, ko['curr_cat'], ko['reason'],
        ], bg_color=C_ALERT_BG, height=26)
        add_ebay_link(ws_alert, i, 2, ko['id'])  # B列=Item ID

# ─────────────────────────────────────────────────────────────
# Sheet 4: ⭐ 準売れ筋（ポテンシャルスコア・優先度追加）
# ─────────────────────────────────────────────────────────────
ws3 = wb.create_sheet('⭐ 準売れ筋')
ws3.sheet_view.showGridLines = False
# A:優先度 B:タイトル C:ItemID D:SKU E:ポテンシャル F:販売数 G:CVR H:PV I:インプ J:在庫 K:掲載日数 L:ウォッチ数 M:現在価格 N:前週カテゴリ
for col, w in zip('ABCDEFGHIJKLMN', [W_RANK, W_TITLE, W_ITEM_ID, W_SKU, W_SCORE,
                                      W_SOLD, W_CVR, W_PV, W_IMPS,
                                      W_QTY, W_DAYS,
                                      W_WATCH, W_PRICE, W_PREV_CAT]):
    ws3.column_dimensions[col].width = w

ws3.merge_cells('A1:N1')
c = ws3['A1']
c.value = f'⭐ 準売れ筋（{準売れ筋_sku_cnt}件SKU）― ポテンシャルスコア順（CVR×PV）― S=最重点10件 / A=重点20件 / B=普通50件 / C=後回し'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_BLUE_HDR)
c.alignment = center()
ws3.row_dimensions[1].height = 24

ws3.merge_cells('A2:N2')
note3 = ws3['A2']
note3.value = ('【優先度の見方】S🔥→今週必ず対処 / A⭐→今週中に確認 / B🌱→来週以降 / C—→後回し  '
               '【ポテンシャル】CVR×PV：高いほど「少し手を加えれば大きく伸びる」商品')
note3.font = Font(name=FONT, size=9, color='0D47A1')
note3.fill = body_fill('E3F2FD')
note3.alignment = left_wrap()
ws3.row_dimensions[2].height = 30

hdr3 = [
    ('優先度', W_RANK), ('商品タイトル', W_TITLE), ('Item ID', W_ITEM_ID),
    ('SKU', W_SKU),
    ('ポテンシャル\n(CVR×PV)', W_SCORE), ('販売数', W_SOLD), ('CVR', W_CVR),
    ('PV', W_PV), ('インプ', W_IMPS), ('在庫数', W_QTY),
    ('掲載日数', W_DAYS),
    ('ウォッチ数\n(API)', W_WATCH), ('現在価格\nUSD(API)', W_PRICE),
    ('前週カテゴリ', W_PREV_CAT),
]
apply_header_row(ws3, 3, hdr3, C_BLUE_HDR, row_height=30)

# 優先度ごとに背景色を変える
PRIORITY_BG = {'S 🔥': 'C8E6C9', 'A ⭐': C_BLUE_BG, 'B 🌱': C_YELLOW_BG, 'C —': C_GRAY_BG}

for i, item in enumerate(準売れ筋, 4):
    sku, watch, _, price, _ = get_item_api_data(item['id'])
    prev_cat  = get_prev_category(item['id'])
    bg        = PRIORITY_BG.get(item['priority'], C_BLUE_BG)
    bold      = item['priority'] == 'S 🔥'   # S優先度は太字で強調
    price_str = f'${price:.2f}' if isinstance(price, float) and price > 0 else '—'
    apply_body_row(ws3, i, [
        item['priority'], item['title'], item['id'],
        sku,
        item['pot_score'], int(item['sold']), f'{item["cvr"]:.1f}%',
        int(item['pv']), f'{item["imps"]:,.0f}', int(item['qty']),
        f'{item["days"]}日',
        watch if watch != '?' else '—', price_str,
        prev_cat,
    ], bg_color=bg, bold=bold)
    add_ebay_link(ws3, i, 3, item['id'])  # C列=Item ID

# ─────────────────────────────────────────────────────────────
# Sheet 5: 🌱 育成候補（v3改訂：PV50以上・要調査除外・上位20件）
# ─────────────────────────────────────────────────────────────
ws4 = wb.create_sheet('🌱 育成候補')
ws4.sheet_view.showGridLines = False
# A:タイトル B:ItemID C:SKU D:PV E:インプ F:CTR G:在庫 H:掲載日数 I:広告状態 J:Organic比 K:ウォッチ数 L:前週カテゴリ
for col, w in zip('ABCDEFGHIJKL', [W_TITLE, W_ITEM_ID, W_SKU, W_PV, W_IMPS, W_CVR,
                                    W_QTY, W_DAYS, W_PROMO, W_ORGANIC,
                                    W_WATCH, W_PREV_CAT]):
    ws4.column_dimensions[col].width = w

ws4.merge_cells('A1:L1')
c = ws4['A1']
c.value = (f'🌱 育成候補（{育成_sku_cnt}件SKU）'
           f'― PV{IKUSEI_PV_THRESHOLD}以上・売上ゼロ・要調査除外 '
           f'― クリックされているが購入に繋がっていない（価格・ページ・送料が原因）')
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_YELLOW_HDR)
c.alignment = center()
ws4.row_dimensions[1].height = 24

ws4.merge_cells('A2:L2')
note4 = ws4['A2']
note4.value = ('【v3改訂】要調査（在庫ツールバグ疑い）は除外済。このリストはページが見られているのに買われていない商品。'
               '確認項目：① 価格（競合比較）② サイズ・在庫 ③ タイトル・写真の質 ④ 送料設定')
note4.font = Font(name=FONT, size=9, color='E65100')
note4.fill = body_fill('FFF9C4')
note4.alignment = left_wrap()
ws4.row_dimensions[2].height = 30

hdr4 = [
    ('商品タイトル', W_TITLE), ('Item ID', W_ITEM_ID),
    ('SKU', W_SKU),
    ('PV', W_PV), ('インプ', W_IMPS), ('CTR', W_CVR),
    ('在庫数', W_QTY), ('掲載日数', W_DAYS), ('広告状態', W_PROMO),
    ('Organic比', W_ORGANIC),
    ('ウォッチ数\n(API)', W_WATCH),
    ('前週カテゴリ', W_PREV_CAT),
]
apply_header_row(ws4, 3, hdr4, C_YELLOW_HDR, txt_color='FFFFFF', row_height=30)

for i, item in enumerate(育成, 4):
    sku, watch, _, _, _ = get_item_api_data(item['id'])
    # Organic比: APIモードではorganic/paid内訳が取れないため「—」表示
    org_ratio = f'{item["org_imps"]/item["imps"]*100:.0f}%' if item['imps'] > 0 and item['org_imps'] > 0 else '—'
    prev_cat  = get_prev_category(item['id'])
    apply_body_row(ws4, i, [
        item['title'], item['id'],
        sku,
        int(item['pv']), f'{item["imps"]:,.0f}', f'{item["ctr"]:.2f}%',
        int(item['qty']), f'{item["days"]}日', item['promo'], org_ratio,
        watch if watch != '?' else '—',
        prev_cat,
    ], bg_color=C_YELLOW_BG)
    add_ebay_link(ws4, i, 2, item['id'])  # B列=Item ID

# ─────────────────────────────────────────────────────────────
# Sheet 6: ⚠️ 要調査（先週カテゴリ追加）
# ─────────────────────────────────────────────────────────────
ws5 = wb.create_sheet('⚠️ 要調査')
ws5.sheet_view.showGridLines = False
# A:タイトル B:ItemID C:SKU D:インプ E:PV F:在庫 G:掲載日数 H:広告状態 I:Organic比 J:ウォッチ数 K:生涯販売数 L:現在価格 M:前週カテゴリ N:調査メモ O:確認状況
for col, w in zip('ABCDEFGHIJKLMNO', [W_TITLE, W_ITEM_ID, W_SKU, W_IMPS, W_PV, W_QTY,
                                       W_DAYS, W_PROMO, W_ORGANIC,
                                       W_WATCH, W_SOLD, W_PRICE,
                                       W_PREV_CAT, W_MEMO, W_MEMO]):
    ws5.column_dimensions[col].width = w

ws5.merge_cells('A1:O1')
c = ws5['A1']
c.value = (f'⚠️ 要調査 TOP50（全{要調査_sku_cnt}件SKU中）'
           f'― インプ500以上・売上ゼロ・掲載90日以上 ― 在庫ツールバグ or 仕入先URL切れ疑い')
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_ORANGE_HDR)
c.alignment = center()
ws5.row_dimensions[1].height = 24

ws5.merge_cells('A2:O2')
note5 = ws5['A2']
note5.value = (f'⚠ 全{要調査_sku_cnt}件SKU/{total_items}件SKU({要調査_sku_cnt/total_items*100:.0f}%)がインプあり売上ゼロ'
               f' — 前週カテゴリで「⚠️ 要調査→今週も要調査」の商品は優先して調査すること'
               f' | 確認①在庫ツールURLが有効か ②仕入先に在庫があるか ③価格・競合確認')
note5.font = Font(name=FONT, size=9, color='BF360C')
note5.fill = body_fill('FFF3E0')
note5.alignment = left_wrap()
ws5.row_dimensions[2].height = 30

hdr5 = [
    ('商品タイトル', W_TITLE), ('Item ID', W_ITEM_ID),
    ('SKU', W_SKU),
    ('インプ', W_IMPS), ('PV', W_PV), ('在庫数', W_QTY),
    ('掲載日数', W_DAYS), ('広告状態', W_PROMO), ('Organic比', W_ORGANIC),
    ('ウォッチ数\n(API)', W_WATCH), ('生涯販売数\n(API)', W_SOLD), ('現在価格\nUSD(API)', W_PRICE),
    ('前週カテゴリ', W_PREV_CAT), ('調査済✅\n（▼選択）', W_CHECK), ('備考\n（記入）', W_MEMO),
]
apply_header_row(ws5, 3, hdr5, C_ORANGE_HDR, row_height=30)

for i, item in enumerate(要調査, 4):
    sku, watch, lsold, price, _ = get_item_api_data(item['id'])
    # Organic比: APIモードではorganic/paid内訳が取れないため「—」表示
    org_ratio = f'{item["org_imps"]/item["imps"]*100:.0f}%' if item['imps'] > 0 and item['org_imps'] > 0 else '—'
    prev_cat  = get_prev_category(item['id'])
    price_str = f'${price:.2f}' if isinstance(price, float) and price > 0 else '—'
    # 前週も要調査だったら強調（オレンジ濃い）
    bg = 'FFCC80' if prev_cat == '⚠️ 要調査' else C_ORANGE_BG
    apply_body_row(ws5, i, [
        item['title'], item['id'],
        sku,
        f'{item["imps"]:,.0f}', int(item['pv']), int(item['qty']),
        f'{item["days"]}日', item['promo'], org_ratio,
        watch if watch != '?' else '—', lsold if lsold != '?' else '—', price_str,
        prev_cat, '', '',
    ], bg_color=bg)
    add_ebay_link(ws5, i, 2, item['id'])  # B列=Item ID

# 要調査シートのN列にドロップダウン（✅選択式チェックボックス）を追加
if len(要調査) > 0:
    dv_req = DataValidation(type='list', formula1='"✅,"', allow_blank=True, showDropDown=False)
    dv_req.sqref = f'N4:N{3 + len(要調査)}'
    ws5.add_data_validation(dv_req)

# ─────────────────────────────────────────────────────────────
# Sheet 7: 🗑 削除候補
# ─────────────────────────────────────────────────────────────
ws6 = wb.create_sheet('🗑 削除候補')
ws6.sheet_view.showGridLines = False

# 重複出品の検出：削除候補（L1+L2）内で同一SKUが複数Item IDに紐づくケース（佐藤大将 2026-04-10 要望）
_削除候補_sku_count = {}
for _it in (削除L1 + 削除L2):
    _s, _, _, _, _ = get_item_api_data(_it['id'])
    if _s and str(_s).strip():
        _削除候補_sku_count[_s] = _削除候補_sku_count.get(_s, 0) + 1

def _dup_warning_xlsx(s):
    n = _削除候補_sku_count.get(s, 0) if s else 0
    return f'⚠️ 重複出品の可能性 ({n}件)' if n >= 2 else ''

# A:タイトル B:ItemID C:SKU D:在庫 E:掲載日数 F:広告状態 G:ウォッチ数 H:生涯販売数 I:前週カテゴリ J:削除判定 K:⚠️重複 L:削除済✅ M:メモ
for col, w in zip('ABCDEFGHIJKLM', [W_TITLE, W_ITEM_ID, W_SKU, W_QTY, W_DAYS, W_PROMO,
                                     W_MEMO, W_MEMO, W_PREV_CAT, W_RANK+4, W_MEMO+2, W_CHECK, W_MEMO]):
    ws6.column_dimensions[col].width = w

ws6.merge_cells('A1:M1')
c = ws6['A1']
c.value = f'🗑 削除候補 — L1即削除: {削除L1_sku_cnt}件（SKU） ／ L2要確認削除: {削除L2_sku_cnt}件（SKU）'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_RED_HDR)
c.alignment = center()
ws6.row_dimensions[1].height = 24

ws6.merge_cells('A2:M2')
note6 = ws6['A2']
note6.value = '⚠️ 削除前に必ずeBayで確認：① ウォッチ数（関心度）② 生涯販売数（過去実績） → 両方ゼロに近い場合のみ削除推奨'
note6.font = Font(name=FONT, size=9, color='B71C1C', bold=True)
note6.fill = body_fill('FFEBEE')
note6.alignment = left_wrap()
ws6.row_dimensions[2].height = 30

apply_header_row(ws6, 3,
    [('【L1：即削除】インプ・PV・売上すべてゼロ（{0}件SKU）'.format(削除L1_sku_cnt), W_TITLE),
     ('Item ID', W_ITEM_ID), ('SKU', W_SKU),
     ('在庫数', W_QTY), ('掲載日数', W_DAYS), ('広告状態', W_PROMO),
     ('ウォッチ数\n(API自動)', W_MEMO), ('生涯販売数\n(API自動)', W_MEMO),
     ('前週カテゴリ', W_PREV_CAT), ('削除判定', W_RANK+4),
     ('⚠️ 重複\n(同SKU)', W_MEMO+2), ('削除済✅\n（▼選択）', W_CHECK),
     ('メモ\n（記入）', W_MEMO)],
    C_RED_HDR, row_height=30)

r = 4
for item in 削除L1:
    prev_cat = get_prev_category(item['id'])
    sku, w, ls, _, judge = get_item_api_data(item['id'])
    # 削除判定に応じて背景色を変える
    if   judge == '🚫 削除NG':  bg = 'FFCC80'   # オレンジ：削除NG
    elif judge == '⚠️ 要確認': bg = 'FFF9C4'   # 黄色：要確認
    else:                        bg = 'FFCDD2'   # 赤系：削除OK
    apply_body_row(ws6, r, [
        item['title'], item['id'],
        sku,
        int(item['qty']), f'{item["days"]}日', item['promo'],
        w, ls, prev_cat, judge,
        _dup_warning_xlsx(sku), '', '',
    ], bg_color=bg)
    add_ebay_link(ws6, r, 2, item['id'])  # B列=Item ID
    r += 1

r += 1

ws6.merge_cells(f'A{r}:M{r}')
c2 = ws6.cell(row=r, column=1,
    value=f'【L2：要確認削除】掲載180日以上・インプ50未満・売上ゼロ（{削除L2_sku_cnt}件SKU）― 削除前に上記の確認必須')
c2.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
c2.fill = hdr_fill(C_RED_HDR)
c2.alignment = left()
ws6.row_dimensions[r].height = 20
r += 1

apply_header_row(ws6, r,
    [('商品タイトル', W_TITLE), ('Item ID', W_ITEM_ID), ('SKU', W_SKU),
     ('インプ', W_IMPS), ('PV', W_PV),
     ('在庫数', W_QTY), ('掲載日数', W_DAYS), ('ウォッチ数\n(API自動)', W_MEMO),
     ('前週カテゴリ', W_PREV_CAT), ('削除判定', W_RANK+4),
     ('⚠️ 重複\n(同SKU)', W_MEMO+2), ('削除済✅\n（▼選択）', W_CHECK),
     ('メモ\n（記入）', W_MEMO)],
    '7B1FA2', row_height=30)
r += 1

for item in 削除L2:
    prev_cat = get_prev_category(item['id'])
    sku, w, _, _, judge = get_item_api_data(item['id'])
    # 削除判定に応じて背景色を変える
    if   judge == '🚫 削除NG':  bg = 'FFCC80'   # オレンジ：削除NG
    elif judge == '⚠️ 要確認': bg = 'FFF9C4'   # 黄色：要確認
    else:                        bg = 'F3E5F5'   # 薄紫：削除OK
    apply_body_row(ws6, r, [
        item['title'], item['id'],
        sku,
        f'{item["imps"]:,.0f}', int(item['pv']),
        int(item['qty']), f'{item["days"]}日', w,
        prev_cat, judge,
        _dup_warning_xlsx(sku), '', '',
    ], bg_color=bg)
    add_ebay_link(ws6, r, 2, item['id'])  # B列=Item ID
    r += 1

# 削除候補シートのL列にドロップダウン（✅選択式チェックボックス）― 重複列追加でK→Lにシフト
if r > 4:
    dv_del = DataValidation(type='list', formula1='"✅,"', allow_blank=True, showDropDown=False)
    dv_del.sqref = f'L4:L{r}'
    ws6.add_data_validation(dv_del)

# ─────────────────────────────────────────────────────────────
# Sheet 8: 📋 週次履歴（蓄積）
# ─────────────────────────────────────────────────────────────
ws7 = wb.create_sheet('📋 週次履歴（蓄積）')
ws7.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHIJ', [14, 22, 8, 13, 13, 8, 8, 8, 8, 10]):
    ws7.column_dimensions[col].width = w

ws7.merge_cells('A1:J1')
c = ws7['A1']
c.value = '📋 週次履歴 ― レポート実行ごとに自動蓄積（前週比較の基礎データ）'
c.font = Font(name=FONT, size=11, bold=True, color='FFFFFF')
c.fill = hdr_fill(C_DARK)
c.alignment = center()
ws7.row_dimensions[1].height = 24

ws7.merge_cells('A2:J2')
note7 = ws7['A2']
note7.value = '→ このシートは毎週レポート実行時に自動更新されます（weekly_history.jsonに保存）。削除しないこと。'
note7.font = Font(name=FONT, size=9, color=C_DARK)
note7.fill = body_fill('E8EAF6')
note7.alignment = left_wrap()
ws7.row_dimensions[2].height = 18

hdr7 = [
    ('レポート日', 14), ('期間', 22), ('注文数', 8),
    ('売上Gross', 13), ('手取り推定', 13), ('CVR', 8), ('CTR', 8),
    ('出品数', 8), ('コア落ち', 8), ('要調査件数', 10),
]
apply_header_row(ws7, 3, hdr7, C_DARK)

# 過去の履歴データを表示（新しい順）
history_row = 4
for date_key in sorted(history.keys(), reverse=True):
    hd = history[date_key]
    apply_body_row(ws7, history_row, [
        date_key,
        hd.get('period', '—'),
        hd.get('total_orders', '—'),
        hd.get('total_gross', '—'),
        hd.get('net_after_plp', '—'),
        hd.get('overall_cvr', '—'),
        hd.get('overall_ctr', '—'),
        hd.get('total_items', '—'),
        '—',
        hd.get('要調査_count', '—'),
    ], bg_color='E3F2FD', bold=False)
    history_row += 1

# 今週のデータを追加（まだ保存前なので直接表示）
apply_body_row(ws7, history_row, [
    TODAY.strftime('%Y/%m/%d') + '（今週）',
    PERIOD_LABEL,
    total_orders,
    f'${total_gross:,.0f}',
    f'${net_after_plp:,.0f}',
    f'{overall_cvr:.3f}%',
    f'{overall_ctr:.3f}%',
    total_items,
    len(コア落ち),
    要調査_sku_cnt,
], bg_color='C8E6C9', bold=True)

# ===== 保存 =====
wb.save(OUTPUT)

# ===== 週次履歴を保存（JSONファイルに書き出し）=====
current_week_data = {
    'period':        PERIOD_LABEL,
    'total_orders':  total_orders,
    'total_gross':   f'${total_gross:,.0f}',
    'net_after_plp': f'${net_after_plp:,.0f}',
    'overall_cvr':   f'{overall_cvr:.3f}%',
    'overall_ctr':   f'{overall_ctr:.3f}%',
    'total_items':   total_items,
    '要調査_count':  要調査_sku_cnt,
    'top15_ids':     list(top15_ids),
    '準売れ筋_ids':  [i['id'] for i in 準売れ筋],
    '育成_ids':      [i['id'] for i in 育成],
    '要調査_ids':    list(要調査_ids),
    '削除L1_ids':    [i['id'] for i in 削除L1],
    '削除L2_ids':    [i['id'] for i in 削除L2],
    'all_ids':       [i['id'] for i in items],
    'per_item': {
        i['id']: {
            'sold':  i['sold'],
            'cvr':   i['cvr'],
            'pv':    i['pv'],
            'imps':  i['imps'],
            'title': i['title'][:60],
            'price': seller_cache.get(i['id'], {}).get('price', 0.0),
            'qty':   seller_cache.get(i['id'], {}).get('qty', 0),
        }
        for i in items
    }
}
history[HISTORY_DATE_KEY] = current_week_data
save_history(history)

# ===== 完了メッセージ =====
print(f'✅ レポート生成完了: {OUTPUT}')
print(f'   💾 週次履歴保存: {HISTORY_FILE}')
print(f'   シート構成（9シート）:')
print(f'     💬 AI総評 / 📊 サマリー / 🔥 コア売れ筋TOP15({len(top15)}件) / 🚨 コア落ち({len(コア落ち)}件)')
print(f'     ⭐ 準売れ筋({準売れ筋_sku_cnt}件SKU) / 🌱 育成候補({育成_sku_cnt}件SKU, PV{IKUSEI_PV_THRESHOLD}+)')
print(f'     ⚠️ 要調査({要調査_sku_cnt}件SKU) / 🗑 削除候補(L1:{削除L1_sku_cnt}+L2:{削除L2_sku_cnt}件SKU) / 📋 週次履歴')
print(f'   🌐 サイト別内訳: {", ".join(s for s in SITE_ORDER if site_stats.get(s, {}).get("orders", 0) > 0)}')
print(f'   📦 商品数（ユニークSKU）: {total_items:,}件 / リスティング数: {total_listings:,}')
print(f'')
print(f'   💡 AIからの提案（{len(suggestions)}件）:')
for s in suggestions:
    print(f'     {s}')

# ===== Google Sheets 直接書き込み（v4機能）=====
# True にするとExcelに加えてGoogleスプレッドシートにも直接書き込む
OUTPUT_GSHEETS = True
if OUTPUT_GSHEETS:
    try:
        from write_gsheets import write_report as gsheets_write
        gsheets_write(globals())
    except ImportError:
        print('⚠️ write_gsheets.py が見つかりません。Google Sheets出力をスキップします。')
    except Exception as e:
        print(f'❌ Google Sheets書き込みエラー: {e}')
        import traceback
        traceback.print_exc()
