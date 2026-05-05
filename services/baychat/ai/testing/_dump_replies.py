# -*- coding: utf-8 -*-
"""Excel から各ケース・各モデルの buyerLanguage / jpnLanguage を抜き出す簡易ダンプ。"""
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

if len(sys.argv) < 2:
    print("Usage: python _dump_replies.py <xlsx_path> [case_id_pattern]")
    sys.exit(1)

xlsx_path = sys.argv[1]
pattern = sys.argv[2] if len(sys.argv) >= 3 else None

wb = openpyxl.load_workbook(xlsx_path)
ws = wb["返信比較"]

headers = [c.value for c in ws[1]]
def col_index(h_list, *candidates):
    for cand in candidates:
        if cand in h_list:
            return h_list.index(cand)
    return None

case_id_col = col_index(headers, "ケース")
model_col = col_index(headers, "モデル")
reply_col = col_index(headers, "英語返信")
jpn_col = col_index(headers, "日本語訳")

for row in ws.iter_rows(min_row=2, values_only=True):
    cid = row[case_id_col]
    if pattern and pattern not in str(cid):
        continue
    model = row[model_col]
    reply = row[reply_col]
    jpn = row[jpn_col] if jpn_col is not None else ""
    print("=" * 70)
    print(f"[{cid}] {model}")
    print("--- ENGLISH REPLY ---")
    print(reply)
    print("--- JAPANESE ---")
    print(jpn)
    print()
