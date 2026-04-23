"""
GPT vs Gemini の応答時間を case ごとに比較して、Geminiが速いケースと GPTが遅いケースを抽出
"""
import sys, io, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
import openpyxl
from collections import defaultdict

EXCEL = r"C:\Users\KEISUKE SHIMOMOTO\Desktop\reffort\services\baychat\ai\testing\results\test_gpt_gemini_v2.3_20260415_183559.xlsx"
CASES = r"C:\Users\KEISUKE SHIMOMOTO\Desktop\reffort\services\baychat\ai\testing\test_cases\extracted_20260415_182039.json"

with open(CASES, "r", encoding="utf-8") as f:
    cases_raw = json.load(f)
cat_map = {c["id"]: c["category"] for c in cases_raw}
buyer_msg_map = {c["id"]: c["buyer_message"] for c in cases_raw}

wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb["スコアサマリー"]
headers = [c.value for c in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    rows.append(dict(zip(headers, row)))

by_case = defaultdict(dict)
for r in rows:
    by_case[r["ケース"]][r["モデル"]] = r

# 速度差分析
records = []
for case_id, models in by_case.items():
    gpt = models.get("GPT-4.1-Mini", {})
    gem = models.get("Gemini 2.5 Flash", {})
    gpt_t = gpt.get("応答時間(秒)", 0) or 0
    gem_t = gem.get("応答時間(秒)", 0) or 0
    diff = gpt_t - gem_t  # 正ならGPTの方が遅い
    records.append({
        "case": case_id,
        "cat": cat_map.get(case_id, "?"),
        "gpt_t": gpt_t,
        "gem_t": gem_t,
        "diff": diff,
        "gpt_score": gpt.get("合計(/25)", 0),
        "gem_score": gem.get("合計(/25)", 0),
        "gpt_tokens": 0,  # token数はExcelに入ってないので後で
        "buyer_len": len(buyer_msg_map.get(case_id, "")),
    })

records.sort(key=lambda x: -x["diff"])

print("=== Geminiが速かった（GPTが遅かった）ケース TOP 10 ===\n")
print(f"{'case':<20s} {'cat':<10s} {'GPT秒':>6s} {'Gem秒':>6s} {'差分':>6s}  {'GPT点':>5s}/{'Gem点':>5s}  buyer文字数")
for r in records[:10]:
    print(f"{r['case']:<20s} {r['cat']:<10s} {r['gpt_t']:>6.2f} {r['gem_t']:>6.2f} {r['diff']:>+6.2f}  {r['gpt_score']:>5d}/{r['gem_score']:>5d}  {r['buyer_len']:>5d}")

print(f"\n逆転件数: GPT遅い={sum(1 for r in records if r['diff']>0)} / Gemini遅い={sum(1 for r in records if r['diff']<0)} / 同等(差0.3秒以内)={sum(1 for r in records if abs(r['diff'])<=0.3)}")

# トークン数との関係（別シートから返信の長さを取る）
ws_rep = wb["返信比較"]
h = [c.value for c in ws_rep[1]]
reply_rows = []
for row in ws_rep.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    reply_rows.append(dict(zip(h, row)))

# モデル×ケース別の返信長
reply_len = defaultdict(dict)
for rr in reply_rows:
    reply_len[rr["ケース"]][rr["モデル"]] = len(rr.get("英語返信", "") or "")

print("\n=== 速度と返信長の関係 ===")
print(f"{'case':<20s} {'GPT秒':>6s} {'GPT文字':>7s}  {'Gem秒':>6s} {'Gem文字':>7s}")
for r in records[:10]:
    g_len = reply_len.get(r["case"], {}).get("GPT-4.1-Mini", 0)
    m_len = reply_len.get(r["case"], {}).get("Gemini 2.5 Flash", 0)
    print(f"{r['case']:<20s} {r['gpt_t']:>6.2f} {g_len:>7d}  {r['gem_t']:>6.2f} {m_len:>7d}")

# GPT最遅TOP5
print("\n=== GPT最遅 TOP 5 ===")
gpt_slow = sorted(records, key=lambda x: -x["gpt_t"])[:5]
for r in gpt_slow:
    print(f"  {r['case']:<22s} cat={r['cat']:<10s}  GPT {r['gpt_t']:.2f}秒 / Gem {r['gem_t']:.2f}秒  buyer={r['buyer_len']}文字")

print("\n=== Gemini最遅 TOP 5 ===")
gem_slow = sorted(records, key=lambda x: -x["gem_t"])[:5]
for r in gem_slow:
    print(f"  {r['case']:<22s} cat={r['cat']:<10s}  Gem {r['gem_t']:.2f}秒 / GPT {r['gpt_t']:.2f}秒  buyer={r['buyer_len']}文字")
