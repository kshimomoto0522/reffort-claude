"""
5モデル比較用の集計スクリプト（GPT-4.1-Mini / GPT-5-Mini / GPT-5-Nano / GPT-4.1-Nano / Gemini 2.5 Flash）
"""
import sys, io, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
import openpyxl
from collections import defaultdict

EXCEL = r"C:\Users\KEISUKE SHIMOMOTO\Desktop\reffort\services\baychat\ai\testing\results\test_gpt_gpt5mini_gpt5nano_gpt41nano_gemini_v2.3_20260415_190920.xlsx"
CASES = r"C:\Users\KEISUKE SHIMOMOTO\Desktop\reffort\services\baychat\ai\testing\test_cases\extracted_20260415_185816.json"

wb = openpyxl.load_workbook(EXCEL, data_only=True)
with open(CASES, "r", encoding="utf-8") as f:
    cases = json.load(f)
cat_map = {c["id"]: c["category"] for c in cases}

ws = wb["スコアサマリー"]
headers = [c.value for c in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    rows.append(dict(zip(headers, row)))

by_model = defaultdict(list)
for r in rows:
    by_model[r["モデル"]].append(r)

MODEL_ORDER = ["GPT-4.1-Mini", "GPT-5-Mini", "GPT-5-Nano", "GPT-4.1-Nano", "Gemini 2.5 Flash"]

print(f"=== 全体結果（{len(rows)}実行、12ケース × 5モデル） ===\n")
print(f"{'モデル':<20s} {'合計':>6s} {'中央値':>6s} {'正確':>5s} {'トーン':>6s} {'完全':>5s} {'アクション':>8s} {'自然':>5s} {'速度(秒)':>9s} {'コスト(¥)':>9s}")
for model in MODEL_ORDER:
    data = by_model.get(model, [])
    if not data:
        continue
    def avg(key): return sum(r.get(key, 0) or 0 for r in data) / len(data)
    totals = [r.get("合計(/25)", 0) or 0 for r in data]
    times = [r.get("応答時間(秒)", 0) or 0 for r in data]
    mid = sorted(totals)[len(totals)//2]
    print(f"{model:<20s} {avg('合計(/25)'):>6.2f} {mid:>6d} {avg('正確性'):>5.2f} {avg('トーン'):>6.2f} {avg('完全性'):>5.2f} {avg('アクション'):>8.2f} {avg('自然さ'):>5.2f} {avg('応答時間(秒)'):>5.2f}(最遅{max(times):.1f}) {avg('コスト(¥)'):>7.3f}")

print("\n=== カテゴリ別スコア (/25) ===\n")
cat_scores = defaultdict(lambda: defaultdict(list))
for r in rows:
    cat = cat_map.get(r["ケース"], "不明")
    cat_scores[cat][r["モデル"]].append(r.get("合計(/25)", 0) or 0)

header_row = f"{'カテゴリ':<10s}  " + "  ".join(f"{m:<16s}" for m in MODEL_ORDER) + "  勝者"
print(header_row)
for cat in ["cancel", "return", "claim", "tracking", "discount", "general"]:
    if cat not in cat_scores:
        continue
    scores_row = []
    best_avg = 0
    winner = ""
    for m in MODEL_ORDER:
        vals = cat_scores[cat].get(m, [])
        a = sum(vals)/len(vals) if vals else 0
        scores_row.append(f"{a:>5.2f}({len(vals)})    ")
        if a > best_avg:
            best_avg = a
            winner = m
    print(f"{cat:<10s}  " + "  ".join(scores_row) + f"  ← {winner}")

# 1ユーザーあたり月額コスト
print("\n=== 1ユーザーあたり月額コスト（5回/日×30日=150回想定） ===")
for model in MODEL_ORDER:
    data = by_model.get(model, [])
    if not data:
        continue
    avg_cost = sum(r.get("コスト(¥)", 0) or 0 for r in data) / len(data)
    per_user = avg_cost * 150
    print(f"  {model:<20s}: ¥{per_user:>5.1f} /ユーザー/月  (¥2,200オプション原価率{per_user/2200*100:.2f}%)")

# 月次コストシミュレーション
print("\n=== 月次コスト（想定 35%契約・5回/日 = 月27,667req） ===")
req = 27667
for model in MODEL_ORDER:
    data = by_model.get(model, [])
    if not data:
        continue
    avg_cost = sum(r.get("コスト(¥)", 0) or 0 for r in data) / len(data)
    print(f"  {model:<20s}: 月額¥{int(avg_cost*req):>6,}")

# 勝敗集計（ケース別に最高スコアを取ったモデル）
print("\n=== 各ケースでの勝者（最高スコア） ===")
by_case = defaultdict(dict)
for r in rows:
    by_case[r["ケース"]][r["モデル"]] = r.get("合計(/25)", 0) or 0
wins = defaultdict(int)
ties = 0
for case_id, models in by_case.items():
    max_score = max(models.values())
    winners = [m for m, s in models.items() if s == max_score]
    if len(winners) > 1:
        ties += 1
        for w in winners:
            wins[w] += 1.0/len(winners)
    else:
        wins[winners[0]] += 1

for m in MODEL_ORDER:
    w = wins.get(m, 0)
    print(f"  {m:<20s}: {w:.1f}勝 ({w/12*100:.1f}%)")

# 速度分析
print("\n=== 速度分析（最速・最遅） ===")
for model in MODEL_ORDER:
    data = by_model.get(model, [])
    if not data:
        continue
    times = [(r["ケース"], r.get("応答時間(秒)", 0) or 0) for r in data]
    times.sort(key=lambda x: x[1])
    print(f"  {model:<20s}: 最速 {times[0][1]:.1f}秒({times[0][0][:20]}) / 最遅 {times[-1][1]:.1f}秒({times[-1][0][:20]})")

# 総合判定（スコア × 速度 × コストの評価）
print("\n=== 総合判定（AI Reply本番採用の観点） ===")
for model in MODEL_ORDER:
    data = by_model.get(model, [])
    if not data:
        continue
    avg_score = sum(r.get("合計(/25)", 0) or 0 for r in data) / len(data)
    avg_time = sum(r.get("応答時間(秒)", 0) or 0 for r in data) / len(data)
    max_time = max(r.get("応答時間(秒)", 0) or 0 for r in data)
    avg_cost = sum(r.get("コスト(¥)", 0) or 0 for r in data) / len(data)
    per_user = avg_cost * 150
    print(f"  {model:<20s}: 品質{avg_score:.1f}/25  速度{avg_time:.1f}秒(最遅{max_time:.1f})  コスト¥{per_user:.0f}/user/月")
