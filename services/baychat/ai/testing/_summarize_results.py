"""
Excelの結果から集計値を取り出して標準出力に出す（分析用）
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
import openpyxl
import json
from collections import defaultdict

EXCEL = r"C:\Users\KEISUKE SHIMOMOTO\Desktop\reffort\services\baychat\ai\testing\results\test_gpt_gemini_v2.3_20260415_183559.xlsx"
CASES = r"C:\Users\KEISUKE SHIMOMOTO\Desktop\reffort\services\baychat\ai\testing\test_cases\extracted_20260415_182039.json"

wb = openpyxl.load_workbook(EXCEL, data_only=True)

# Load case categories
with open(CASES, "r", encoding="utf-8") as f:
    cases = json.load(f)
cat_map = {c["id"]: c["category"] for c in cases}

# Summary sheet
ws = wb["スコアサマリー"]
headers = [c.value for c in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    rows.append(dict(zip(headers, row)))

# Group by model
by_model = defaultdict(list)
for r in rows:
    by_model[r["モデル"]].append(r)

print(f"=== 全体結果（{len(rows)}実行） ===\n")

for model, data in by_model.items():
    totals = [r.get("合計(/25)", 0) for r in data]
    accs = [r.get("正確性", 0) for r in data]
    tones = [r.get("トーン", 0) for r in data]
    comps = [r.get("完全性", 0) for r in data]
    acts = [r.get("アクション", 0) for r in data]
    nats = [r.get("自然さ", 0) for r in data]
    times = [r.get("応答時間(秒)", 0) for r in data]
    costs = [r.get("コスト(¥)", 0) for r in data]

    def avg(lst): return sum(lst)/len(lst) if lst else 0

    print(f"【{model}】 {len(data)}件")
    print(f"  合計スコア: {avg(totals):.2f} / 25  (中央値 {sorted(totals)[len(totals)//2]})")
    print(f"  正確性:     {avg(accs):.2f} / 5")
    print(f"  トーン:     {avg(tones):.2f} / 5")
    print(f"  完全性:     {avg(comps):.2f} / 5")
    print(f"  アクション: {avg(acts):.2f} / 5")
    print(f"  自然さ:     {avg(nats):.2f} / 5")
    print(f"  応答時間:   {avg(times):.2f} 秒 (最速 {min(times):.1f}秒 / 最遅 {max(times):.1f}秒)")
    print(f"  1回コスト:  ¥{avg(costs):.3f}")
    print()

# Category breakdown
print("=== カテゴリ別スコア (avg/25) ===\n")
cat_scores = defaultdict(lambda: defaultdict(list))
for r in rows:
    cat = cat_map.get(r["ケース"], "不明")
    cat_scores[cat][r["モデル"]].append(r.get("合計(/25)", 0))

print(f"{'カテゴリ':<10s} {'件数':>4s}  " + "  ".join(f"{m:<20s}" for m in by_model.keys()) + "  勝者")
for cat in ["cancel", "return", "claim", "tracking", "discount", "general"]:
    if cat not in cat_scores:
        continue
    scores_row = []
    winner = ""
    best_avg = 0
    counts = {m: len(v) for m, v in cat_scores[cat].items()}
    for m in by_model.keys():
        vals = cat_scores[cat].get(m, [])
        a = sum(vals)/len(vals) if vals else 0
        scores_row.append(f"{a:>5.2f}/25 ({len(vals)}件)      ")
        if a > best_avg:
            best_avg = a
            winner = m
    print(f"{cat:<10s} {sum(counts.values()):>4d}  " + "  ".join(scores_row) + f"  ← {winner}")

# Win rate: per-case wins
print("\n=== 1対1対決（ケース別勝敗） ===")
by_case = defaultdict(dict)
for r in rows:
    by_case[r["ケース"]][r["モデル"]] = r.get("合計(/25)", 0)

wins = defaultdict(int)
ties = 0
for case_id, models in by_case.items():
    if len(models) < 2:
        continue
    items = list(models.items())
    if items[0][1] > items[1][1]:
        wins[items[0][0]] += 1
    elif items[1][1] > items[0][1]:
        wins[items[1][0]] += 1
    else:
        ties += 1
total_matches = sum(wins.values()) + ties
for m, w in wins.items():
    print(f"  {m}: {w}勝 ({w/total_matches*100:.1f}%)")
print(f"  引き分け: {ties} ({ties/total_matches*100:.1f}%)")

# Monthly cost simulation
print("\n=== 月次コストシミュレーション ===")
# Assumption: 527ユーザー全員がAI Reply使う × 5回/日 × 30日
# But realistic: AI Reply オプション契約者（仮定: 35%） × 5回/日 × 30日
scenarios = [
    ("保守的 (20%契約・3回/日)", 527 * 0.20 * 3 * 30),
    ("想定   (35%契約・5回/日)", 527 * 0.35 * 5 * 30),
    ("強気   (50%契約・8回/日)", 527 * 0.50 * 8 * 30),
]
for label, req_count in scenarios:
    print(f"\n  {label}: 月間{int(req_count):,}リクエスト")
    for model, data in by_model.items():
        avg_cost = sum(r.get("コスト(¥)", 0) for r in data) / len(data)
        total = avg_cost * req_count
        print(f"    {model}: 月額¥{int(total):,} (1回¥{avg_cost:.3f})")

# Per-user monthly cost
print("\n=== 1ユーザーあたり月額コスト（5回/日×30日=150回想定） ===")
for model, data in by_model.items():
    avg_cost = sum(r.get("コスト(¥)", 0) for r in data) / len(data)
    per_user = avg_cost * 150
    print(f"  {model}: ¥{per_user:.1f} /ユーザー/月  (AI Replyオプション¥2,200に対して原価率{per_user/2200*100:.2f}%)")
