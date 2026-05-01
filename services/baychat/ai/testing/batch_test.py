"""
バッチテストエンジン — 複数モデル×複数プロンプトで一括テスト＆AI採点
====================================================================
テストケース（JSON）を読み込み、指定したモデルとプロンプトの全組み合わせで
AI Replyを生成し、AI審判で品質を自動採点し、Excelレポートを出力する。

使い方:
  python batch_test.py                           # デフォルト設定で実行
  python batch_test.py --models gpt gemini       # GPTとGeminiだけ
  python batch_test.py --prompt-versions 2.2 2.3 # v2.2 vs v2.3
  python batch_test.py --cases test_cases/extracted_20260414.json  # DB抽出ケース
"""

import json
import sys
import os
import io
import re
import time
import argparse
from datetime import datetime

# Windows環境の文字化け対策
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# パス設定
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)  # services/baychat/ai/
RESULTS_DIR = os.path.join(SCRIPT_DIR, "results")

# .envの読み込み
from dotenv import dotenv_values
env = dotenv_values(os.path.join(BASE_DIR, ".env"))

OPENAI_API_KEY = env.get("OPENAI_API_KEY")
GEMINI_API_KEY = env.get("GEMINI_API_KEY")
ANTHROPIC_API_KEY = env.get("ANTHROPIC_API_KEY")


# ============================================================
# モデル定義
# ============================================================
# (ID, 表示名, プロバイダー, モデルID, input$/1M, output$/1M)
AVAILABLE_MODELS = {
    "gpt": ("GPT-4.1-Mini", "openai", "gpt-4.1-mini", 0.40, 1.60),
    "gpt5mini": ("GPT-5-Mini", "openai", "gpt-5-mini", 0.25, 2.00),
    "gpt5nano": ("GPT-5-Nano", "openai", "gpt-5-nano", 0.05, 0.40),
    "gpt41nano": ("GPT-4.1-Nano", "openai", "gpt-4.1-nano", 0.10, 0.40),
    "gpt4omini": ("GPT-4o-Mini", "openai", "gpt-4o-mini", 0.15, 0.60),
    "gemini": ("Gemini 2.5 Flash", "gemini", "gemini-2.5-flash", 0.30, 2.50),
    "claude": ("Claude Haiku 4.5", "anthropic", "claude-haiku-4-5-20251001", 0.80, 4.00),
}

JPY_RATE = 150  # 1ドル = 150円


# ============================================================
# プロンプト読み込み
# ============================================================

def load_prompt_from_md(version):
    """prompt_admin_v{version}.md の ```ブロック``` からプロンプト本文を取り出す"""
    md_file = os.path.join(BASE_DIR, f"prompt_admin_v{version}.md")
    if not os.path.exists(md_file):
        raise FileNotFoundError(f"プロンプトファイルが見つかりません: {md_file}")
    with open(md_file, "r", encoding="utf-8") as f:
        content = f.read()
    match = re.search(r"```\n(.*?)```", content, re.DOTALL)
    if not match:
        raise ValueError(f"プロンプト本文が見つかりません: {md_file}")
    return match.group(1).strip()


def is_admin_prompt(content):
    """developerメッセージがadminプロンプトかどうかを判定"""
    return "ROLE" in content and "CONVERSATION STAGE DETECTION" in content and "TONE" in content


def replace_admin_prompt(messages, new_prompt):
    """messagesリストのadminプロンプトを差し替える"""
    replaced = []
    for msg in messages:
        if msg["role"] == "developer" and is_admin_prompt(msg.get("content", "")):
            replaced.append({"role": "developer", "content": "\n" + new_prompt + "\n"})
        else:
            replaced.append(msg)
    return replaced


def get_latest_buyer_message(messages):
    """最後のuserメッセージを取得"""
    for msg in reversed(messages):
        if msg["role"] == "user":
            return msg["content"]
    return ""


def get_product_info(messages):
    """最初のdeveloperメッセージ（商品情報JSON）を取得"""
    for msg in messages:
        if msg["role"] == "developer":
            content = msg.get("content", "")
            if not is_admin_prompt(content):
                try:
                    return json.loads(content)
                except json.JSONDecodeError:
                    return content
    return {}


def get_conversation_history(messages):
    """user/assistant/systemメッセージだけの会話履歴を取得"""
    history = []
    for msg in messages:
        if msg["role"] in ("user", "assistant", "system"):
            history.append(msg)
    return history


# ============================================================
# モデル呼び出し関数（既存のcompare_models.pyから継承）
# ============================================================

def convert_to_gemini_format(messages):
    """OpenAI形式 → Gemini API形式に変換"""
    system_parts = []
    contents = []
    for msg in messages:
        role = msg.get("role", "")
        content = msg.get("content", "")
        if role in ("developer", "system"):
            system_parts.append(content)
        elif role == "user":
            contents.append({"role": "user", "parts": [{"text": content}]})
        elif role == "assistant":
            contents.append({"role": "model", "parts": [{"text": content}]})
    return "\n\n".join(system_parts), contents


def convert_to_claude_format(messages):
    """OpenAI形式 → Claude API形式に変換"""
    system_parts = []
    raw_messages = []
    for msg in messages:
        role = msg.get("role", "")
        content = msg.get("content", "")
        if role in ("developer", "system"):
            system_parts.append(content)
        elif role in ("user", "assistant"):
            raw_messages.append({"role": role, "content": content})
    # 同じroleが連続する場合は結合
    merged = []
    for msg in raw_messages:
        if merged and merged[-1]["role"] == msg["role"]:
            merged[-1]["content"] += "\n" + msg["content"]
        else:
            merged.append(dict(msg))
    return "\n\n".join(system_parts), merged


def call_model(provider, model_id, messages, response_format=None, temperature_override=None):
    """
    指定プロバイダーのAPIを呼び出す。
    返り値: (result_text, input_tokens, output_tokens, elapsed_seconds)

    引数:
      response_format: OpenAI chat.completions互換のresponse_format (json_schema等)
                       本番再現時は payload_builder.get_production_response_format() を渡す
      temperature_override: temperatureを上書き。本番再現時は 0.2
    """
    if provider == "openai":
        from openai import OpenAI
        client = OpenAI(api_key=OPENAI_API_KEY)
        # GPT-5系はtemperature=1のみ対応（0.7指定不可）＋reasoning_effort=minimalで高速化
        is_gpt5 = model_id.startswith("gpt-5")
        start = time.time()
        kwargs = {"model": model_id, "messages": messages}
        if is_gpt5:
            # reasoning_effort を最小にして推論負荷を抑える（本番運用想定に合わせる）
            kwargs["reasoning_effort"] = "minimal"
        else:
            # 本番再現モードでは temperature_override=0.2 を使用
            kwargs["temperature"] = temperature_override if temperature_override is not None else 0.7
        if response_format is not None:
            kwargs["response_format"] = response_format
        response = client.chat.completions.create(**kwargs)
        elapsed = time.time() - start
        return (response.choices[0].message.content,
                response.usage.prompt_tokens,
                response.usage.completion_tokens, elapsed)

    elif provider == "gemini":
        import google.genai as genai
        from google.genai import types
        client = genai.Client(api_key=GEMINI_API_KEY)
        system_instruction, contents = convert_to_gemini_format(messages)
        start = time.time()
        response = client.models.generate_content(
            model=model_id, contents=contents,
            config=types.GenerateContentConfig(
                system_instruction=system_instruction,
                temperature=0.7,
                response_mime_type="application/json",
            )
        )
        elapsed = time.time() - start
        return (response.text,
                response.usage_metadata.prompt_token_count,
                response.usage_metadata.candidates_token_count, elapsed)

    elif provider == "anthropic":
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        system_text, claude_messages = convert_to_claude_format(messages)
        start = time.time()
        response = client.messages.create(
            model=model_id, max_tokens=1024,
            system=system_text, messages=claude_messages, temperature=1.0,
        )
        elapsed = time.time() - start
        return (response.content[0].text,
                response.usage.input_tokens,
                response.usage.output_tokens, elapsed)

    else:
        raise ValueError(f"未知のプロバイダー: {provider}")


# ============================================================
# テストケース読み込み
# ============================================================

def load_test_cases(path=None):
    """
    テストケースを読み込む。
    - path指定なし → services/baychat/ai/ 内の既存gpt_request_*.jsonを使用
    - path指定あり → そのJSONファイルを読み込む
    """
    cases = []

    if path and os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            # リスト形式 → 各要素がテストケース
            for i, case in enumerate(data):
                if isinstance(case, dict) and "input" in case:
                    # 元IDがあれば保持。なければcase_N
                    case_id = case.get("id") or f"case_{i+1}"
                    # テストケースのメタ情報（category, buyer_message, buyer_message_ja など）を保持
                    cases.append({
                        "id": case_id,
                        "messages": case["input"],
                        "category": case.get("category"),
                        "buyer_message": case.get("buyer_message"),
                        "buyer_message_ja": case.get("buyer_message_ja"),
                        "buyer_ebay": case.get("buyer_ebay"),
                    })
                elif isinstance(case, list):
                    cases.append({"id": f"case_{i+1}", "messages": case})
        elif isinstance(data, dict) and "input" in data:
            cases.append({"id": "case_1", "messages": data["input"]})
        print(f"テストケース読み込み: {path} ({len(cases)}件)")
    else:
        # 既存のgpt_request_*.jsonを使用
        files = sorted([
            f for f in os.listdir(BASE_DIR)
            if f.startswith("gpt_request") and f.endswith(".json")
        ])
        for filename in files:
            filepath = os.path.join(BASE_DIR, filename)
            try:
                with open(filepath, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                if isinstance(raw, dict) and "input" in raw:
                    messages = raw["input"]
                else:
                    messages = raw
                cases.append({"id": filename, "messages": messages})
            except Exception as e:
                print(f"⚠️ {filename} の読み込み失敗: {e}")
        print(f"既存テストケース読み込み: {len(cases)}件")

    return cases


# ============================================================
# Excel出力
# ============================================================

def save_results_to_excel(all_results, output_path):
    """
    テスト結果をExcelファイルに保存する。
    openpyxlが必要（なければCSVにフォールバック）。
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        use_excel = True
    except ImportError:
        use_excel = False
        print("⚠️ openpyxlが未インストール。CSV形式で出力します。")

    if use_excel:
        _save_as_excel(all_results, output_path)
    else:
        _save_as_csv(all_results, output_path.replace(".xlsx", ".csv"))


_ILLEGAL_XLSX_RE = re.compile(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]')

def _xlsx_safe(value):
    """openpyxlのIllegalCharacterError回避：制御文字を除去"""
    if isinstance(value, str):
        return _ILLEGAL_XLSX_RE.sub('', value)
    return value


def _save_as_excel(all_results, output_path):
    """openpyxlでExcelファイルを作成"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # ===== シート1: スコアサマリー =====
    ws_summary = wb.active
    ws_summary.title = "スコアサマリー"

    # ヘッダー
    headers = ["ケース", "モデル", "プロンプト", "正確性", "トーン", "完全性",
               "アクション", "自然さ", "合計(/25)", "応答時間(秒)", "コスト(¥)", "サマリー"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # データ行
    row = 2
    for result in all_results:
        score = result.get("score", {})
        ws_summary.cell(row=row, column=1, value=result.get("case_id", ""))
        ws_summary.cell(row=row, column=2, value=result.get("model_name", ""))
        ws_summary.cell(row=row, column=3, value=result.get("prompt_version", ""))
        ws_summary.cell(row=row, column=4, value=score.get("accuracy", 0))
        ws_summary.cell(row=row, column=5, value=score.get("tone", 0))
        ws_summary.cell(row=row, column=6, value=score.get("completeness", 0))
        ws_summary.cell(row=row, column=7, value=score.get("action_clarity", 0))
        ws_summary.cell(row=row, column=8, value=score.get("naturalness", 0))
        ws_summary.cell(row=row, column=9, value=score.get("total", 0))
        ws_summary.cell(row=row, column=10, value=round(result.get("elapsed", 0), 1))
        ws_summary.cell(row=row, column=11, value=round(result.get("cost_jpy", 0), 2))
        ws_summary.cell(row=row, column=12, value=score.get("summary_ja", ""))

        # スコアが低い場合は赤くハイライト
        for col_idx in range(4, 9):
            cell = ws_summary.cell(row=row, column=col_idx)
            if cell.value and cell.value <= 2:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif cell.value and cell.value >= 4:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        row += 1

    # 列幅調整
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 20
    ws_summary.column_dimensions["C"].width = 10
    ws_summary.column_dimensions["L"].width = 40

    # ===== シート2: 返信比較 =====
    ws_replies = wb.create_sheet("返信比較")
    reply_headers = ["ケース", "バイヤーメッセージ", "モデル", "プロンプト",
                     "英語返信", "日本語訳", "合計スコア"]
    for col, header in enumerate(reply_headers, 1):
        cell = ws_replies.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row = 2
    for result in all_results:
        ws_replies.cell(row=row, column=1, value=_xlsx_safe(result.get("case_id", "")))
        ws_replies.cell(row=row, column=2, value=_xlsx_safe(result.get("buyer_message", "")[:200]))
        ws_replies.cell(row=row, column=3, value=_xlsx_safe(result.get("model_name", "")))
        ws_replies.cell(row=row, column=4, value=_xlsx_safe(result.get("prompt_version", "")))
        ws_replies.cell(row=row, column=5, value=_xlsx_safe(result.get("buyer_reply", "")))
        ws_replies.cell(row=row, column=6, value=_xlsx_safe(result.get("jpn_reply", "")))
        ws_replies.cell(row=row, column=7, value=result.get("score", {}).get("total", 0))
        row += 1

    # 列幅
    ws_replies.column_dimensions["B"].width = 50
    ws_replies.column_dimensions["E"].width = 60
    ws_replies.column_dimensions["F"].width = 40

    # ===== シート3: モデル別集計 =====
    ws_agg = wb.create_sheet("モデル別集計")
    # モデル×プロンプト別の平均スコアを計算
    from collections import defaultdict
    agg = defaultdict(lambda: {"scores": [], "costs": [], "times": []})

    for result in all_results:
        key = f"{result.get('model_name', '')} | v{result.get('prompt_version', '')}"
        score = result.get("score", {})
        if score.get("total", 0) > 0:
            agg[key]["scores"].append(score)
            agg[key]["costs"].append(result.get("cost_jpy", 0))
            agg[key]["times"].append(result.get("elapsed", 0))

    agg_headers = ["モデル × プロンプト", "テスト件数", "平均スコア(/25)",
                   "平均正確性", "平均トーン", "平均完全性", "平均アクション", "平均自然さ",
                   "平均応答時間(秒)", "平均コスト(¥)", "月額推定(¥)"]
    for col, header in enumerate(agg_headers, 1):
        cell = ws_agg.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row = 2
    for key, data in sorted(agg.items()):
        n = len(data["scores"])
        avg = lambda field: sum(s.get(field, 0) for s in data["scores"]) / n

        ws_agg.cell(row=row, column=1, value=key)
        ws_agg.cell(row=row, column=2, value=n)
        ws_agg.cell(row=row, column=3, value=round(avg("total"), 1))
        ws_agg.cell(row=row, column=4, value=round(avg("accuracy"), 1))
        ws_agg.cell(row=row, column=5, value=round(avg("tone"), 1))
        ws_agg.cell(row=row, column=6, value=round(avg("completeness"), 1))
        ws_agg.cell(row=row, column=7, value=round(avg("action_clarity"), 1))
        ws_agg.cell(row=row, column=8, value=round(avg("naturalness"), 1))
        ws_agg.cell(row=row, column=9, value=round(sum(data["times"]) / n, 1))
        ws_agg.cell(row=row, column=10, value=round(sum(data["costs"]) / n, 3))

        # 月額推定（527ユーザー × 35% × 5回/日 × 30日）
        monthly_requests = 527 * 0.35 * 5 * 30
        avg_cost_per_req = sum(data["costs"]) / n
        monthly_cost = avg_cost_per_req * monthly_requests
        ws_agg.cell(row=row, column=11, value=round(monthly_cost, 0))
        row += 1

    ws_agg.column_dimensions["A"].width = 35

    # 保存
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"\nExcel保存完了: {output_path}")


def _save_as_csv(all_results, output_path):
    """CSVフォールバック"""
    import csv
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["ケース", "モデル", "プロンプト", "正確性", "トーン",
                         "完全性", "アクション", "自然さ", "合計", "応答時間", "コスト¥", "サマリー"])
        for r in all_results:
            s = r.get("score", {})
            writer.writerow([
                r.get("case_id"), r.get("model_name"), r.get("prompt_version"),
                s.get("accuracy", 0), s.get("tone", 0), s.get("completeness", 0),
                s.get("action_clarity", 0), s.get("naturalness", 0), s.get("total", 0),
                round(r.get("elapsed", 0), 1), round(r.get("cost_jpy", 0), 3),
                s.get("summary_ja", "")
            ])
    print(f"\nCSV保存完了: {output_path}")


# ============================================================
# メイン処理
# ============================================================

def run_batch_test(model_ids, prompt_versions, cases, judge_model="openai",
                   enable_judge=True, use_production_payload=False,
                   forced_template=True, tone="polite",
                   default_seller_name="rioxxrinaxjapan", description=""):
    """
    バッチテストのメインループ。

    引数:
      model_ids: テスト対象モデルのID一覧 (例: ["gpt", "gemini"])
      prompt_versions: テスト対象プロンプトバージョン (例: ["2.2", "2.3"])
      cases: テストケース一覧 (load_test_cases()の返り値)
      judge_model: AI審判に使うモデル
      enable_judge: AI採点を行うかどうか
      use_production_payload: Trueなら本番ペイロード構造で送信（temperature=0.2・json_schema strict）
      forced_template: FORCED TEMPLATEブロックを入れるか（A/B比較用）
      tone: "polite" | "friendly" | "apologetic"
      default_seller_name: FORCED TEMPLATE用のseller_name。TO=セラーID想定
      description: 画面側補足入力(任意)。空なら補足ブロックを挿入しない

    返り値:
      list: 全結果のリスト
    """
    # プロンプト読み込み
    prompts = {}
    for v in prompt_versions:
        prompts[v] = load_prompt_from_md(v)
        print(f"  プロンプト v{v} 読み込み完了")

    # 利用可能なモデルをフィルタ
    models = []
    for mid in model_ids:
        if mid not in AVAILABLE_MODELS:
            print(f"⚠️ 未知のモデルID: {mid} — スキップ")
            continue
        label, provider, model_id, in_price, out_price = AVAILABLE_MODELS[mid]
        # APIキーチェック
        if provider == "openai" and not OPENAI_API_KEY:
            print(f"⚠️ {label}: OPENAI_API_KEY 未設定 — スキップ")
            continue
        if provider == "gemini" and not GEMINI_API_KEY:
            print(f"⚠️ {label}: GEMINI_API_KEY 未設定 — スキップ")
            continue
        if provider == "anthropic" and not ANTHROPIC_API_KEY:
            print(f"⚠️ {label}: ANTHROPIC_API_KEY 未設定 — スキップ")
            continue
        models.append((mid, label, provider, model_id, in_price, out_price))

    if not models:
        print("❌ 利用可能なモデルがありません。.envを確認してください。")
        return []

    # AI審判のインポート（必要な場合のみ）
    if enable_judge:
        from ai_judge import judge_reply

    # テスト実行
    all_results = []
    total_combinations = len(cases) * len(models) * len(prompt_versions)
    current = 0

    print(f"\n{'='*70}")
    print(f"  バッチテスト開始")
    print(f"  ケース: {len(cases)}件 × モデル: {len(models)}種 × プロンプト: {len(prompt_versions)}版")
    print(f"  合計: {total_combinations}回のテスト")
    print(f"  AI採点: {'ON（' + judge_model + '）' if enable_judge else 'OFF'}")
    print(f"{'='*70}\n")

    # 本番再現モードで使うpayload_builder
    if use_production_payload:
        from payload_builder import (
            build_production_payload,
            get_production_response_format,
            PRODUCTION_TEMPERATURE,
        )
        prod_response_format = get_production_response_format()
        prod_temperature = PRODUCTION_TEMPERATURE
        print(f"  [本番再現モード] temperature={prod_temperature} / json_schema strict / "
              f"forced_template={forced_template} / tone={tone}")

    for case in cases:
        case_id = case["id"]
        messages_orig = case.get("messages") or case.get("input") or []
        # test_cases/extracted_*.json は 'input' キー、旧形式は 'messages'。どちらも許容。
        # case辞書自体は本番ペイロードビルダー用に input キーで渡す
        case_for_builder = {"input": messages_orig, **{k: v for k, v in case.items() if k not in ("messages", "input")}}

        # バイヤーの最新メッセージ
        buyer_msg = get_latest_buyer_message(messages_orig)
        product_info = get_product_info(messages_orig)
        conv_history = get_conversation_history(messages_orig)

        # 本番再現モードのFORCED TEMPLATEで使うbuyer_name（TO=バイヤーID想定）
        buyer_name_for_template = case.get("buyer_ebay") or case.get("buyer_message_ja", "") or ""

        for version in prompt_versions:
            if use_production_payload:
                # 本番と同じ構造で組み立て
                # v2.5 は FORCED_TEMPLATE 強制除外 + admin_prompt 内プレースホルダ置換
                messages = build_production_payload(
                    test_case=case_for_builder,
                    admin_prompt_text=prompts[version],
                    tone=tone,
                    buyer_name=buyer_name_for_template,
                    seller_name=default_seller_name,
                    description=description,
                    include_forced_template=forced_template,
                    prompt_version=version,
                )
            else:
                # 旧来: 既存messagesのadmin promptだけ差し替え
                messages = replace_admin_prompt(messages_orig, prompts[version])

            for mid, label, provider, model_id, in_price, out_price in models:
                current += 1
                print(f"[{current}/{total_combinations}] {case_id} | {label} | v{version} ...", end="", flush=True)

                result = {
                    "case_id": case_id,
                    "model_id": mid,
                    "model_name": label,
                    "prompt_version": version,
                    "buyer_message": buyer_msg,
                }

                try:
                    # AI Reply生成（本番再現モード時は temperature=0.2 + json_schema strict）
                    if use_production_payload and provider == "openai":
                        result_text, in_tok, out_tok, elapsed = call_model(
                            provider, model_id, messages,
                            response_format=prod_response_format,
                            temperature_override=prod_temperature,
                        )
                    else:
                        result_text, in_tok, out_tok, elapsed = call_model(provider, model_id, messages)

                    # コスト計算
                    cost_usd = (in_tok / 1_000_000 * in_price) + (out_tok / 1_000_000 * out_price)
                    cost_jpy = cost_usd * JPY_RATE

                    # JSON解析（buyerLanguage / jpnLanguage を取り出す）
                    try:
                        result_json = json.loads(result_text)
                        buyer_reply = result_json.get("buyerLanguage", result_text)
                        jpn_reply = result_json.get("jpnLanguage", "")
                    except json.JSONDecodeError:
                        buyer_reply = result_text
                        jpn_reply = "（JSONパース失敗）"

                    result.update({
                        "buyer_reply": buyer_reply,
                        "jpn_reply": jpn_reply,
                        "input_tokens": in_tok,
                        "output_tokens": out_tok,
                        "elapsed": elapsed,
                        "cost_usd": cost_usd,
                        "cost_jpy": cost_jpy,
                        "raw_output": result_text,
                    })

                    print(f" {elapsed:.1f}秒 | {in_tok}+{out_tok}tok | ¥{cost_jpy:.3f}")

                    # AI採点
                    if enable_judge:
                        print(f"  → AI採点中...", end="", flush=True)
                        try:
                            score = judge_reply(
                                buyer_message=buyer_msg,
                                ai_reply=buyer_reply,
                                product_info=product_info,
                                conversation_history=conv_history,
                                judge_model=judge_model
                            )
                            result["score"] = score
                            print(f" スコア: {score.get('total', '?')}/25 — {score.get('summary_ja', '')}")
                        except Exception as e:
                            result["score"] = {"total": 0, "summary_ja": f"採点エラー: {e}"}
                            print(f" 採点エラー: {e}")

                except Exception as e:
                    result["error"] = str(e)
                    print(f" エラー: {e}")

                all_results.append(result)
                time.sleep(1)  # レート制限対策

    return all_results


# ============================================================
# メインエントリーポイント
# ============================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="BayChat AI Reply バッチテスト")
    parser.add_argument("--models", nargs="+", default=["gpt", "gemini"],
                        help="テストするモデル (gpt / gemini / claude)")
    parser.add_argument("--prompt-versions", nargs="+", default=["2.3"],
                        help="テストするプロンプトバージョン (例: 2.2 2.3)")
    parser.add_argument("--cases", type=str, default=None,
                        help="テストケースJSONファイルのパス")
    parser.add_argument("--judge", type=str, default="openai",
                        help="AI採点に使うモデル (claude / openai / gemini)")
    parser.add_argument("--no-judge", action="store_true",
                        help="AI採点をスキップ（生成結果の比較のみ）")
    parser.add_argument("--limit", type=int, default=None,
                        help="テストケース数の上限")
    parser.add_argument("--production-payload", action="store_true",
                        help="本番ペイロード構造で送信（temperature=0.2・json_schema strict・4 developer blocks）")
    parser.add_argument("--no-forced-template", action="store_true",
                        help="本番再現モード時、FORCED TEMPLATEブロックを外す（A/B比較用）")
    parser.add_argument("--tone", type=str, default="polite",
                        choices=["polite", "friendly", "apologetic"],
                        help="tone設定（本番再現モード時）")
    parser.add_argument("--seller-name", type=str, default="rioxxrinaxjapan",
                        help="FORCED TEMPLATE用のseller_name")
    parser.add_argument("--description", type=str, default="",
                        help="画面側(任意)補足入力。空ならsellerSetting補足ブロックを省略")

    args = parser.parse_args()

    # テストケース読み込み
    cases = load_test_cases(args.cases)
    if args.limit:
        cases = cases[:args.limit]

    if not cases:
        print("❌ テストケースが見つかりません。")
        sys.exit(1)

    # バッチテスト実行
    results = run_batch_test(
        model_ids=args.models,
        prompt_versions=args.prompt_versions,
        cases=cases,
        judge_model=args.judge,
        enable_judge=not args.no_judge,
        use_production_payload=args.production_payload,
        forced_template=not args.no_forced_template,
        tone=args.tone,
        default_seller_name=args.seller_name,
        description=args.description,
    )

    if results:
        # Excel出力
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        models_str = "_".join(args.models)
        prompts_str = "_".join(f"v{v}" for v in args.prompt_versions)
        mode_str = ""
        if args.production_payload:
            mode_str = "_prodON" if not args.no_forced_template else "_prodOFF"
        output_path = os.path.join(RESULTS_DIR, f"test_{models_str}_{prompts_str}{mode_str}_{timestamp}.xlsx")

        save_results_to_excel(results, output_path)

        # サマリー出力
        print(f"\n{'='*70}")
        print(f"  テスト完了サマリー")
        print(f"{'='*70}")
        scored = [r for r in results if r.get("score", {}).get("total", 0) > 0]
        if scored:
            avg_total = sum(r["score"]["total"] for r in scored) / len(scored)
            print(f"  平均スコア: {avg_total:.1f} / 25")

        total_cost = sum(r.get("cost_jpy", 0) for r in results)
        print(f"  合計コスト: ¥{total_cost:.2f}")
        print(f"  結果ファイル: {output_path}")

    print("\n完了。")
