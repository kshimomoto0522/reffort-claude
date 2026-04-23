# -*- coding: utf-8 -*-
"""
既存40ケースExcel から購入前9ケース分を除去し、新9ケースExcelをマージする。
"""
import sys
from openpyxl import load_workbook, Workbook
from copy import copy

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

OLD_PATH = "results/test_gpt_gemini_gpt5nano_v2.6_prodON_20260422_004318.xlsx"
NEW_PATH = "results/test_gpt_gemini_gpt5nano_v2.6_prodON_20260422_011512.xlsx"
OUT_PATH = "results/test_gpt_gemini_gpt5nano_v2.6_merged_20260422.xlsx"

PREPURCHASE_IDS = {
    "synth_A1_cam_stock",
    "synth_A2_jacket_sizing",
    "synth_A3_guitar_ship_time",
    "synth_A4_pokemon_condition",
    "synth_G4_cam_german",
    "synth_H1_bag_price_offer",
    "synth_H2_cards_bulk",
    "synth_H3_watch_re_nego",
    "synth_H4_sneakers_reoffer",
}


def merge_sheet(old_wb, new_wb, sheet_name, id_col_name):
    """指定シートを合体"""
    out_ws = old_wb[sheet_name]
    headers = [c.value for c in out_ws[1]]
    id_col_idx = headers.index(id_col_name) + 1  # 1-indexed

    # 既存の行をフィルタ（購入前9件を削除）
    rows_to_delete = []
    for row_idx in range(2, out_ws.max_row + 1):
        cid = out_ws.cell(row=row_idx, column=id_col_idx).value
        if cid in PREPURCHASE_IDS:
            rows_to_delete.append(row_idx)

    # 降順で削除（index ずれ防止）
    for ri in reversed(rows_to_delete):
        out_ws.delete_rows(ri, 1)

    # 新WBから該当行を追加
    new_ws = new_wb[sheet_name]
    for row_idx in range(2, new_ws.max_row + 1):
        values = [new_ws.cell(row=row_idx, column=ci).value for ci in range(1, len(headers) + 1)]
        out_ws.append(values)

    print(f"  {sheet_name}: 旧{len(rows_to_delete)}件削除 / 新{new_ws.max_row - 1}件追加")


def main():
    print("[1] Excel合体開始")
    old_wb = load_workbook(OLD_PATH)
    new_wb = load_workbook(NEW_PATH)

    print(f"  シート一覧: {old_wb.sheetnames}")

    # スコアサマリー
    merge_sheet(old_wb, new_wb, "スコアサマリー", "ケース")
    # 返信比較
    if "返信比較" in old_wb.sheetnames:
        # 返信比較のid列名を確認する必要があるがおそらく同じ
        try:
            merge_sheet(old_wb, new_wb, "返信比較", "ケースID")
        except ValueError:
            try:
                merge_sheet(old_wb, new_wb, "返信比較", "ケース")
            except Exception as e:
                print(f"  返信比較マージ失敗: {e}")

    old_wb.save(OUT_PATH)
    print(f"[2] 保存完了: {OUT_PATH}")


if __name__ == "__main__":
    main()
