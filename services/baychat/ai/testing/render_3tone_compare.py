"""
3トーン横並び比較HTMLレンダラー (POLITE / APOLOGY / ASSERTIVE)
====================================================================
同じバイヤーメッセージに対して、3つのトーンでどう違うかを横並びで比較。
"""
import json
import os
import sys
import io
import argparse
from datetime import datetime
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(SCRIPT_DIR, "results")


def load_excel_results(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    summary_ws = wb["スコアサマリー"]
    replies_ws = wb["返信比較"]

    score_data = {}
    for row in summary_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        case_id, model, prompt, accuracy, tone, completeness, action, naturalness, total, elapsed, cost, summary = row[:12]
        score_data[(case_id, model)] = {
            "score_total": total,
            "summary_ja": summary or "",
            "elapsed": elapsed,
            "cost_jpy": cost,
        }

    reply_data = {}
    for row in replies_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        case_id, buyer_msg, model, prompt, buyer_reply, jpn_reply, total = row[:7]
        reply_data[(case_id, model)] = {
            "buyer_reply": buyer_reply or "",
            "jpn_reply": jpn_reply or "",
        }

    merged = {}
    for key in score_data:
        merged[key] = {**score_data[key], **reply_data.get(key, {})}
    return merged


def load_cases(cases_json_path):
    with open(cases_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return {c["id"]: c for c in data}


def html_escape(text):
    if not text:
        return ""
    return (str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<title>BayChat AI Reply — 3トーン横並び比較 (POLITE / APOLOGY / ASSERTIVE)</title>
<style>
  * { box-sizing: border-box; }
  body {
    font-family: -apple-system, "Hiragino Sans", "Yu Gothic UI", "Meiryo", sans-serif;
    background: #f3f4f6;
    margin: 0;
    padding: 24px;
    color: #1f2937;
    line-height: 1.7;
  }
  h1 { font-size: 22px; color: #6b21a8; margin: 0 0 8px 0; }
  h2 { font-size: 17px; color: #374151; margin: 30px 0 8px 0; padding-bottom: 6px; border-bottom: 2px solid #e5e7eb; }
  h3 { font-size: 14px; color: #6b7280; margin: 12px 0 4px 0; font-weight: 600; }
  .container { max-width: 1900px; margin: 0 auto; }
  .meta { font-size: 12px; color: #6b7280; margin-bottom: 18px; }
  .legend {
    background: #fff;
    border: 1px solid #e5e7eb;
    border-radius: 10px;
    padding: 14px 18px;
    margin-bottom: 24px;
    font-size: 13px;
  }
  .legend strong { color: #6b21a8; }
  .legend table { width: 100%; margin-top: 8px; border-collapse: collapse; }
  .legend td { padding: 5px 10px; border: 1px solid #e5e7eb; vertical-align: top; }
  .legend td:first-child { width: 130px; font-weight: 600; }
  .case-card {
    background: #fff;
    border: 1px solid #e5e7eb;
    border-radius: 12px;
    padding: 20px;
    margin-bottom: 24px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04);
  }
  .buyer-msg {
    background: #faf5ff;
    border-left: 4px solid #6d28d9;
    padding: 12px 16px;
    border-radius: 6px;
    margin: 10px 0 16px 0;
    font-size: 13.5px;
  }
  .buyer-msg .label { font-size: 11px; color: #6b21a8; font-weight: 600; margin-bottom: 4px; }
  .scenario-note {
    background: #fffbeb;
    border-left: 4px solid #f59e0b;
    padding: 8px 14px;
    font-size: 12.5px;
    color: #78350f;
    border-radius: 4px;
    margin: 8px 0 12px 0;
  }
  .three-col {
    display: grid;
    grid-template-columns: 1fr 1fr 1fr;
    gap: 12px;
    margin-top: 8px;
  }
  .col {
    border: 1px solid #e5e7eb;
    border-radius: 8px;
    padding: 12px;
    background: #fafafa;
    min-width: 0;
  }
  .col.POLITE { border-top: 4px solid #3b82f6; }
  .col.APOLOGY { border-top: 4px solid #10b981; }
  .col.ASSERTIVE { border-top: 4px solid #dc2626; }
  .col-header {
    font-size: 13px;
    font-weight: 700;
    margin-bottom: 8px;
  }
  .col.POLITE .col-header { color: #1d4ed8; }
  .col.APOLOGY .col-header { color: #047857; }
  .col.ASSERTIVE .col-header { color: #b91c1c; }
  .model-block {
    background: #fff;
    border: 1px solid #e5e7eb;
    border-radius: 6px;
    padding: 10px 12px;
    margin-bottom: 10px;
    font-size: 12.5px;
  }
  .model-name {
    font-size: 11px;
    color: #6b7280;
    font-weight: 600;
    margin-bottom: 6px;
  }
  .reply-en {
    white-space: pre-wrap;
    word-wrap: break-word;
    font-family: -apple-system, sans-serif;
    font-size: 12px;
    color: #1f2937;
    background: #fff;
    border-radius: 4px;
    padding: 8px;
    border: 1px solid #f3f4f6;
    margin-bottom: 6px;
    line-height: 1.6;
  }
  .reply-ja {
    white-space: pre-wrap;
    word-wrap: break-word;
    font-size: 11.5px;
    color: #4b5563;
    padding: 6px 8px;
    background: #f9fafb;
    border-radius: 4px;
    border-left: 2px solid #c4b5fd;
    line-height: 1.6;
  }
  .meta-row {
    display: flex;
    gap: 8px;
    margin-top: 6px;
    font-size: 10.5px;
    color: #6b7280;
    flex-wrap: wrap;
  }
  .badge {
    background: #ede9fe;
    color: #5b21b6;
    padding: 1px 6px;
    border-radius: 3px;
    font-weight: 600;
  }
  .badge.elapsed { background: #fef3c7; color: #92400e; }
  .badge.cost { background: #dbeafe; color: #1e40af; }
  .judge-summary {
    background: #fef9c3;
    border-left: 3px solid #ca8a04;
    padding: 6px 10px;
    margin-top: 6px;
    font-size: 11px;
    color: #713f12;
    border-radius: 3px;
  }
  .case-meta {
    font-size: 12px;
    color: #6b7280;
    margin-bottom: 8px;
  }
  .case-id {
    font-weight: 700;
    color: #6b21a8;
  }
</style>
</head>
<body>
<div class="container">
  <h1>🎭 BayChat AI Reply — 3トーン横並び比較</h1>
  <p class="meta">テスト走行: {timestamp} / プロンプト: v2.3_baseline_natural4_principle (iter10)</p>

  <div class="legend">
    <strong>3トーンの定義</strong>
    <table>
      <tr>
        <td style="background:#dbeafe;">丁寧 (POLITE)</td>
        <td>標準的なCSビジネス文体。「Thank you very much」「We appreciate」「Kindly」など。Close: "Best regards," / "Kind regards,"</td>
      </tr>
      <tr>
        <td style="background:#d1fae5;">謝罪 (APOLOGY)</td>
        <td>POLITE より格式高く深い共感。「We sincerely apologize」「We understand how disappointing」など。Close: "Sincerely," / "With our apologies,"</td>
      </tr>
      <tr>
        <td style="background:#fee2e2;">主張 (ASSERTIVE)</td>
        <td>礼儀正しく毅然。謝罪語(sorry/apologize/regret/unfortunately)・感情共感フレーズ全廃。eBayプラットフォーム/国際取引慣行の権威を借りる。Close: "Best regards," / "Kind regards,"</td>
      </tr>
    </table>
    <p style="margin-top:8px; font-size:12px; color:#6b7280;">
      <strong>確認観点</strong>: 同じクレームに対して3トーンがどう違うか。トーンによる事実情報の変化はNG（fact-invariant）。voice/form のみ変化すべし。
    </p>
  </div>

{cases_html}

</div>
</body>
</html>"""


def render(cases_path, xlsx_polite, xlsx_apology, xlsx_assertive, output_path):
    cases = load_cases(cases_path)
    res_p = load_excel_results(xlsx_polite)
    res_a = load_excel_results(xlsx_apology)
    res_s = load_excel_results(xlsx_assertive)

    cases_html = []
    for case_id, case in cases.items():
        buyer_msg = case.get("buyer_message", "")
        buyer_msg_ja = case.get("buyer_message_ja", "")
        situation = case.get("situation", "")
        emotion = case.get("emotion", "")

        models = sorted(set(model for (cid, model) in res_p.keys() if cid == case_id))

        col_html = {"POLITE": [], "APOLOGY": [], "ASSERTIVE": []}
        for model in models:
            for label, results in [("POLITE", res_p), ("APOLOGY", res_a), ("ASSERTIVE", res_s)]:
                r = results.get((case_id, model), {})
                buyer_reply = (r.get("buyer_reply") or "").strip()
                jpn_reply = (r.get("jpn_reply") or "").strip()
                score = r.get("score_total", "?")
                summary = (r.get("summary_ja") or "").strip()
                elapsed = r.get("elapsed", 0)
                cost = r.get("cost_jpy", 0)

                block = f"""<div class="model-block">
  <div class="model-name">📦 {model}</div>
  <div class="reply-en">{html_escape(buyer_reply)}</div>
  <div class="reply-ja">{html_escape(jpn_reply)}</div>
  <div class="meta-row">
    <span class="badge">スコア {score}/25</span>
    <span class="badge elapsed">{elapsed:.1f}秒</span>
    <span class="badge cost">¥{cost:.3f}</span>
  </div>
  {f'<div class="judge-summary">{html_escape(summary)}</div>' if summary else ''}
</div>"""
                col_html[label].append(block)

        case_html = f"""<div class="case-card">
  <div class="case-meta">
    <span class="case-id">{case_id}</span> | 感情: {html_escape(emotion)}
  </div>
  <div class="scenario-note">{html_escape(situation)}</div>
  <div class="buyer-msg">
    <div class="label">バイヤーメッセージ (英)</div>
    {html_escape(buyer_msg)}
    <div class="label" style="margin-top:6px;">日本語訳</div>
    {html_escape(buyer_msg_ja)}
  </div>
  <div class="three-col">
    <div class="col POLITE">
      <div class="col-header">🔵 丁寧 (POLITE)</div>
      {''.join(col_html['POLITE'])}
    </div>
    <div class="col APOLOGY">
      <div class="col-header">🟢 謝罪 (APOLOGY)</div>
      {''.join(col_html['APOLOGY'])}
    </div>
    <div class="col ASSERTIVE">
      <div class="col-header">🔴 主張 (ASSERTIVE)</div>
      {''.join(col_html['ASSERTIVE'])}
    </div>
  </div>
</div>"""
        cases_html.append(case_html)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    html = HTML_TEMPLATE.replace("{timestamp}", timestamp).replace("{cases_html}", "\n".join(cases_html))

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"HTML生成完了: {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--cases", required=True)
    parser.add_argument("--polite", required=True, help="POLITE Excel")
    parser.add_argument("--apology", required=True, help="APOLOGY Excel")
    parser.add_argument("--assertive", required=True, help="ASSERTIVE Excel")
    parser.add_argument("--output")
    args = parser.parse_args()

    if not args.output:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = os.path.join(RESULTS_DIR, f"compare_3tone_{ts}.html")

    render(args.cases, args.polite, args.apology, args.assertive, args.output)
