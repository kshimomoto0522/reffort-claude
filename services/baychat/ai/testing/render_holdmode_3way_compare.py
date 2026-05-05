"""
保留モード必要性検証 — 3パターン横並び比較HTMLレンダラー
====================================================================
A: 即答（補足なし）   B: 補足「現在確認中…」   C: 保留モード ON
の 3 パターンを同じケースで横並びに表示する。
"""
import json
import os
import sys
import io
from datetime import datetime
import openpyxl

# Windows 環境の文字化け対策
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(SCRIPT_DIR, "results")


def load_excel_results(xlsx_path):
    """batch_test の出力 Excel から結果を辞書化する。
    返り値: {(case_id, model_name): {"buyer_reply": "", "jpn_reply": "", "score_total": 0, "summary_ja": "", "elapsed": 0}}"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    summary_ws = wb["スコアサマリー"]
    replies_ws = wb["返信比較"]

    # スコアサマリーから case_id/model/score 取得
    score_data = {}
    for row in summary_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        case_id, model, prompt, accuracy, tone, completeness, action, naturalness, total, elapsed, cost, summary = row[:12]
        score_data[(case_id, model)] = {
            "score_total": total,
            "summary_ja": summary or "",
            "elapsed": elapsed,
            "cost_jpy": cost,
        }

    # 返信比較から buyer_reply/jpn_reply 取得
    reply_data = {}
    for row in replies_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        case_id, buyer_msg, model, prompt, buyer_reply, jpn_reply, total = row[:7]
        reply_data[(case_id, model)] = {
            "buyer_reply": buyer_reply or "",
            "jpn_reply": jpn_reply or "",
        }

    # マージ
    merged = {}
    for key in score_data:
        merged[key] = {**score_data[key], **reply_data.get(key, {})}
    return merged


def load_cases(cases_json_path):
    """テストケース JSON を読み込んで {case_id: case_dict} を返す"""
    with open(cases_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return {c["id"]: c for c in data}


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<title>保留モード必要性検証 — 3パターン比較</title>
<style>
  * { box-sizing: border-box; }
  body {
    font-family: -apple-system, "Hiragino Sans", "Yu Gothic UI", "Meiryo", sans-serif;
    background: #f3f4f6;
    margin: 0;
    padding: 24px;
    color: #1f2937;
    line-height: 1.7;
  }
  h1 { font-size: 22px; color: #6b21a8; margin: 0 0 8px 0; }
  h2 { font-size: 17px; color: #374151; margin: 30px 0 8px 0; padding-bottom: 6px; border-bottom: 2px solid #e5e7eb; }
  h3 { font-size: 14px; color: #6b7280; margin: 12px 0 4px 0; font-weight: 600; }
  .container { max-width: 1700px; margin: 0 auto; }
  .meta { font-size: 12px; color: #6b7280; margin-bottom: 18px; }
  .legend {
    background: #fff;
    border: 1px solid #e5e7eb;
    border-radius: 10px;
    padding: 14px 18px;
    margin-bottom: 24px;
    font-size: 13px;
  }
  .legend strong { color: #6b21a8; }
  .legend table { width: 100%; margin-top: 8px; border-collapse: collapse; }
  .legend td { padding: 5px 10px; border: 1px solid #e5e7eb; vertical-align: top; }
  .legend td:first-child { width: 100px; font-weight: 600; }
  .case-card {
    background: #fff;
    border: 1px solid #e5e7eb;
    border-radius: 12px;
    padding: 20px;
    margin-bottom: 24px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04);
  }
  .buyer-msg {
    background: #faf5ff;
    border-left: 4px solid #6d28d9;
    padding: 12px 16px;
    border-radius: 6px;
    margin: 10px 0 16px 0;
    font-size: 13.5px;
  }
  .buyer-msg .label { font-size: 11px; color: #6b21a8; font-weight: 600; margin-bottom: 4px; }
  .scenario-note {
    background: #fffbeb;
    border-left: 4px solid #f59e0b;
    padding: 8px 14px;
    font-size: 12.5px;
    color: #78350f;
    border-radius: 4px;
    margin: 8px 0 12px 0;
  }
  .three-col {
    display: grid;
    grid-template-columns: 1fr 1fr 1fr;
    gap: 12px;
    margin-top: 8px;
  }
  .col {
    border: 1px solid #e5e7eb;
    border-radius: 8px;
    padding: 12px;
    background: #fafafa;
    min-width: 0;
  }
  .col.A { border-top: 4px solid #3b82f6; }
  .col.B { border-top: 4px solid #f59e0b; }
  .col.C { border-top: 4px solid #10b981; }
  .col-header {
    font-size: 12px;
    font-weight: 700;
    margin-bottom: 8px;
  }
  .col.A .col-header { color: #1d4ed8; }
  .col.B .col-header { color: #b45309; }
  .col.C .col-header { color: #047857; }
  .model-block {
    background: #fff;
    border: 1px solid #e5e7eb;
    border-radius: 6px;
    padding: 10px 12px;
    margin-bottom: 10px;
    font-size: 12.5px;
  }
  .model-name {
    font-size: 11px;
    color: #6b7280;
    font-weight: 600;
    margin-bottom: 6px;
  }
  .reply-en {
    white-space: pre-wrap;
    word-wrap: break-word;
    font-family: -apple-system, sans-serif;
    font-size: 12px;
    color: #1f2937;
    background: #fff;
    border-radius: 4px;
    padding: 8px;
    border: 1px solid #f3f4f6;
    margin-bottom: 6px;
    line-height: 1.6;
  }
  .reply-ja {
    white-space: pre-wrap;
    word-wrap: break-word;
    font-size: 11.5px;
    color: #4b5563;
    padding: 6px 8px;
    background: #f9fafb;
    border-radius: 4px;
    border-left: 2px solid #c4b5fd;
    line-height: 1.6;
  }
  .meta-row {
    display: flex;
    gap: 8px;
    margin-top: 6px;
    font-size: 10.5px;
    color: #6b7280;
    flex-wrap: wrap;
  }
  .badge {
    background: #ede9fe;
    color: #5b21b6;
    padding: 1px 6px;
    border-radius: 3px;
    font-weight: 600;
  }
  .badge.elapsed { background: #fef3c7; color: #92400e; }
  .badge.cost { background: #dbeafe; color: #1e40af; }
  .judge-summary {
    background: #fef9c3;
    border-left: 3px solid #ca8a04;
    padding: 6px 10px;
    margin-top: 6px;
    font-size: 11px;
    color: #713f12;
    border-radius: 3px;
  }
</style>
</head>
<body>
<div class="container">
  <h1>🔍 保留モード必要性検証 — 3パターン横並び比較</h1>
  <p class="meta">テスト走行: {timestamp}</p>

  <div class="legend">
    <strong>3パターンの違い</strong>
    <table>
      <tr>
        <td style="background:#dbeafe;">A. 即答（補足なし）</td>
        <td>通常生成。バイヤー質問に AI が即答する。</td>
      </tr>
      <tr>
        <td style="background:#fef3c7;">B. 補足「確認中…」</td>
        <td>「現在確認中のため、後ほど改めてご連絡します」を補足欄に入力して生成。</td>
      </tr>
      <tr>
        <td style="background:#d1fae5;">C. 保留モード ON</td>
        <td>HOLD_MODE_BLOCK を末尾挿入。専用ガイドで「実質的な質問に答えない・確認後連絡する旨だけ返す」。</td>
      </tr>
    </table>
    <p style="margin-top:8px; font-size:12px; color:#6b7280;">
      <strong>検証観点</strong>: B（補足だけ）で十分か? それとも C（専用モード）でないと達成できない違いがあるか? を実生成結果で社長判断していただく。
    </p>
  </div>

{cases_html}

</div>
</body>
</html>"""


def render(cases_path, xlsx_a, xlsx_b, xlsx_c, output_path):
    cases = load_cases(cases_path)
    res_a = load_excel_results(xlsx_a)
    res_b = load_excel_results(xlsx_b)
    res_c = load_excel_results(xlsx_c)

    cases_html = []
    for case_id, case in cases.items():
        buyer_msg = case.get("buyer_message", "")
        buyer_msg_ja = case.get("buyer_message_ja", "")
        situation = case.get("situation", "")

        # モデル一覧（res_a に存在するものを使う）
        models = sorted(set(model for (cid, model) in res_a.keys() if cid == case_id))

        # 3列のHTML 生成
        col_html = {"A": [], "B": [], "C": []}
        for model in models:
            for label, results in [("A", res_a), ("B", res_b), ("C", res_c)]:
                r = results.get((case_id, model), {})
                buyer_reply = (r.get("buyer_reply") or "").strip()
                jpn_reply = (r.get("jpn_reply") or "").strip()
                score = r.get("score_total", "?")
                summary = (r.get("summary_ja") or "").strip()
                elapsed = r.get("elapsed", 0)
                cost = r.get("cost_jpy", 0)

                block = f"""<div class="model-block">
  <div class="model-name">📦 {model}</div>
  <div class="reply-en">{html_escape(buyer_reply)}</div>
  <div class="reply-ja">{html_escape(jpn_reply)}</div>
  <div class="meta-row">
    <span class="badge">スコア {score}/25</span>
    <span class="badge elapsed">⏱ {elapsed}s</span>
    <span class="badge cost">¥{cost}</span>
  </div>
  {f'<div class="judge-summary">📝 {html_escape(summary)}</div>' if summary else ''}
</div>"""
                col_html[label].append(block)

        case_html = f"""
<div class="case-card">
  <h2>{html_escape(case_id)}</h2>
  <div class="scenario-note">📌 シーン: {html_escape(situation)}</div>
  <div class="buyer-msg">
    <div class="label">バイヤーメッセージ</div>
    <div>{html_escape(buyer_msg)}</div>
    <div style="color:#6b7280; font-size:12px; margin-top:4px;">{html_escape(buyer_msg_ja)}</div>
  </div>

  <div class="three-col">
    <div class="col A">
      <div class="col-header">A. 即答（補足なし）</div>
      {''.join(col_html['A'])}
    </div>
    <div class="col B">
      <div class="col-header">B. 補足「確認中のため、後ほど…」</div>
      {''.join(col_html['B'])}
    </div>
    <div class="col C">
      <div class="col-header">C. 保留モード ON（HOLD_MODE_BLOCK 注入）</div>
      {''.join(col_html['C'])}
    </div>
  </div>
</div>"""
        cases_html.append(case_html)

    html = HTML_TEMPLATE.replace(
        "{timestamp}", datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ).replace(
        "{cases_html}", "\n".join(cases_html)
    )
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"HTML 生成完了: {output_path}")


def html_escape(s):
    if s is None:
        return ""
    return (str(s)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;"))


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--cases", required=True)
    p.add_argument("--a", required=True, help="パターンA Excel")
    p.add_argument("--b", required=True, help="パターンB Excel")
    p.add_argument("--c", required=True, help="パターンC Excel")
    p.add_argument("--out", default=None)
    args = p.parse_args()

    if args.out is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.out = os.path.join(RESULTS_DIR, f"holdmode_3way_compare_{ts}.html")

    render(args.cases, args.a, args.b, args.c, args.out)
