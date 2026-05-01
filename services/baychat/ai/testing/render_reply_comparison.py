"""
AI Reply モデル比較結果レンダラー（v2: feedback textarea + 日本語訳表示対応）
============================================================================
BayChatブランドカラー（紫ベース）・英日併記・シンプルイズザベスト。

新機能:
  - バイヤーメッセージの日本語訳を必ず表示（test_casesから読み込み）
  - 各ケース下に「社長の理想フィードバック」textarea
  - localStorage自動保存
  - 最下部に「全フィードバックをJSONダウンロード」ボタン
  - 速度逆転ケースの分析セクション

使い方:
    python render_reply_comparison.py --excel results/test_xxx.xlsx --cases test_cases/xxx.json
"""

import os
import sys
import io
import json
import argparse
from datetime import datetime
from collections import defaultdict
import html as html_lib

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(SCRIPT_DIR, "results")

# ============================================================
# BayChat ブランドカラー（紫ベース）
# ============================================================
COLOR_PRIMARY = "#7c3aed"
COLOR_PRIMARY_DARK = "#5b21b6"
COLOR_PRIMARY_LIGHT = "#a78bfa"
COLOR_PRIMARY_BG = "#f5f3ff"
COLOR_PRIMARY_BG2 = "#ede9fe"
COLOR_TEXT = "#1f2937"
COLOR_TEXT_SUB = "#6b7280"
COLOR_BORDER = "#e5e7eb"


def esc(text):
    if text is None:
        return ""
    return html_lib.escape(str(text))


def read_excel(excel_path):
    """openpyxlでExcelを読み込み、全行をdictリスト化"""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    results = []
    ws = wb["スコアサマリー"]
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        results.append(dict(zip(headers, row)))

    ws_rep = wb["返信比較"]
    rep_headers = [cell.value for cell in ws_rep[1]]
    rep_rows = []
    for row in ws_rep.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rep_rows.append(dict(zip(rep_headers, row)))

    reply_map = {}
    for r in rep_rows:
        key = (r.get("ケース"), r.get("モデル"))
        reply_map[key] = {
            "buyer_message": r.get("バイヤーメッセージ") or "",
            "buyer_reply": r.get("英語返信") or "",
            "jpn_reply": r.get("日本語訳") or "",
        }

    for r in results:
        key = (r.get("ケース"), r.get("モデル"))
        if key in reply_map:
            r.update(reply_map[key])

    ws_agg = wb["モデル別集計"]
    agg_headers = [cell.value for cell in ws_agg[1]]
    agg_rows = []
    for row in ws_agg.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        agg_rows.append(dict(zip(agg_headers, row)))

    return results, agg_rows


def build_agg_table(agg_rows):
    """モデル別集計テーブル + 勝者バッジ"""
    if not agg_rows:
        return "<p>集計データなし</p>"

    scores = {r["モデル × プロンプト"]: r.get("平均スコア(/25)", 0) or 0 for r in agg_rows}
    times = {r["モデル × プロンプト"]: r.get("平均応答時間(秒)", 0) or 0 for r in agg_rows}
    costs = {r["モデル × プロンプト"]: r.get("平均コスト(¥)", 0) or 0 for r in agg_rows}

    best_score = max(scores, key=scores.get) if scores else None
    best_speed = min({k: v for k, v in times.items() if v > 0},
                     key=lambda k: times[k]) if any(times.values()) else None
    best_cost = min({k: v for k, v in costs.items() if v > 0},
                    key=lambda k: costs[k]) if any(costs.values()) else None

    table_rows = ""
    for r in agg_rows:
        m = r["モデル × プロンプト"]
        badges = ""
        if m == best_score: badges += ' <span class="badge-winner">品質◎</span>'
        if m == best_speed: badges += ' <span class="badge-winner">速度◎</span>'
        if m == best_cost: badges += ' <span class="badge-winner">コスト◎</span>'

        table_rows += f"""
        <tr>
            <td class="model-name"><strong>{esc(m)}</strong>{badges}</td>
            <td class="num">{r.get('テスト件数', 0)}</td>
            <td class="num score">{(r.get('平均スコア(/25)', 0) or 0):.1f}</td>
            <td class="num">{(r.get('平均正確性', 0) or 0):.1f}</td>
            <td class="num">{(r.get('平均トーン', 0) or 0):.1f}</td>
            <td class="num">{(r.get('平均完全性', 0) or 0):.1f}</td>
            <td class="num">{(r.get('平均アクション', 0) or 0):.1f}</td>
            <td class="num">{(r.get('平均自然さ', 0) or 0):.1f}</td>
            <td class="num">{(r.get('平均応答時間(秒)', 0) or 0):.1f}秒</td>
            <td class="num">¥{(r.get('平均コスト(¥)', 0) or 0):.3f}</td>
            <td class="num">¥{int(r.get('月額推定(¥)', 0) or 0):,}</td>
        </tr>"""

    return f"""
    <table class="agg-table">
        <thead>
            <tr>
                <th>モデル</th><th>件数</th><th>合計<br>(/25)</th>
                <th>正確<br>性</th><th>トーン</th><th>完全<br>性</th>
                <th>アクション</th><th>自然さ</th>
                <th>応答<br>時間</th><th>1回<br>コスト</th><th>月額<br>推定</th>
            </tr>
        </thead>
        <tbody>{table_rows}</tbody>
    </table>
    """


def build_category_matrix(results, cat_map, models):
    """カテゴリ × モデル の平均スコアマトリクス"""
    cat_scores = defaultdict(lambda: defaultdict(list))
    for r in results:
        cat = cat_map.get(r["ケース"], "不明")
        total = r.get("合計(/25)", 0) or 0
        if total > 0:
            cat_scores[cat][r["モデル"]].append(total)

    if not cat_scores:
        return ""

    header = "<tr><th>カテゴリ</th>" + "".join(f"<th>{esc(m)}</th>" for m in models) + "<th>勝者</th></tr>"
    body = ""
    for cat in ["cancel", "return", "claim", "tracking", "discount", "general"]:
        if cat not in cat_scores:
            continue
        avgs = {}
        row = f"<tr><td><strong>{esc(cat)}</strong></td>"
        for m in models:
            vals = cat_scores[cat].get(m, [])
            a = sum(vals) / len(vals) if vals else 0
            avgs[m] = a
            row += f'<td class="num">{a:.1f}<br><small>({len(vals)}件)</small></td>'
        winner = max(avgs, key=avgs.get) if avgs else "—"
        row += f'<td><strong style="color:{COLOR_PRIMARY_DARK}">{esc(winner)}</strong></td></tr>'
        body += row

    return f"""
    <h3 style="margin-top:28px">📂 カテゴリ別スコア比較</h3>
    <table class="agg-table">
        <thead>{header}</thead>
        <tbody>{body}</tbody>
    </table>
    """


def build_speed_analysis(results, models):
    """
    速度逆転分析セクション:
    各ケースで「もっとも速いモデル」と「もっとも遅いモデル」の差分を出す
    """
    by_case = defaultdict(dict)
    for r in results:
        by_case[r["ケース"]][r["モデル"]] = r.get("応答時間(秒)", 0) or 0

    diffs = []
    for case_id, mts in by_case.items():
        if len(mts) < 2:
            continue
        fastest = min(mts, key=mts.get)
        slowest = max(mts, key=mts.get)
        diff = mts[slowest] - mts[fastest]
        diffs.append((case_id, fastest, mts[fastest], slowest, mts[slowest], diff))

    diffs.sort(key=lambda x: -x[5])  # 差分大きい順

    rows = ""
    for case_id, fast, ft, slow, st, d in diffs[:10]:
        rows += f"""
        <tr>
            <td>{esc(case_id)}</td>
            <td>{esc(fast)} <small>({ft:.1f}秒)</small></td>
            <td>{esc(slow)} <small>({st:.1f}秒)</small></td>
            <td class="num"><strong>+{d:.1f}秒</strong></td>
        </tr>"""

    return f"""
    <h3 style="margin-top:28px">⏱ 速度ばらつき（差分TOP10）</h3>
    <p style="font-size:13px;color:{COLOR_TEXT_SUB}">
    同じケースでもモデルによって応答時間が大きく違うケースの一覧。
    「モデル平均」だけでは見えない外れ値の傾向がわかる。
    </p>
    <table class="agg-table">
        <thead><tr><th>ケース</th><th>最速モデル</th><th>最遅モデル</th><th>差分</th></tr></thead>
        <tbody>{rows}</tbody>
    </table>
    """


def build_conversation_history(input_messages, history_ja=None):
    """会話履歴をタイムライン形式でレンダリング（英語 + 日本語訳併記・役割ごとにラベル）
    history_ja = user/assistantメッセージの日本語訳リスト（出現順）。systemは除外。
    """
    if not input_messages:
        return ""
    history_ja = history_ja or []

    product_info_json = None

    # 最後のバイヤー(user)メッセージのインデックスを特定（= AIが返信する対象）
    last_user_idx = -1
    for i, msg in enumerate(input_messages):
        if msg.get("role") == "user":
            last_user_idx = i

    items_html = ""
    ua_counter = 0  # user/assistantメッセージの出現カウンタ（history_ja対応用）
    for i, msg in enumerate(input_messages):
        role = msg.get("role", "")
        content = msg.get("content", "") or ""

        # developerロール = プロンプト本体。1個目(商品情報JSON)だけ表示、以降は省略
        if role == "developer":
            if product_info_json is None and content.strip().startswith("{"):
                product_info_json = content
            continue

        if role == "system":
            items_html += f"""
            <div class="msg msg-system">
                <span class="msg-role">📦 eBayイベント</span>
                <div class="msg-body">{esc(content)}</div>
            </div>
            """
        elif role == "assistant":
            ja = history_ja[ua_counter] if ua_counter < len(history_ja) else ""
            ua_counter += 1
            ja_block = f'<div class="msg-body-ja"><span class="ja-label">🇯🇵</span> {esc(ja)}</div>' if ja else ""
            items_html += f"""
            <div class="msg msg-assistant">
                <span class="msg-role">🏪 セラー（過去の返信）</span>
                <div class="msg-body">{esc(content)}</div>
                {ja_block}
            </div>
            """
        elif role == "user":
            ja = history_ja[ua_counter] if ua_counter < len(history_ja) else ""
            ua_counter += 1
            is_last = (i == last_user_idx)
            extra_class = " msg-user-latest" if is_last else ""
            extra_label = " 🔻 AIが返信する対象メッセージ" if is_last else ""
            ja_block = f'<div class="msg-body-ja"><span class="ja-label">🇯🇵</span> {esc(ja)}</div>' if ja else ""
            items_html += f"""
            <div class="msg msg-user{extra_class}">
                <span class="msg-role">👤 バイヤー{extra_label}</span>
                <div class="msg-body">{esc(content)}</div>
                {ja_block}
            </div>
            """

    product_block = ""
    if product_info_json:
        try:
            info = json.loads(product_info_json)
            product_block = f"""
            <div class="product-info">
                <span class="product-label">🛒 商品情報（本番DBのログから抽出・eBay APIは未使用）</span>
                <div class="product-grid">
                    <div><b>Title:</b> {esc(info.get('Title', '—'))}</div>
                    <div><b>SKU:</b> {esc(info.get('SKU', '—'))}</div>
                    <div><b>Order ID:</b> {esc(info.get('orderID', '—'))}</div>
                    <div><b>Price:</b> {esc(info.get('CurrentPrice', '—'))} {esc(info.get('Currency', ''))}</div>
                </div>
            </div>
            """
        except Exception:
            pass

    return f"""
    <div class="history-block">
        <div class="history-header">💬 会話履歴全体（AIにはこの内容が全部渡されている）</div>
        {product_block}
        <div class="history-timeline">{items_html}</div>
    </div>
    """


def build_case_card(case_id, category, buyer_en, buyer_ja, rows_by_model, input_messages, winners, history_ja=None):
    """
    1ケース分のカード
    winners = dict {"score": モデル名, "speed": モデル名, "cost": モデル名}  ← このケースでの勝者
    history_ja = user/assistant各メッセージの日本語訳リスト（順序は input_messages 内の出現順）
    """
    history_ja = history_ja or []
    buyer_block = f"""
    <div class="buyer-block">
        <div class="block-label">📌 バイヤーの最新メッセージ（AIが返信すべき対象）</div>
        <div class="bilingual">
            <div class="en">
                <div class="lang-label">🇬🇧 バイヤー原文（英語）</div>
                <pre>{esc(buyer_en)}</pre>
            </div>
            <div class="ja">
                <div class="lang-label">🇯🇵 日本語訳（GPT-4.1-Mini自動翻訳）</div>
                <pre>{esc(buyer_ja) if buyer_ja else '（翻訳未生成）'}</pre>
            </div>
        </div>
    </div>
    """

    history_html = build_conversation_history(input_messages, history_ja=history_ja)

    # モデル名をdata属性として安全に使えるID化
    def _slug(s): return "".join(c if c.isalnum() else "-" for c in (s or ""))
    case_slug = _slug(case_id)

    reply_cards = ""
    for model_name, row in rows_by_model.items():
        total = row.get("合計(/25)", 0) or 0
        buyer_reply = row.get("buyer_reply", "") or ""
        jpn_reply = row.get("jpn_reply", "") or ""
        elapsed = row.get("応答時間(秒)", 0) or 0
        cost = row.get("コスト(¥)", 0) or 0
        summary = row.get("サマリー", "") or ""

        # このケース内での勝者バッジ
        case_badges = ""
        if model_name == winners.get("score"):
            case_badges += '<span class="case-winner-badge win-score">品質◎</span>'
        if model_name == winners.get("speed"):
            case_badges += '<span class="case-winner-badge win-speed">速度◎</span>'
        if model_name == winners.get("cost"):
            case_badges += '<span class="case-winner-badge win-cost">コスト◎</span>'

        scores = {
            "正確性": row.get("正確性", 0) or 0,
            "トーン": row.get("トーン", 0) or 0,
            "完全性": row.get("完全性", 0) or 0,
            "アクション": row.get("アクション", 0) or 0,
            "自然さ": row.get("自然さ", 0) or 0,
        }
        score_pills = "".join(
            f'<span class="score-pill">{esc(k)} <b>{v}</b>/5</span>'
            for k, v in scores.items()
        )

        reply_cards += f"""
        <div class="reply-card" data-case-id="{esc(case_id)}" data-model-name="{esc(model_name)}">
            <div class="reply-header">
                <div class="model-badge">{esc(model_name)}</div>
                <div class="case-winners">{case_badges}</div>
                <div class="score-total">
                    <span class="meta-elapsed">⏱ <b>{elapsed:.1f}</b>秒</span>
                    <span class="meta-cost">💴 <b>¥{cost:.3f}</b></span>
                </div>
            </div>
            <div class="bilingual">
                <div class="en">
                    <div class="lang-label">🇬🇧 AI返信（英語・バイヤー送信用）</div>
                    <pre class="reply-en">{esc(buyer_reply)}</pre>
                </div>
                <div class="ja">
                    <div class="lang-label">🇯🇵 日本語訳（セラー確認用）</div>
                    <pre class="reply-ja">{esc(jpn_reply)}</pre>
                </div>
            </div>
        </div>
        """

    supplemental_block = f"""
    <div class="supplemental-block" data-case-id="{esc(case_id)}">
        <label class="supplemental-label">
            🎯 補足情報（sellerSetting）<small>このケース全モデル共通枠。ここに「セラーが本当に伝えたい補足情報」を日本語で書いて「補足込みで再生成」を押すと、3モデル全部で再走します。空欄なら補足なし生成。</small>
        </label>
        <textarea class="supplemental-input"
                  data-case-id="{esc(case_id)}"
                  placeholder="例: 新品で未使用、保証書・原箱・papers全て同梱します。タグも付いています。"
                  rows="3"></textarea>
        <div class="supplemental-actions">
            <button class="regen-btn" onclick="regenerateCase('{esc(case_id)}', this)">
                🔁 補足込みで再生成（3モデル）
            </button>
            <span class="regen-status"></span>
        </div>
    </div>

    <div class="feedback-block">
        <label class="feedback-label">
            ✏️ 社長の理想フィードバック <small>(このケースで本当に欲しかった返信内容・NG理由・トーン指示等を日本語で書く。自動保存・最下部からJSONダウンロード可能)</small>
        </label>
        <textarea class="feedback-input"
                  data-case-id="{esc(case_id)}"
                  placeholder="例: 返品の件は『Have you already started the return process via eBay? If not, please open a return request so I can issue a return label』と具体的な手順を提示してほしい。謝罪は簡潔に1文でOK。"
                  rows="4"></textarea>
    </div>
    """
    # 旧コードとの互換のため変数名を維持
    feedback_textarea = supplemental_block

    return f"""
    <article class="case-card" data-case-id="{esc(case_id)}">
        <div class="case-header">
            <h3>📩 {esc(case_id)}</h3>
            <span class="category-pill">{esc(category)}</span>
        </div>
        {history_html}
        {buyer_block}
        <div class="reply-grid">{reply_cards}</div>
        {feedback_textarea}
    </article>
    """


def render_comparison_html(excel_path, test_cases_path, top_n=12, tone=None):
    """Excelから結果を読み込み、サマリー + 全ケース詳細 + feedbackのHTML生成

    tone: "polite" | "friendly" | "apologetic" | None
          指定するとHTML title / header / 出力ファイル名にトーンラベルを表示
    """
    # トーンラベル定義
    tone_label_map = {
        "polite": ("POLITE", "丁寧", "#5b21b6"),
        "friendly": ("FRIENDLY", "フレンドリー", "#0d9488"),
        "apologetic": ("APOLOGY", "謝罪", "#b45309"),
    }
    tone_info = tone_label_map.get(tone) if tone else None
    tone_en = tone_info[0] if tone_info else ""
    tone_ja = tone_info[1] if tone_info else ""
    tone_color = tone_info[2] if tone_info else "#5b21b6"
    tone_suffix_title = f" — トーン {tone_en} ({tone_ja})" if tone_info else ""

    results, agg_rows = read_excel(excel_path)

    # test_casesから追加情報を取得
    case_meta = {}
    if test_cases_path and os.path.exists(test_cases_path):
        with open(test_cases_path, "r", encoding="utf-8") as f:
            cases = json.load(f)
        for c in cases:
            case_meta[c["id"]] = {
                "category": c.get("category", "不明"),
                "buyer_message_en": c.get("buyer_message", ""),
                "buyer_message_ja": c.get("buyer_message_ja", ""),
                "input": c.get("input", []),  # 会話履歴全体
                "history_ja": c.get("history_ja", []),  # user/assistant各メッセージの日本語訳
            }

    cat_map = {cid: m["category"] for cid, m in case_meta.items()}

    # モデル一覧
    models = sorted(set(r["モデル"] for r in results))

    # ケースごとにグルーピング
    by_case = defaultdict(dict)
    for r in results:
        by_case[r["ケース"]][r["モデル"]] = r

    # 詳細表示はすべてのケース（top_n=12）
    all_case_ids = sorted(by_case.keys())
    selected_cases = all_case_ids[:top_n]

    # サマリーセクション構築
    agg_html = build_agg_table(agg_rows)
    cat_html = build_category_matrix(results, cat_map, models)
    speed_html = build_speed_analysis(results, models)

    case_cards_html = ""
    for case_id in selected_cases:
        meta = case_meta.get(case_id, {})
        # このケース内での勝者を算出
        rows_by_model = by_case[case_id]
        case_scores = {m: (r.get("合計(/25)", 0) or 0) for m, r in rows_by_model.items()}
        case_times = {m: (r.get("応答時間(秒)", 0) or 999) for m, r in rows_by_model.items()}
        case_costs = {m: (r.get("コスト(¥)", 0) or 999) for m, r in rows_by_model.items()}
        winners = {
            "score": max(case_scores, key=case_scores.get) if case_scores else None,
            "speed": min(case_times, key=case_times.get) if case_times else None,
            "cost": min(case_costs, key=case_costs.get) if case_costs else None,
        }
        case_cards_html += build_case_card(
            case_id,
            meta.get("category", "不明"),
            meta.get("buyer_message_en", ""),
            meta.get("buyer_message_ja", ""),
            rows_by_model,
            meta.get("input", []),
            winners,
            history_ja=meta.get("history_ja", []),
        )

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<title>BayChat AI Reply モデル比較レポート{tone_suffix_title}</title>
<style>
    body {{
        font-family: 'Yu Gothic UI', 'Segoe UI', 'Hiragino Kaku Gothic ProN', sans-serif;
        background: {COLOR_PRIMARY_BG};
        color: {COLOR_TEXT};
        margin: 0; padding: 0;
    }}
    .container {{ max-width: 1600px; margin: 0 auto; padding: 32px; }}
    header {{
        background: linear-gradient(135deg, {COLOR_PRIMARY} 0%, {COLOR_PRIMARY_DARK} 100%);
        color: white; padding: 28px 32px; border-radius: 16px; margin-bottom: 24px;
        box-shadow: 0 6px 20px rgba(124,58,237,0.25);
    }}
    header h1 {{ margin: 0 0 4px; font-size: 24px; }}
    header .subtitle {{ opacity: 0.88; font-size: 14px; }}

    h2 {{
        color: {COLOR_PRIMARY_DARK};
        border-bottom: 3px solid {COLOR_PRIMARY};
        padding-bottom: 6px; margin-top: 32px;
    }}
    h3 {{ color: {COLOR_PRIMARY_DARK}; }}

    .summary-section {{
        background: white; padding: 24px; border-radius: 12px;
        border: 1px solid {COLOR_BORDER};
    }}

    table.agg-table {{
        width: 100%; border-collapse: collapse; margin-top: 12px;
        font-size: 13px;
    }}
    table.agg-table th, table.agg-table td {{
        border: 1px solid {COLOR_BORDER}; padding: 8px 10px;
        text-align: left; vertical-align: middle;
    }}
    table.agg-table th {{
        background: {COLOR_PRIMARY_BG2}; color: {COLOR_PRIMARY_DARK};
        font-weight: 600; text-align: center;
    }}
    table.agg-table td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
    table.agg-table td.num.score {{ font-weight: 700; color: {COLOR_PRIMARY_DARK}; font-size: 15px; }}
    table.agg-table td.model-name {{ white-space: nowrap; }}

    .badge-winner {{
        background: {COLOR_PRIMARY}; color: white; padding: 2px 8px;
        border-radius: 10px; font-size: 11px; margin-left: 4px;
    }}

    /* 補足情報 + 再生成ボタン */
    .supplemental-block {{
        background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%);
        border: 2px solid #f59e0b;
        border-radius: 10px; padding: 16px; margin-top: 16px;
    }}
    .supplemental-label {{
        display: block; font-weight: 700; color: #78350f;
        margin-bottom: 8px; font-size: 13px;
    }}
    .supplemental-label small {{ font-weight: 400; color: #92400e; }}
    .supplemental-input {{
        width: 100%; box-sizing: border-box;
        border: 1px solid #f59e0b; border-radius: 6px;
        padding: 10px 12px; font-size: 13px;
        font-family: inherit; resize: vertical;
        background: white;
    }}
    .supplemental-input:focus {{
        outline: none; border-color: #d97706;
        box-shadow: 0 0 0 3px rgba(245,158,11,0.2);
    }}
    .supplemental-actions {{
        margin-top: 10px; display: flex; align-items: center; gap: 12px;
    }}
    .regen-btn {{
        background: #d97706; color: white; border: none;
        padding: 10px 18px; border-radius: 8px; font-weight: 700;
        font-size: 14px; cursor: pointer;
        transition: all 0.2s;
        box-shadow: 0 2px 6px rgba(217,119,6,0.3);
    }}
    .regen-btn:hover {{
        background: #b45309; transform: translateY(-1px);
        box-shadow: 0 4px 10px rgba(217,119,6,0.4);
    }}
    .regen-btn:disabled {{
        background: #9ca3af; cursor: wait;
        transform: none; box-shadow: none;
    }}
    .regen-status {{
        font-size: 13px; color: #78350f; font-weight: 600;
    }}
    .regen-status.error {{ color: #dc2626; }}
    .regen-status.success {{ color: #15803d; }}

    /* 再生成中の reply-card のオーバーレイ */
    .reply-card.regenerating {{ position: relative; opacity: 0.6; }}
    .reply-card.regenerating::after {{
        content: "再生成中..."; position: absolute;
        top: 50%; left: 50%; transform: translate(-50%, -50%);
        background: {COLOR_PRIMARY}; color: white;
        padding: 8px 16px; border-radius: 8px; font-weight: 700;
    }}
    .reply-card.regenerated {{
        animation: highlight 1.5s ease-out;
    }}
    @keyframes highlight {{
        0% {{ background: #fef3c7; }}
        100% {{ background: white; }}
    }}
    .meta-elapsed.fast b {{ color: #15803d; }}
    .meta-elapsed.slow b {{ color: #dc2626; }}

    .case-card {{
        background: white; margin-top: 20px; padding: 20px;
        border-radius: 12px; border: 1px solid {COLOR_BORDER};
        box-shadow: 0 2px 6px rgba(0,0,0,0.04);
    }}
    .case-header {{ display: flex; align-items: center; gap: 10px; margin-bottom: 12px; }}
    .case-header h3 {{ margin: 0; font-size: 16px; }}
    .category-pill {{
        background: {COLOR_PRIMARY_BG2}; color: {COLOR_PRIMARY_DARK};
        padding: 3px 10px; border-radius: 10px; font-size: 12px;
    }}

    .buyer-block {{
        background: {COLOR_PRIMARY_BG}; padding: 14px;
        border-left: 4px solid {COLOR_PRIMARY}; border-radius: 8px;
        margin-bottom: 14px;
    }}
    .block-label {{
        font-size: 12px; font-weight: 700; color: {COLOR_PRIMARY_DARK};
        margin-bottom: 8px;
    }}

    /* 会話履歴全体のタイムライン */
    .history-block {{
        margin-bottom: 16px; background: #fafafa;
        border: 1px solid {COLOR_BORDER}; border-radius: 8px; padding: 14px;
    }}
    .history-header {{
        font-weight: 700; color: {COLOR_PRIMARY_DARK};
        font-size: 13px; padding: 2px 0 8px;
        border-bottom: 1px solid {COLOR_BORDER};
        margin-bottom: 8px;
    }}
    /* タイムライン：最大450pxでスクロール表示 */
    .history-timeline {{
        max-height: 450px;
        overflow-y: auto;
        padding: 10px 4px 10px 4px;
        background: white;
        border-radius: 6px;
        border: 1px solid {COLOR_BORDER};
    }}
    .history-timeline::-webkit-scrollbar {{
        width: 10px;
    }}
    .history-timeline::-webkit-scrollbar-thumb {{
        background: {COLOR_PRIMARY_LIGHT};
        border-radius: 5px;
    }}
    .history-timeline::-webkit-scrollbar-track {{
        background: {COLOR_PRIMARY_BG};
    }}
    .product-info {{
        background: white; border: 1px solid {COLOR_BORDER};
        border-radius: 6px; padding: 10px 12px; margin: 10px 0;
        font-size: 12px;
    }}
    .product-label {{
        display: block; color: {COLOR_PRIMARY_DARK};
        font-weight: 700; margin-bottom: 6px; font-size: 11px;
    }}
    .product-grid {{
        display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
        gap: 6px;
    }}
    .msg {{
        margin: 6px 0; padding: 8px 12px;
        border-radius: 8px; font-size: 12px;
    }}
    .msg-role {{
        display: inline-block; font-weight: 700; font-size: 11px;
        margin-bottom: 4px;
    }}
    .msg-body {{
        white-space: pre-wrap; word-break: break-word;
        font-size: 12px; line-height: 1.55;
    }}
    .msg-system {{
        background: #eef2ff; border-left: 3px solid #6366f1; color: #4338ca;
        font-size: 11px;
    }}
    .msg-assistant {{
        background: #f0fdf4; border-left: 3px solid #22c55e;
        margin-left: 30px;  /* セラー側はインデント */
    }}
    .msg-assistant .msg-role {{ color: #15803d; }}
    .msg-user {{
        background: {COLOR_PRIMARY_BG2}; border-left: 3px solid {COLOR_PRIMARY};
    }}
    .msg-user .msg-role {{ color: {COLOR_PRIMARY_DARK}; }}
    /* 最終バイヤーメッセージは強調 */
    .msg-user-latest {{
        background: #fef3c7 !important;
        border-left: 4px solid #f59e0b !important;
        box-shadow: 0 0 0 2px #fbbf24 inset;
    }}
    .msg-user-latest .msg-role {{ color: #b45309 !important; }}
    /* 履歴の日本語訳ブロック（各メッセージの下に小さく表示） */
    .msg-body-ja {{
        margin-top: 6px; padding-top: 6px;
        border-top: 1px dashed {COLOR_BORDER};
        font-size: 11px; line-height: 1.5;
        color: {COLOR_TEXT_SUB};
        white-space: pre-wrap; word-break: break-word;
    }}
    .msg-body-ja .ja-label {{
        display: inline-block; font-size: 10px;
        color: {COLOR_PRIMARY_DARK}; font-weight: 700;
        margin-right: 4px;
    }}

    /* ケース内勝者バッジ */
    .case-winners {{ display: flex; gap: 4px; }}
    .case-winner-badge {{
        padding: 2px 8px; border-radius: 10px;
        font-size: 10px; font-weight: 700; color: white;
    }}
    .win-score {{ background: #7c3aed; }}
    .win-speed {{ background: #0ea5e9; }}
    .win-cost {{ background: #f59e0b; }}

    .bilingual {{
        display: grid; grid-template-columns: 1fr; gap: 12px;
    }}
    .bilingual pre {{
        background: white; border: 1px solid {COLOR_BORDER};
        border-radius: 6px; padding: 10px; margin: 4px 0 0;
        white-space: pre-wrap; word-break: break-word; font-size: 13px;
        font-family: inherit; max-height: 260px; overflow-y: auto;
    }}
    .lang-label {{
        font-size: 11px; color: {COLOR_TEXT_SUB}; font-weight: 600;
    }}

    .reply-grid {{
        display: grid; grid-template-columns: repeat(auto-fit, minmax(480px, 1fr));
        gap: 12px;
    }}
    .reply-card {{
        background: {COLOR_PRIMARY_BG}; padding: 14px;
        border-radius: 10px; border: 1px solid {COLOR_PRIMARY_LIGHT};
    }}
    .reply-header {{
        display: flex; justify-content: space-between; align-items: center;
        margin-bottom: 8px; flex-wrap: wrap; gap: 6px;
    }}
    .model-badge {{
        background: {COLOR_PRIMARY}; color: white; padding: 4px 10px;
        border-radius: 8px; font-size: 13px; font-weight: 600;
    }}
    .score-total {{ font-size: 12px; color: {COLOR_TEXT_SUB}; }}
    .score-total b {{ color: {COLOR_PRIMARY_DARK}; font-size: 14px; }}
    .score-detail {{ margin-bottom: 10px; }}
    .score-pill {{
        display: inline-block; background: white; border: 1px solid {COLOR_BORDER};
        padding: 2px 8px; border-radius: 10px; font-size: 11px; margin: 2px 2px 0 0;
    }}
    .score-pill b {{ color: {COLOR_PRIMARY_DARK}; }}
    .judge-summary {{
        margin-top: 10px; padding: 8px 10px; background: white;
        border: 1px dashed {COLOR_PRIMARY_LIGHT}; border-radius: 6px;
        font-size: 12px; color: {COLOR_TEXT_SUB};
    }}

    /* フィードバック入力欄 */
    .feedback-block {{
        margin-top: 16px; padding: 14px;
        background: #fefce8;  /* ここだけ淡いクリーム色で目立たせる（注意喚起ではなく「入力エリア」として） */
        border: 2px dashed {COLOR_PRIMARY_LIGHT};
        border-radius: 10px;
    }}
    .feedback-label {{
        display: block; font-size: 13px; font-weight: 600;
        color: {COLOR_PRIMARY_DARK}; margin-bottom: 6px;
    }}
    .feedback-label small {{
        font-weight: 400; color: {COLOR_TEXT_SUB}; font-size: 11px;
    }}
    .feedback-input {{
        width: 100%; box-sizing: border-box; padding: 10px;
        border: 1px solid {COLOR_BORDER}; border-radius: 6px;
        font-family: inherit; font-size: 13px;
        resize: vertical;
    }}
    .feedback-input:focus {{
        outline: none; border-color: {COLOR_PRIMARY};
        box-shadow: 0 0 0 3px rgba(124,58,237,0.15);
    }}

    /* フローティングボタン */
    .floating-download {{
        position: fixed; bottom: 24px; right: 24px;
        background: {COLOR_PRIMARY}; color: white;
        padding: 14px 22px; border-radius: 30px; border: none;
        font-size: 14px; font-weight: 600; cursor: pointer;
        box-shadow: 0 6px 20px rgba(124,58,237,0.4);
        z-index: 1000;
    }}
    .floating-download:hover {{ background: {COLOR_PRIMARY_DARK}; }}
    .floating-status {{
        position: fixed; bottom: 80px; right: 24px;
        background: white; border: 1px solid {COLOR_PRIMARY_LIGHT};
        padding: 6px 12px; border-radius: 20px; font-size: 12px;
        color: {COLOR_PRIMARY_DARK}; z-index: 1000;
    }}

    footer {{
        text-align: center; margin-top: 48px; padding-bottom: 80px;
        color: {COLOR_TEXT_SUB}; font-size: 12px;
    }}
</style>
</head>
<body>
<div class="container">
    <header>
        <h1>BayChat AI Reply モデル比較レポート{tone_suffix_title}</h1>
        <div class="subtitle">
            {('<span style="background: ' + tone_color + '; color: white; padding: 4px 12px; border-radius: 6px; font-weight: 700; font-size: 16px; margin-right: 12px;">トーン: ' + tone_en + ' / ' + tone_ja + '</span>') if tone_info else ''}
            {len(models)}モデル × {len(selected_cases)}ケース | 生成日時: {timestamp}<br>
            社長のフィードバックは各ケース下の入力欄に書いてください。自動保存されます。
        </div>
    </header>

    <section class="summary-section">
        <h2>📊 モデル別集計</h2>
        {agg_html}
        {cat_html}
        {speed_html}
    </section>

    <h2>📌 全{len(selected_cases)}ケース（返信内容 + フィードバック入力欄）</h2>
    <p style="color:{COLOR_TEXT_SUB}">
    各ケースの一番下に「社長の理想フィードバック」入力欄があります。書き込むと即座にブラウザに自動保存。
    最下部の <strong>紫ボタン</strong> から全フィードバックをJSONでダウンロードできます。
    </p>
    {case_cards_html}

    <footer>
        BayChat AI Reply テストレポート / ソース: {esc(os.path.basename(excel_path))}<br>
        フィードバックは localStorage（このブラウザ）に自動保存されます。
    </footer>
</div>

<div class="floating-status" id="saveStatus">💾 変更なし</div>
<button class="floating-download" onclick="downloadFeedback()">
    📥 全フィードバックをJSONダウンロード
</button>

<script>
// localStorage キー（このレポート固有）
const STORAGE_KEY = 'baychat_feedback_{timestamp.replace(" ", "_").replace(":", "").replace("-", "")}';
const statusEl = document.getElementById('saveStatus');

// 初期化：保存済みフィードバックを復元
function restoreFeedback() {{
    const saved = localStorage.getItem(STORAGE_KEY);
    if (!saved) return;
    try {{
        const data = JSON.parse(saved);
        document.querySelectorAll('.feedback-input').forEach(ta => {{
            const cid = ta.dataset.caseId;
            if (data[cid]) ta.value = data[cid];
        }});
        const count = Object.keys(data).filter(k => data[k]).length;
        statusEl.textContent = `💾 保存済み ${{count}}件`;
    }} catch (e) {{ console.error(e); }}
}}

// 自動保存
function saveFeedback() {{
    const data = {{}};
    document.querySelectorAll('.feedback-input').forEach(ta => {{
        const cid = ta.dataset.caseId;
        if (ta.value.trim()) data[cid] = ta.value;
    }});
    localStorage.setItem(STORAGE_KEY, JSON.stringify(data));
    const count = Object.keys(data).length;
    statusEl.textContent = `💾 保存 ${{count}}件 (${{new Date().toLocaleTimeString('ja-JP')}})`;
}}

document.querySelectorAll('.feedback-input').forEach(ta => {{
    ta.addEventListener('input', saveFeedback);
}});

// ダウンロード
function downloadFeedback() {{
    const data = {{
        exported_at: new Date().toISOString(),
        report_source: '{esc(os.path.basename(excel_path))}',
        feedback: {{}}
    }};
    document.querySelectorAll('.feedback-input').forEach(ta => {{
        const cid = ta.dataset.caseId;
        if (ta.value.trim()) data.feedback[cid] = ta.value;
    }});
    const blob = new Blob([JSON.stringify(data, null, 2)], {{type: 'application/json'}});
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    const ts = new Date().toISOString().replace(/[:.]/g, '-').slice(0, 19);
    a.download = `baychat_feedback_${{ts}}.json`;
    a.click();
    URL.revokeObjectURL(url);
}}

restoreFeedback();

// ==========================================================
// 補足情報 + 再生成API
// ==========================================================
const REGENERATE_API = 'http://127.0.0.1:8765/api/regenerate';
const TONE = '{tone or "polite"}';
const PROMPT_VERSION = '2.3_baseline_natural2';
const TEST_CASE_PATH = '{(test_cases_path.replace(os.sep, "/").split("testing/")[-1]) if test_cases_path else "test_cases/category_02_natural5_subset.json"}';

// 速度目標（秒）
const SPEED_TARGET_STANDARD = 3.0;  // 標準
const SPEED_TARGET_COMPLEX = 6.0;   // 複雑
const SPEED_TARGET_APOLOGY = 8.0;   // 謝罪

function colorizeElapsed(span, elapsed) {{
    span.classList.remove('fast', 'slow');
    const target = (TONE === 'apologetic') ? SPEED_TARGET_APOLOGY : SPEED_TARGET_STANDARD;
    if (elapsed <= target) span.classList.add('fast');
    else if (elapsed > SPEED_TARGET_COMPLEX) span.classList.add('slow');
}}

async function regenerateCase(caseId, btn) {{
    const block = document.querySelector(`.supplemental-block[data-case-id="${{caseId}}"]`);
    if (!block) return;
    const ta = block.querySelector('.supplemental-input');
    const statusEl = block.querySelector('.regen-status');
    const supplemental = (ta.value || '').trim();

    // 同じケースの全 reply-card を取得（同 case-card 内）
    const caseCard = block.closest('.case-card');
    const replyCards = caseCard.querySelectorAll('.reply-card');

    // ローディング表示
    btn.disabled = true;
    btn.textContent = '⏳ 生成中...';
    statusEl.classList.remove('error','success');
    statusEl.textContent = `${{replyCards.length}}モデルで再生成中...`;
    replyCards.forEach(c => c.classList.add('regenerating'));

    try {{
        const resp = await fetch(REGENERATE_API, {{
            method: 'POST',
            headers: {{ 'Content-Type': 'application/json' }},
            body: JSON.stringify({{
                case_id: caseId,
                supplemental: supplemental,
                tone: TONE,
                prompt_version: PROMPT_VERSION,
                test_case_path: TEST_CASE_PATH,
            }}),
        }});
        if (!resp.ok) throw new Error(`HTTP ${{resp.status}}: ${{await resp.text()}}`);
        const data = await resp.json();

        // 各モデルの結果を画面更新
        let updated = 0;
        let errors = [];
        data.results.forEach(r => {{
            if (r.error) {{ errors.push(`${{r.model_name||r.model_id}}: ${{r.error}}`); return; }}
            const card = caseCard.querySelector(`.reply-card[data-model-name="${{r.model_name}}"]`);
            if (!card) return;
            card.querySelector('.reply-en').textContent = r.buyer_reply;
            card.querySelector('.reply-ja').textContent = r.jpn_reply;
            const elapsedEl = card.querySelector('.meta-elapsed');
            elapsedEl.innerHTML = `⏱ <b>${{r.elapsed.toFixed(1)}}</b>秒`;
            colorizeElapsed(elapsedEl, r.elapsed);
            const costEl = card.querySelector('.meta-cost');
            costEl.innerHTML = `💴 <b>¥${{r.cost_jpy.toFixed(3)}}</b>`;
            card.classList.remove('regenerating');
            card.classList.add('regenerated');
            setTimeout(() => card.classList.remove('regenerated'), 1500);
            updated++;
        }});

        replyCards.forEach(c => c.classList.remove('regenerating'));
        if (errors.length > 0) {{
            statusEl.classList.add('error');
            statusEl.textContent = `⚠️ ${{updated}}件成功 / ${{errors.length}}件エラー: ${{errors.join('; ')}}`;
        }} else {{
            statusEl.classList.add('success');
            statusEl.textContent = `✅ ${{updated}}モデル再生成完了`;
        }}
    }} catch (e) {{
        replyCards.forEach(c => c.classList.remove('regenerating'));
        statusEl.classList.add('error');
        statusEl.textContent = `❌ 失敗: ${{e.message}}（result_server.py が起動しているか確認してください）`;
        console.error(e);
    }} finally {{
        btn.disabled = false;
        btn.textContent = '🔁 補足込みで再生成（3モデル）';
    }}
}}

// 補足情報の自動保存（localStorageに別キーで）
const SUPP_KEY = 'baychat_supplemental_{timestamp.replace(" ", "_").replace(":", "").replace("-", "")}';
function restoreSupplemental() {{
    const saved = localStorage.getItem(SUPP_KEY);
    if (!saved) return;
    try {{
        const data = JSON.parse(saved);
        document.querySelectorAll('.supplemental-input').forEach(ta => {{
            const cid = ta.dataset.caseId;
            if (data[cid]) ta.value = data[cid];
        }});
    }} catch (e) {{ console.error(e); }}
}}
function saveSupplemental() {{
    const data = {{}};
    document.querySelectorAll('.supplemental-input').forEach(ta => {{
        const cid = ta.dataset.caseId;
        if (ta.value.trim()) data[cid] = ta.value;
    }});
    localStorage.setItem(SUPP_KEY, JSON.stringify(data));
}}
document.querySelectorAll('.supplemental-input').forEach(ta => {{
    ta.addEventListener('input', saveSupplemental);
}});
restoreSupplemental();

// 初期表示の応答時間を色付け
document.querySelectorAll('.meta-elapsed').forEach(el => {{
    const m = el.textContent.match(/([\d.]+)秒/);
    if (m) colorizeElapsed(el, parseFloat(m[1]));
}});
</script>
</body>
</html>
"""

    timestamp_fn = datetime.now().strftime("%Y%m%d_%H%M%S")
    tone_fn_suffix = f"_{tone}" if tone else ""
    outpath = os.path.join(RESULTS_DIR, f"comparison{tone_fn_suffix}_{timestamp_fn}.html")
    with open(outpath, "w", encoding="utf-8") as f:
        f.write(html)

    return outpath


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", type=str, required=True)
    parser.add_argument("--cases", type=str, required=True)
    parser.add_argument("--top", type=int, default=12)
    parser.add_argument("--tone", type=str, default=None,
                        choices=["polite", "friendly", "apologetic"],
                        help="トーン（HTML title / header / ファイル名にラベル表示）")
    args = parser.parse_args()

    html_path = render_comparison_html(args.excel, args.cases, args.top, tone=args.tone)
    print(f"HTML生成完了: {html_path}")
    os.startfile(html_path)
