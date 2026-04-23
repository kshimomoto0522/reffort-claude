"""
AI Reply v2.5 モデル比較ダッシュボード
============================================================================
社長向け。1画面で全体像が掴めるシンプルなHTMLダッシュボード。

BayChatブランドカラー（紫ベース）・信号機色禁止・英日併記・情報階層化。

使い方:
    python render_v25_dashboard.py --excel results/test_xxx.xlsx --cases test_cases/xxx.json
"""

import os
import sys
import io
import json
import argparse
from datetime import datetime
from collections import defaultdict
import html as html_lib

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(SCRIPT_DIR, "results")

# BayChat ブランドカラー（紫ベース）
C_PRIMARY = "#5a189a"
C_SECONDARY = "#9d4edd"
C_ACCENT = "#c77dff"
C_LIGHT = "#f3e8ff"
C_LIGHTER = "#faf5ff"
C_TEXT = "#2d2d2d"
C_TEXT_SUB = "#6b7280"
C_BORDER = "#e0d5ef"


def esc(text):
    if text is None:
        return ""
    return html_lib.escape(str(text))


def read_excel(path):
    """Excel結果を辞書リストとして読み込む"""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb["スコアサマリー"] if "スコアサマリー" in wb.sheetnames else wb.active

    # ヘッダー
    headers = [cell.value for cell in ws[1]]
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = dict(zip(headers, row))
        results.append(row_dict)

    # 返信内容シート読み込み
    replies = {}
    if "返信内容" in wb.sheetnames:
        ws_r = wb["返信内容"]
        r_headers = [c.value for c in ws_r[1]]
        for row in ws_r.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rd = dict(zip(r_headers, row))
            key = (rd.get("ケースID") or rd.get("case_id"), rd.get("モデル") or rd.get("model_name"))
            replies[key] = rd

    return results, replies


def load_cases(path):
    """テストケースJSONを読み込む"""
    if not path or not os.path.exists(path):
        return {}
    cases = json.load(open(path, encoding="utf-8"))
    return {c["id"]: c for c in cases}


def render_score_bar(score, max_score=25):
    """スコアバー（紫グラデーション・信号機色NG）"""
    pct = round(score / max_score * 100)
    return f'''
    <div class="score-bar">
      <div class="score-bar-fill" style="width:{pct}%">{pct}%</div>
    </div>'''


def render_grade(score, max_score=25):
    """記号評価（◎/○/△/▲）"""
    pct = score / max_score
    if pct >= 0.85:
        return "◎"
    if pct >= 0.70:
        return "○"
    if pct >= 0.55:
        return "△"
    return "▲"


def render_dashboard(excel_path, cases_path=None, output_path=None):
    results, replies = read_excel(excel_path)
    cases = load_cases(cases_path)

    # モデル別集計
    by_model = defaultdict(list)
    for r in results:
        model = r.get("モデル") or r.get("model_name")
        by_model[model].append(r)

    # ケース別集計（スコア降順で並べ替え）
    case_ids = []
    seen = set()
    for r in results:
        cid = r.get("ケースID") or r.get("case_id")
        if cid and cid not in seen:
            case_ids.append(cid)
            seen.add(cid)

    # モデル比較カード用データ
    model_summaries = []
    for model, rows in by_model.items():
        scores = [r.get("合計スコア") or 0 for r in rows if r.get("合計スコア")]
        elapsed = [r.get("応答時間(秒)") or 0 for r in rows if r.get("応答時間(秒)")]
        costs = [r.get("コスト(円)") or 0 for r in rows if r.get("コスト(円)")]
        model_summaries.append({
            "model": model,
            "avg_score": sum(scores) / len(scores) if scores else 0,
            "avg_elapsed": sum(elapsed) / len(elapsed) if elapsed else 0,
            "total_cost": sum(costs),
            "count": len(rows),
        })
    model_summaries.sort(key=lambda x: x["avg_score"], reverse=True)

    # 勝者（最高スコア）
    best_model = model_summaries[0]["model"] if model_summaries else "-"

    # 全体平均
    all_scores = [r.get("合計スコア") or 0 for r in results if r.get("合計スコア")]
    overall_avg = sum(all_scores) / len(all_scores) if all_scores else 0
    total_cost = sum([r.get("コスト(円)") or 0 for r in results])

    # 改善点の集約
    improvements = []
    critical_issues = []
    for r in results:
        imps = r.get("改善提案") or ""
        crits = r.get("重大な問題") or ""
        if imps:
            improvements.append(imps)
        if crits:
            critical_issues.append(crits)

    # メタ情報
    meta_models = " / ".join(by_model.keys())
    meta_prompt = results[0].get("プロンプト") or results[0].get("prompt_version") or "-" if results else "-"
    case_count = len(case_ids)
    exec_date = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ============================================================
    # HTML生成
    # ============================================================
    html_parts = []

    # CSS
    html_parts.append(f'''<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<title>AI Reply v{meta_prompt} 比較ダッシュボード</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{
    font-family: 'Yu Gothic UI', 'Segoe UI', -apple-system, sans-serif;
    background: #fafafa;
    color: {C_TEXT};
    margin: 0;
    line-height: 1.7;
  }}
  .hero {{
    background: linear-gradient(135deg, {C_PRIMARY} 0%, {C_SECONDARY} 100%);
    color: white;
    padding: 32px 40px;
    text-align: center;
  }}
  .hero h1 {{ margin: 0 0 8px; font-size: 24px; font-weight: 600; }}
  .hero .meta {{ font-size: 13px; opacity: 0.9; }}
  .hero .meta span {{ margin: 0 8px; }}

  .container {{ max-width: 1100px; margin: 0 auto; padding: 28px 24px 60px; }}

  h2 {{
    color: {C_PRIMARY};
    border-left: 5px solid {C_SECONDARY};
    padding-left: 14px;
    font-size: 20px;
    margin: 36px 0 16px;
  }}
  h2:first-of-type {{ margin-top: 0; }}

  /* 総合ボックス */
  .overall {{
    background: white;
    border: 2px solid {C_ACCENT};
    border-radius: 12px;
    padding: 24px 28px;
    margin-bottom: 32px;
  }}
  .overall-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 20px;
    margin-top: 12px;
  }}
  .overall-item {{ text-align: center; }}
  .overall-item .label {{ font-size: 12px; color: {C_TEXT_SUB}; margin-bottom: 4px; }}
  .overall-item .value {{ font-size: 28px; font-weight: 700; color: {C_PRIMARY}; }}
  .overall-item .sub {{ font-size: 12px; color: {C_TEXT_SUB}; margin-top: 2px; }}

  /* モデル比較カード */
  .cards {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
    gap: 20px;
    margin-top: 12px;
  }}
  .card {{
    background: white;
    border: 2px solid {C_BORDER};
    border-radius: 12px;
    padding: 20px 24px;
    transition: border-color 0.2s;
  }}
  .card.best {{ border-color: {C_SECONDARY}; box-shadow: 0 4px 12px rgba(157, 78, 221, 0.15); }}
  .card h3 {{
    margin: 0 0 4px;
    color: {C_PRIMARY};
    font-size: 16px;
    font-weight: 600;
  }}
  .card .badge-best {{
    display: inline-block;
    background: {C_SECONDARY};
    color: white;
    font-size: 11px;
    padding: 2px 8px;
    border-radius: 12px;
    margin-left: 6px;
    font-weight: 500;
  }}
  .card .score-main {{
    font-size: 32px;
    font-weight: 700;
    color: {C_PRIMARY};
    text-align: center;
    margin: 12px 0 4px;
  }}
  .card .score-main .grade {{ font-size: 20px; margin-left: 6px; }}
  .card .score-sub {{
    font-size: 12px;
    color: {C_TEXT_SUB};
    text-align: center;
    margin-bottom: 12px;
  }}
  .score-bar {{
    background: {C_LIGHT};
    border-radius: 8px;
    height: 20px;
    overflow: hidden;
    margin: 10px 0;
  }}
  .score-bar-fill {{
    height: 100%;
    background: linear-gradient(90deg, {C_PRIMARY} 0%, {C_ACCENT} 100%);
    display: flex;
    align-items: center;
    justify-content: center;
    color: white;
    font-size: 11px;
    font-weight: 600;
    transition: width 0.4s ease;
  }}
  .card-stats {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 8px;
    margin-top: 14px;
    padding-top: 14px;
    border-top: 1px dashed {C_BORDER};
  }}
  .card-stat {{ text-align: center; font-size: 12px; }}
  .card-stat .label {{ color: {C_TEXT_SUB}; }}
  .card-stat .value {{ color: {C_PRIMARY}; font-weight: 600; font-size: 14px; margin-top: 2px; }}

  /* ケース別アコーディオン */
  .accordion {{ margin-top: 12px; }}
  .case-item {{
    background: white;
    border: 1px solid {C_BORDER};
    border-radius: 10px;
    margin-bottom: 12px;
    overflow: hidden;
  }}
  .case-header {{
    padding: 14px 20px;
    cursor: pointer;
    display: flex;
    justify-content: space-between;
    align-items: center;
    background: white;
    user-select: none;
  }}
  .case-header:hover {{ background: {C_LIGHTER}; }}
  .case-header .title {{ font-weight: 600; color: {C_PRIMARY}; }}
  .case-header .category {{
    display: inline-block;
    background: {C_LIGHT};
    color: {C_PRIMARY};
    font-size: 11px;
    padding: 2px 10px;
    border-radius: 10px;
    margin-left: 8px;
    font-weight: 500;
  }}
  .case-header .scores-mini {{ font-size: 12px; color: {C_TEXT_SUB}; }}
  .case-header .caret {{ transition: transform 0.2s; color: {C_ACCENT}; font-size: 14px; margin-left: 10px; }}
  .case-item.open .caret {{ transform: rotate(180deg); }}
  .case-body {{
    padding: 0 20px;
    max-height: 0;
    overflow: hidden;
    transition: max-height 0.3s ease, padding 0.3s ease;
  }}
  .case-item.open .case-body {{
    max-height: 10000px;
    padding: 20px;
    border-top: 1px dashed {C_BORDER};
  }}

  .buyer-msg {{
    background: {C_LIGHTER};
    border-left: 4px solid {C_ACCENT};
    padding: 12px 16px;
    border-radius: 6px;
    margin-bottom: 16px;
  }}
  .buyer-msg .label {{
    font-size: 11px;
    color: {C_PRIMARY};
    font-weight: 600;
    margin-bottom: 4px;
  }}
  .buyer-msg .text {{ font-size: 13px; white-space: pre-wrap; }}
  .buyer-msg .ja {{
    font-size: 12px;
    color: {C_TEXT_SUB};
    margin-top: 6px;
    border-top: 1px dashed {C_BORDER};
    padding-top: 6px;
  }}

  .replies-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
    gap: 14px;
  }}
  .reply-card {{
    background: white;
    border: 1.5px solid {C_BORDER};
    border-radius: 10px;
    padding: 14px 16px;
  }}
  .reply-card.winner {{ border-color: {C_SECONDARY}; background: {C_LIGHTER}; }}
  .reply-card h4 {{
    margin: 0 0 4px;
    font-size: 13px;
    color: {C_PRIMARY};
    display: flex;
    justify-content: space-between;
    align-items: center;
  }}
  .reply-card .score {{
    font-size: 16px;
    font-weight: 700;
    color: {C_PRIMARY};
  }}
  .reply-card .reply-text {{
    font-size: 12px;
    white-space: pre-wrap;
    line-height: 1.6;
    color: {C_TEXT};
    margin: 8px 0;
    max-height: 180px;
    overflow-y: auto;
    padding: 8px;
    background: #fafafa;
    border-radius: 4px;
  }}
  .reply-card .reply-ja {{
    font-size: 11px;
    color: {C_TEXT_SUB};
    white-space: pre-wrap;
    margin-top: 6px;
    padding: 8px;
    background: {C_LIGHT};
    border-radius: 4px;
  }}
  .reply-card .judge {{
    font-size: 11px;
    color: {C_TEXT_SUB};
    margin-top: 6px;
    padding: 6px 8px;
    background: {C_LIGHTER};
    border-left: 3px solid {C_SECONDARY};
    border-radius: 4px;
  }}

  .proposals {{
    background: {C_LIGHTER};
    border-left: 5px solid {C_SECONDARY};
    padding: 20px 24px;
    border-radius: 8px;
    margin-top: 12px;
  }}
  .proposals h3 {{
    margin: 0 0 10px;
    color: {C_PRIMARY};
    font-size: 15px;
  }}
  .proposals ul {{ margin: 4px 0; padding-left: 20px; font-size: 13px; }}
  .proposals li {{ margin-bottom: 6px; }}

  .footer {{
    text-align: center;
    color: {C_TEXT_SUB};
    font-size: 11px;
    margin-top: 40px;
    padding-top: 20px;
    border-top: 1px dashed {C_BORDER};
  }}
</style>
</head>
<body>

<div class="hero">
  <h1>🧪 AI Reply v{esc(meta_prompt)} モデル比較</h1>
  <div class="meta">
    <span>📅 {exec_date}</span>
    <span>|</span>
    <span>🎯 {case_count}ケース × {len(by_model)}モデル</span>
    <span>|</span>
    <span>🧩 {esc(meta_models)}</span>
  </div>
</div>

<div class="container">
''')

    # 総合サマリー
    html_parts.append(f'''
  <h2>📊 総合サマリー</h2>
  <div class="overall">
    <div class="overall-grid">
      <div class="overall-item">
        <div class="label">全体平均スコア</div>
        <div class="value">{overall_avg:.1f}<span style="font-size:16px; color:{C_TEXT_SUB};"> / 25</span></div>
        <div class="sub">{render_grade(overall_avg)} {int(overall_avg/25*100)}%</div>
      </div>
      <div class="overall-item">
        <div class="label">総合最高モデル</div>
        <div class="value" style="font-size:18px; padding-top:10px;">{esc(best_model)}</div>
        <div class="sub">平均 {model_summaries[0]['avg_score']:.1f} / 25</div>
      </div>
      <div class="overall-item">
        <div class="label">テスト総件数</div>
        <div class="value">{len(results)}</div>
        <div class="sub">ケース × モデル</div>
      </div>
      <div class="overall-item">
        <div class="label">合計コスト</div>
        <div class="value">¥{total_cost:.2f}</div>
        <div class="sub">今回の実行分</div>
      </div>
    </div>
  </div>
''')

    # モデル比較カード
    html_parts.append('<h2>🏆 モデル比較</h2>\n  <div class="cards">\n')
    for ms in model_summaries:
        is_best = ms["model"] == best_model
        best_badge = f'<span class="badge-best">最高</span>' if is_best else ''
        html_parts.append(f'''
    <div class="card {'best' if is_best else ''}">
      <h3>{esc(ms['model'])}{best_badge}</h3>
      <div class="score-main">{ms['avg_score']:.1f}<span style="font-size:14px; color:{C_TEXT_SUB};"> / 25</span> <span class="grade">{render_grade(ms['avg_score'])}</span></div>
      <div class="score-sub">{int(ms['avg_score']/25*100)}% / {ms['count']}件</div>
      {render_score_bar(ms['avg_score'])}
      <div class="card-stats">
        <div class="card-stat">
          <div class="label">平均応答</div>
          <div class="value">{ms['avg_elapsed']:.1f}秒</div>
        </div>
        <div class="card-stat">
          <div class="label">総コスト</div>
          <div class="value">¥{ms['total_cost']:.2f}</div>
        </div>
      </div>
    </div>
''')
    html_parts.append('  </div>\n')

    # ケース別詳細（アコーディオン）
    html_parts.append('<h2>📋 ケース別詳細</h2>\n  <div class="accordion">\n')
    for idx, case_id in enumerate(case_ids):
        case_meta = cases.get(case_id, {})
        category = case_meta.get("category", "-")
        buyer_msg = case_meta.get("buyer_message", "")
        buyer_msg_ja = case_meta.get("buyer_message_ja", "")

        # このケースの結果を取得
        case_results = [r for r in results if (r.get("ケースID") or r.get("case_id")) == case_id]
        case_scores = [r.get("合計スコア") or 0 for r in case_results]
        avg = sum(case_scores) / len(case_scores) if case_scores else 0
        best_r = max(case_results, key=lambda r: r.get("合計スコア") or 0) if case_results else None
        best_model_for_case = best_r.get("モデル") if best_r else "-"

        # スコアミニ表示
        scores_mini = " / ".join([f"{(r.get('モデル') or '')[:10]}:{r.get('合計スコア')}" for r in case_results])

        html_parts.append(f'''
    <div class="case-item" id="case-{idx}">
      <div class="case-header" onclick="toggleCase({idx})">
        <div>
          <span class="title">ケース{idx+1}</span>
          <span class="category">{esc(category)}</span>
        </div>
        <div class="scores-mini">平均 {avg:.1f}／勝者: {esc(best_model_for_case)} <span class="caret">▼</span></div>
      </div>
      <div class="case-body">
        <div class="buyer-msg">
          <div class="label">📨 バイヤーメッセージ</div>
          <div class="text">{esc(buyer_msg)}</div>
          {f'<div class="ja">🇯🇵 {esc(buyer_msg_ja)}</div>' if buyer_msg_ja else ''}
        </div>
        <div class="replies-grid">
''')
        for r in case_results:
            model_name = r.get("モデル") or r.get("model_name") or ""
            is_winner = (r is best_r)
            score = r.get("合計スコア") or 0
            reply_text = r.get("返信(英語)") or r.get("buyer_reply") or ""
            reply_ja = r.get("返信(日本語)") or r.get("jpn_reply") or ""
            judge_comment = r.get("審判コメント") or r.get("summary_ja") or ""
            elapsed = r.get("応答時間(秒)")
            cost = r.get("コスト(円)")

            html_parts.append(f'''
          <div class="reply-card {'winner' if is_winner else ''}">
            <h4>
              <span>{esc(model_name)}{' 🏆' if is_winner else ''}</span>
              <span class="score">{score}<span style="font-size:11px;color:{C_TEXT_SUB};"> / 25</span></span>
            </h4>
            <div style="font-size:11px; color:{C_TEXT_SUB};">
              {f'{elapsed:.1f}秒' if elapsed else ''} {f'| ¥{cost:.3f}' if cost else ''}
            </div>
            <div class="reply-text">{esc(reply_text)}</div>
            {f'<div class="reply-ja">🇯🇵 {esc(reply_ja)}</div>' if reply_ja else ''}
            {f'<div class="judge">💬 {esc(judge_comment)}</div>' if judge_comment else ''}
          </div>
''')

        html_parts.append('        </div>\n      </div>\n    </div>\n')

    html_parts.append('  </div>\n')

    # 改善提案サマリ
    if improvements or critical_issues:
        html_parts.append(f'''
  <h2>💡 審判による改善提案サマリ</h2>
  <div class="proposals">
''')
        if critical_issues:
            unique_crits = list(dict.fromkeys(critical_issues))[:5]
            html_parts.append('    <h3>🔴 重大な問題</h3>\n    <ul>\n')
            for c in unique_crits:
                html_parts.append(f'      <li>{esc(c)}</li>\n')
            html_parts.append('    </ul>\n')
        if improvements:
            unique_imps = list(dict.fromkeys(improvements))[:5]
            html_parts.append('    <h3 style="margin-top:16px;">✨ 改善提案</h3>\n    <ul>\n')
            for i in unique_imps:
                html_parts.append(f'      <li>{esc(i)}</li>\n')
            html_parts.append('    </ul>\n')
        html_parts.append('  </div>\n')

    # Footer
    html_parts.append(f'''
  <div class="footer">
    BayChat AI Reply v{esc(meta_prompt)} テストダッシュボード | 生成: {exec_date}
  </div>
</div>

<script>
  function toggleCase(idx) {{
    const item = document.getElementById('case-' + idx);
    item.classList.toggle('open');
  }}

  // 最初のケースをデフォルトで開く
  document.addEventListener('DOMContentLoaded', () => {{
    const first = document.getElementById('case-0');
    if (first) first.classList.add('open');
  }});
</script>

</body>
</html>
''')

    html_content = "".join(html_parts)

    # 出力ファイルパス
    if not output_path:
        base = os.path.splitext(os.path.basename(excel_path))[0]
        output_path = os.path.join(RESULTS_DIR, f"dashboard_{base}.html")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"ダッシュボード生成完了: {output_path}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True, help="Excel結果ファイル")
    parser.add_argument("--cases", help="テストケースJSON（オプション・バイヤーメッセージ表示用）")
    parser.add_argument("--output", help="出力HTMLパス（省略時は自動）")
    args = parser.parse_args()

    excel = args.excel
    if not os.path.isabs(excel):
        excel = os.path.join(SCRIPT_DIR, excel)
    cases = args.cases
    if cases and not os.path.isabs(cases):
        cases = os.path.join(SCRIPT_DIR, cases)

    render_dashboard(excel, cases, args.output)
